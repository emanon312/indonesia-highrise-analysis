# -*- coding: utf-8 -*-
"""扫描另一个 agent 产出的 13 个 xlsx，输出结构化摘要(JSON, UTF-8)供对比。
只读，不修改任何文件。"""
import json, glob, os
from openpyxl import load_workbook

SRC = r"analysis/data/other_agent_lists"  # 对方 agent 提供的各城高层清单 xlsx，未入库，需自备

def summarize(path):
    info = {"文件": os.path.basename(path)}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        info["错误"] = repr(e); return info
    sheets = wb.sheetnames
    info["sheet名"] = sheets
    ws = wb[sheets[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        info["错误"] = "主sheet空"; wb.close(); return info
    header = [str(c) if c is not None else "" for c in rows[0]]
    data = rows[1:]
    info["列数"] = len(header)
    info["列名"] = header
    info["数据行数"] = len(data)
    # 找高度列（列名含 高/height/m）
    hi = next((i for i,h in enumerate(header) if ("高度" in h) or ("height" in h.lower()) or h.strip() in ("高","米","m")), None)
    if hi is not None:
        hs = [r[hi] for r in data if hi < len(r) and isinstance(r[hi], (int,float))]
        if hs:
            info["高度列"] = header[hi]
            info["高度范围"] = [round(min(hs),1), round(max(hs),1)]
            info["高度≥32数"] = sum(1 for v in hs if v>=32)
    # 找名称列
    ni = next((i for i,h in enumerate(header) if ("名称" in h) or ("name" in h.lower()) or ("楼" in h) or ("建筑物" in h)), None)
    if ni is not None:
        info["名称列"] = header[ni]
        info["名称样例"] = [str(r[ni]) for r in data if ni < len(r) and r[ni] and str(r[ni]).strip()][:5]
    # 前2行原始数据样例（便于看格式）
    info["首2行"] = [[ (str(v)[:40] if v is not None else None) for v in r[:len(header)]] for r in data[:2]]
    # 其余sheet行数
    info["其余sheet行"] = {s: max(0, sum(1 for _ in wb[s].iter_rows(values_only=True))-1) for s in sheets[1:]}
    wb.close()
    return info

results = []
for p in sorted(glob.glob(os.path.join(SRC, "*高层建筑.xlsx"))):
    results.append(summarize(p))

out = r"analysis\data\inspect_other.json"
with open(out, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=1)
print("WROTE", out, "files=", len(results))
