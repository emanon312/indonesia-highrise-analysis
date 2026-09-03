# -*- coding: utf-8 -*-
"""摸清另一agent各文件每个sheet的逐栋字段结构(只读)。
对每个文件的每个sheet：输出 sheet名/列数/表头/1行样例/数据行数。"""
import json, glob, os
from openpyxl import load_workbook

SRC = r"analysis/data/other_agent_lists"  # 对方 agent 提供的各城高层清单 xlsx，未入库，需自备

def dump(path):
    info = {"文件": os.path.basename(path), "sheets": []}
    wb = load_workbook(path, read_only=True, data_only=True)
    for s in wb.sheetnames:
        ws = wb[s]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i >= 3:  # 只取前4行够看表头+样例
                break
        n = sum(1 for _ in ws.iter_rows(values_only=True))
        # 找出看起来是表头的那一行（含 Lat/Lon/Height/高度/经/纬 关键字）
        def cell(r): return [ (str(v)[:22] if v is not None else None) for v in r ]
        info["sheets"].append({
            "名": s,
            "总行": n,
            "前4行": [cell(r) for r in rows],
        })
    wb.close()
    return info

results = [dump(p) for p in sorted(glob.glob(os.path.join(SRC, "*高层建筑.xlsx")))]
out = r"analysis\data\inspect_other_tiers.json"
with open(out, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=1)
print("WROTE", out)
