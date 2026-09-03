# -*- coding: utf-8 -*-
"""把另一agent的"漏网楼"(我们没有的)补进10城xlsx的「高层清单」sheet。
口径统一到我们4米/层、阈值≥32m；空间匹配>30m视为漏网；补进的行明确标注第二来源。
不动雅加达、棉兰。幂等：已合并过的城自动跳过。"""
import json, os, math
from copy import copy
from openpyxl import load_workbook

SRC = r"analysis/data/other_agent_lists"  # 对方 agent 提供的各城高层清单 xlsx，未入库，需自备
OUT = r"analysis\output"
DIST = 30.0       # 匹配阈值(米)
SELF_DEDUP = 15.0 # 补进的楼彼此去重(米)

CITY = {"Bandung":"万隆","Bekasi":"勿加泗","Tangerang":"唐格朗","Bogor":"茂物",
        "Semarang":"三宝垄","Depok":"德波","Surabaya":"泗水","Makassar":"望加锡",
        "Batam":"巴淡","Palembang":"巨港"}

def find_header_row(ws, maxscan=6):
    rows = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows[:maxscan]):
        cells = [str(c).lower() if c is not None else "" for c in r]
        if any(("纬" in c) or ("lat" in c) for c in cells) and \
           any(("高度" in c) or ("height" in c) for c in cells):
            return i, rows
    return 0, rows

def col_idx(header, kind):
    for i, h in enumerate(header):
        hl = str(h).lower() if h is not None else ""
        if "中心" in hl: continue
        if kind=="lat" and (("纬" in hl) or ("lat" in hl)): return i
        if kind=="lon" and (("经" in hl) or ("lon" in hl)): return i
        if kind=="h" and (("高度" in hl) or ("height" in hl)): return i
        if kind=="conf" and (("置信" in hl) or ("conf" in hl)): return i
        if kind=="area" and (("面积" in hl) or ("area" in hl)): return i
    return None

def their_buildings(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "28" in s), None)
    ws = wb[sheet]; hidx, rows = find_header_row(ws)
    header = [str(c) if c is not None else "" for c in rows[hidx]]
    ci = {k: col_idx(header, k) for k in ("lat","lon","h","conf","area")}
    out = []
    for r in rows[hidx+1:]:
        try:
            lat=float(r[ci["lat"]]); lon=float(r[ci["lon"]]); h=float(r[ci["h"]])
        except (TypeError, ValueError, IndexError): continue
        conf = None; area = None
        if ci["conf"] is not None and ci["conf"]<len(r):
            try: conf=float(r[ci["conf"]])
            except (TypeError,ValueError): pass
        if ci["area"] is not None and ci["area"]<len(r):
            try: area=float(r[ci["area"]])
            except (TypeError,ValueError): pass
        out.append({"lat":lat,"lon":lon,"h":h,"conf":conf,"area":area})
    wb.close()
    return out

def dist_m(lat1,lon1,lat2,lon2):
    dy=(lat2-lat1)*111320.0
    dx=(lon2-lon1)*111320.0*math.cos(math.radians(lat1))
    return math.sqrt(dx*dx+dy*dy)

def tier_label(h):
    if h>=80: return "≥80米(约20层)"
    if h>=64: return "≥64米(约16层)"
    if h>=48: return "≥48米(约12层)"
    if h>=40: return "≥40米(约10层)"
    return "≥32米(约8层)"

def hcol(header, name):
    if name=="建筑面积":
        return next(i for i,h in enumerate(header) if h.startswith("建筑面积"))
    if name=="高度(米)":
        return next(i for i,h in enumerate(header) if h=="高度(米)")
    if name=="层数":
        return next(i for i,h in enumerate(header) if h=="层数")
    return next(i for i,h in enumerate(header) if h==name)

log=[]
for en, cn in CITY.items():
    ourp = os.path.join(OUT, f"{cn}高层建筑清单.xlsx")
    wb = load_workbook(ourp)            # 可写，保留样式
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]]
    idx = {n: hcol(header, n) for n in
           ["序号","楼宇名称","名称来源","纬度","经度","高度(米)","高度来源","层数","层数来源","高度分档","用途分类","建筑面积","备注"]}
    data = rows[1:]
    # 幂等保护：已合并过则跳过
    src_i = idx["高度来源"]
    if any("第二来源" in str(r[src_i]) for r in data if r[src_i] is not None):
        log.append({"城市":cn,"状态":"已合并过-跳过"}); wb.close(); continue
    # 我们现有坐标
    li, oi = idx["纬度"], idx["经度"]
    ours = [(float(r[li]),float(r[oi])) for r in data
            if isinstance(r[li],(int,float)) and isinstance(r[oi],(int,float))]
    # 无名行的"名称来源"约定(取众数)
    ni, nsi = idx["楼宇名称"], idx["名称来源"]
    nameless = [str(r[nsi]) for r in data if (r[ni] is None or str(r[ni]).strip()=="")]
    nameless_label = max(set(nameless), key=nameless.count) if nameless else "无"
    # 他们的≥32m楼，逐栋判漏网
    theirs = [b for b in their_buildings(os.path.join(SRC, f"{en}_高层建筑.xlsx")) if b["h"]>=32]
    added=[]
    for b in theirs:
        if any(dist_m(b["lat"],b["lon"],la,lo)<=DIST for (la,lo) in ours):
            continue  # 已有
        if any(dist_m(b["lat"],b["lon"],a["lat"],a["lon"])<=SELF_DEDUP for a in added):
            continue  # 本轮自重复
        added.append(b)
    # 样式模板=第2行(首个数据行)各列
    tmpl = {c: ws.cell(row=2, column=c+1) for c in range(len(header))}
    start_seq = len(data) + 1
    for k, b in enumerate(added):
        rr = ws.max_row + 1
        hr = round(b["h"], 1)   # 先舍入，分档/层数都基于显示值，避免边界不自洽
        vals = {
            "序号": start_seq + k,
            "楼宇名称": None,
            "名称来源": nameless_label,
            "纬度": round(b["lat"],6),
            "经度": round(b["lon"],6),
            "高度(米)": hr,
            "高度来源": "Google估算(第二来源)",
            "层数": int(round(hr/4)),
            "层数来源": "估算",
            "高度分档": tier_label(hr),
            "用途分类": None,
            "建筑面积": round(b["area"],1) if b["area"] is not None else None,
            "备注": f"交叉补充·第二来源ML·置信度{b['conf']:.2f}" if b["conf"] is not None else "交叉补充·第二来源ML",
        }
        for n, ci_ in idx.items():
            cell = ws.cell(row=rr, column=ci_+1, value=vals[n])
            try: cell._style = copy(tmpl[ci_]._style)
            except Exception: pass
    # 说明sheet追加注记
    if "说明" in wb.sheetnames:
        wsd = wb["说明"]
        base = wsd.max_row + 2
        notes = [
            "【交叉补充说明】",
            f"本清单已用第二来源(另一套Google ML估算)交叉补充我方遗漏的高层 {len(added)} 栋。",
            "匹配方法：第二来源≥32米的楼，与本清单做坐标空间匹配，距离>30米视为我方遗漏并补入。",
            "补入行特征：高度来源标注「Google估算(第二来源)」，无楼名，层数按4米/层换算，备注含原始置信度。",
            "局限：第二来源同为ML估算且无楼名，补入楼多因高度估算差异/边界而被我方原口径漏判，仍属下限估计。",
        ]
        for j, t in enumerate(notes):
            wsd.cell(row=base+j, column=1, value=t)
    wb.save(ourp)
    log.append({"城市":cn,"现有":len(data),"他们≥32m":len(theirs),"补入":len(added),"补后":len(data)+len(added)})

o = r"analysis\data\merge_log.json"
json.dump(log, open(o,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
tot = sum(x.get("补入",0) for x in log)
print("EN          our  their32  added  total")
for x in log:
    if "补入" in x:
        print(f"{x['城市']:<6}{x['现有']:>6}{x['他们≥32m']:>8}{x['补入']:>7}{x['补后']:>7}")
    else:
        print(f"{x['城市']}  {x['状态']}")
print("TOTAL ADDED =", tot)
print("WROTE", o)
