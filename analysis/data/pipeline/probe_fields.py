# -*- coding: utf-8 -*-
"""探查现有xlsx的字段约定：表头 + 含名行 + 无名行各一，写UTF-8文件。"""
import json
from openpyxl import load_workbook
p = r"analysis\output\望加锡高层建筑清单.xlsx"
wb = load_workbook(p, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
header = [str(c) if c is not None else "" for c in rows[0]]
data = rows[1:]
# 名称来源列
si = next(i for i,h in enumerate(header) if "名称来源" in h)
named = next((r for r in data if r[si] not in (None,"空","")), None)
unnamed = next((r for r in data if r[si] in ("空",None,"")), None)
out = {
  "header": header,
  "含名行": [str(v) for v in named] if named else None,
  "无名行": [str(v) for v in unnamed] if unnamed else None,
  "分档去重": sorted({str(r[next(i for i,h in enumerate(header) if "分档" in h)]) for r in data}),
  "名称来源去重": sorted({str(r[si]) for r in data}),
  "高度来源去重": sorted({str(r[next(i for i,h in enumerate(header) if "高度来源" in h)]) for r in data}),
  "层数来源去重": sorted({str(r[next(i for i,h in enumerate(header) if "层数来源" in h)]) for r in data}),
  "用途分类去重": sorted({str(r[next(i for i,h in enumerate(header) if "用途" in h)]) for r in data})[:10],
}
o = r"analysis\data\probe_fields.json"
json.dump(out, open(o,"w",encoding="utf-8"), ensure_ascii=False, indent=1)
print("WROTE", o)
