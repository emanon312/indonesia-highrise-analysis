# -*- coding: utf-8 -*-
"""12城高层建筑汇总：按累计阈值(≥32/40/48/64/80米)统计各城栋数。
11个估算城从各自xlsx的「高度分档」列读per-bucket后累加；雅加达用官方真值。
输出UTF-8 markdown文件供阅读(终端GBK会乱码故写文件)。"""
import os, re
from openpyxl import load_workbook

OUT = r"analysis\output"

# 11 估算城：中文名（文件名= {cn}高层建筑清单.xlsx）
EST = ["泗水","万隆","唐格朗","望加锡","棉兰","三宝垄","巴淡","德波","勿加泗","茂物","巨港"]

# 雅加达官方真值（累计），来自交接状态
JKT = {"name":"雅加达","ge32":7220,"ge40":5261,"ge48":4226,"ge64":2943,"ge80":2105,"src":"官方真值"}

def band_to_thr(label):
    """从分档标签提取阈值数字，如「≥80米(约20层)」→80"""
    m = re.search(r"(80|64|48|40|32)", str(label))
    return int(m.group(1)) if m else None

def city_cumulative(cn):
    path = os.path.join(OUT, f"{cn}高层建筑清单.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]]
    # 高度分档列下标
    bi = next((i for i,h in enumerate(header) if "分档" in h), None)
    buckets = {32:0,40:0,48:0,64:0,80:0}
    for r in rows[1:]:
        thr = band_to_thr(r[bi]) if bi is not None and bi < len(r) else None
        if thr in buckets:
            buckets[thr] += 1
    wb.close()
    # 累计：≥t = 所有 >=t 的桶之和
    cum = {}
    for t in (32,40,48,64,80):
        cum[t] = sum(v for k,v in buckets.items() if k >= t)
    return cum

rows_out = []
for cn in EST:
    cum = city_cumulative(cn)
    rows_out.append({"name":cn,"ge32":cum[32],"ge40":cum[40],"ge48":cum[48],"ge64":cum[64],"ge80":cum[80],"src":"Google ML估算(下限)"})

# 雅加达
rows_out.append(JKT)
# 按 ≥32 降序
rows_out.sort(key=lambda x:x["ge32"], reverse=True)

lines = []
lines.append("# 印尼12城高层建筑汇总（按累计阈值统计栋数）\n")
lines.append("> 口径：4米/层。≥32米=8层 / ≥40=10 / ≥48=12 / ≥64=16 / ≥80=20。")
lines.append("> 雅加达为官方真值标杆；其余11城为 Google ML 估算**下限**（越高漏越多，雅加达实测各档捕获率53→23%）。")
lines.append("> 其中10城已用第二来源(另一套Google ML估算)交叉补充我方遗漏高层共+425栋(坐标>30米视为遗漏)；棉兰因第二来源缺坐标未补。\n")
lines.append("| 城市 | ≥32米(8层) | ≥40米(10层) | ≥48米(12层) | ≥64米(16层) | ≥80米(20层) | 数据来源 |")
lines.append("|------|-----------:|-----------:|-----------:|-----------:|-----------:|----------|")
tot = {"ge32":0,"ge40":0,"ge48":0,"ge64":0,"ge80":0}
for r in rows_out:
    lines.append(f"| {r['name']} | {r['ge32']} | {r['ge40']} | {r['ge48']} | {r['ge64']} | {r['ge80']} | {r['src']} |")
    for k in tot: tot[k]+=r[k]
lines.append(f"| **合计** | **{tot['ge32']}** | **{tot['ge40']}** | **{tot['ge48']}** | **{tot['ge64']}** | **{tot['ge80']}** | — |")

md = "\n".join(lines) + "\n"
out_path = os.path.join(OUT, "12城高层建筑汇总.md")
with open(out_path, "w", encoding="utf-8") as f:
    f.write(md)
print("WROTE", out_path)
