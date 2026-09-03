# -*- coding: utf-8 -*-
# 棉兰高层建筑清单验收脚本，结果以 JSON 输出
import json
import openpyxl

PATH = r"analysis\output\棉兰高层建筑清单.xlsx"
W, S, E, N = 98.62, 3.52, 98.74, 3.66

EXPECTED_SHEETS = ["高层清单", "地标对照", "说明"]
EXPECTED_COLS = ["序号", "楼宇名称", "名称来源", "纬度", "经度", "高度(米)",
                 "高度来源", "层数", "层数来源", "高度分档", "用途分类", "建筑面积", "备注"]

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
sheets = wb.sheetnames

res = {"anomalies": []}
anomalies = res["anomalies"]

# 1 sheets_ok
sheets_ok = sheets == EXPECTED_SHEETS
if not sheets_ok:
    anomalies.append("sheet 名称/数量不符，实际为: " + " / ".join(sheets))

main = wb[sheets[0]]
rows = list(main.iter_rows(values_only=True))
header = list(rows[0]) if rows else []
data = rows[1:] if len(rows) > 1 else []

# 2 columns_ok
header_clean = [str(c).strip() if c is not None else "" for c in header]
columns_ok = (len(header_clean) == 13) and (header_clean == EXPECTED_COLS)
if not columns_ok:
    anomalies.append("列名不符。实际列(%d): %s" % (len(header_clean), header_clean))

# 列索引
def idx(name):
    return EXPECTED_COLS.index(name) if name in EXPECTED_COLS else None

# 用实际表头匹配以防顺序异常，但优先用预期顺序
col = {}
for nm in EXPECTED_COLS:
    if nm in header_clean:
        col[nm] = header_clean.index(nm)

# 3 count
count = len(data)

# helper
def get(row, name):
    i = col.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]

def to_float(v):
    try:
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return None
        return float(v)
    except Exception:
        return None

# 4 tier_sum_ok
from collections import Counter
tier_counter = Counter()
for r in data:
    t = get(r, "高度分档")
    tier_counter[str(t).strip() if t is not None else ""] += 1
tier_sum = sum(tier_counter.values())
tier_sum_ok = (tier_sum == count)

# 5 band_consistent  口径 4米/层
# 阈值: >=80 ->"≥80", 64<=h<80 ->"≥64", 48<=h<64 ->"≥48", 40<=h<48 ->"≥40", 32<=h<40 ->"≥32"
def expected_band(h):
    if h is None:
        return None
    if h >= 80:
        return "≥80"
    if h >= 64:
        return "≥64"
    if h >= 48:
        return "≥48"
    if h >= 40:
        return "≥40"
    if h >= 32:
        return "≥32"
    return "<32"

import re
def norm_band(b):
    # 提取数字，统一为 "≥NN"
    if b is None:
        return None
    s = str(b)
    m = re.search(r"(\d+)", s)
    if m:
        return "≥" + m.group(1)
    return s.strip()

band_bad = 0
band_examples = []
for ri, r in enumerate(data):
    h = to_float(get(r, "高度(米)"))
    b = norm_band(get(r, "高度分档"))
    eb = expected_band(h)
    if h is None:
        band_bad += 1
        if len(band_examples) < 5:
            band_examples.append("行%d 高度缺失 分档=%s" % (ri+2, b))
        continue
    if b != eb:
        band_bad += 1
        if len(band_examples) < 5:
            band_examples.append("行%d 高度%.1f 期望%s 实际%s" % (ri+2, h, eb, b))
band_consistent = (band_bad == 0)
if not band_consistent:
    anomalies.append("分档不自洽行数=%d。示例: %s" % (band_bad, "; ".join(band_examples)))

# 6 all_height_ge32
heights = [to_float(get(r, "高度(米)")) for r in data]
below32 = [h for h in heights if h is not None and h < 32]
missing_h = sum(1 for h in heights if h is None)
all_height_ge32 = (len(below32) == 0 and missing_h == 0)
if below32:
    anomalies.append("有 %d 行高度<32米，最小=%.1f" % (len(below32), min(below32)))
if missing_h:
    anomalies.append("有 %d 行高度缺失" % missing_h)

# 高度异常大 >250
big = [h for h in heights if h is not None and h > 250]
if big:
    anomalies.append("有 %d 行高度>250米(二三线城市可疑)，最大=%.1f" % (len(big), max(big)))
maxh = max([h for h in heights if h is not None], default=None)

# 7 coords_in_bbox_pct
in_box = 0
valid_coord = 0
for r in data:
    lat = to_float(get(r, "纬度"))
    lon = to_float(get(r, "经度"))
    if lat is None or lon is None:
        continue
    valid_coord += 1
    if (S - 0.05) <= lat <= (N + 0.05) and (W - 0.05) <= lon <= (E + 0.05):
        in_box += 1
coords_in_bbox_pct = round(100.0 * in_box / count, 1) if count else 0.0
if coords_in_bbox_pct < 80:
    anomalies.append("仅 %.1f%% 坐标落在 bbox(±0.05)内，疑似坐标越界/错误" % coords_in_bbox_pct)

# 8 landmark
lm_rows = 0
lm_has_ctbuh = False
if "地标对照" in sheets:
    lm = wb["地标对照"]
    lmrows = list(lm.iter_rows(values_only=True))
    if lmrows:
        lm_data = lmrows[1:]
        lm_rows = len([r for r in lm_data if any(c is not None and str(c).strip() != "" for c in r)])
        joined = " ".join(str(c) for r in lmrows for c in r if c is not None)
        lm_has_ctbuh = "CTBUH" in joined.upper()
if lm_rows == 0:
    anomalies.append("地标对照 sheet 为空")
if not lm_has_ctbuh:
    anomalies.append("地标对照 sheet 未含 CTBUH 来源")

# 9 disclosure_ok
disclosure_ok = False
if "说明" in sheets:
    sm = wb["说明"]
    txt = " ".join(str(c) for r in sm.iter_rows(values_only=True) for c in r if c is not None)
    disclosure_ok = any(k in txt for k in ["下限", "估算", "捕获率"])
if not disclosure_ok:
    anomalies.append("说明 sheet 缺少 下限/估算/捕获率 等诚实标注")

# 10 duplicate_rows  楼名+纬度+经度
seen = Counter()
for r in data:
    key = (str(get(r, "楼宇名称")).strip() if get(r, "楼宇名称") is not None else "",
           get(r, "纬度"), get(r, "经度"))
    seen[key] += 1
duplicate_rows = sum(c - 1 for c in seen.values() if c > 1)
if duplicate_rows > 0:
    anomalies.append("发现 %d 行重复(楼名+坐标完全相同)" % duplicate_rows)

# 11 named_count
named_count = sum(1 for r in data
                  if get(r, "楼宇名称") is not None and str(get(r, "楼宇名称")).strip() != "")

wb.close()

# pass 判定: 格式类 1,2,4,5,6 全过 且 无严重异常
severe = (not all_height_ge32) or (not band_consistent) or (not tier_sum_ok) \
         or (not columns_ok) or (not sheets_ok) or (lm_rows == 0) or bool(big) or (duplicate_rows > 0)
verdict_pass = sheets_ok and columns_ok and tier_sum_ok and band_consistent and all_height_ge32 and not severe

out = {
    "city": "棉兰(medan)",
    "pass": bool(verdict_pass),
    "count": count,
    "checks": {
        "sheets_ok": bool(sheets_ok),
        "columns_ok": bool(columns_ok),
        "tier_sum_ok": bool(tier_sum_ok),
        "band_consistent": bool(band_consistent),
        "all_height_ge32": bool(all_height_ge32),
        "coords_in_bbox_pct": coords_in_bbox_pct,
        "landmark_rows": lm_rows,
        "landmark_has_ctbuh": bool(lm_has_ctbuh),
        "disclosure_ok": bool(disclosure_ok),
        "duplicate_rows": duplicate_rows,
        "named_count": named_count,
    },
    "extra": {
        "sheets": sheets,
        "tier_counts": dict(tier_counter),
        "max_height": maxh,
        "below32": len(below32),
        "missing_height": missing_h,
        "valid_coord": valid_coord,
        "in_box": in_box,
        "band_bad": band_bad,
    },
    "anomalies": anomalies,
}
print(json.dumps(out, ensure_ascii=True))
