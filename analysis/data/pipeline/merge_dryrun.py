# -*- coding: utf-8 -*-
"""试算：另一agent的楼里，有多少是我们漏掉的(只读，不改任何文件)。
逻辑：他们≥28m逐栋 → 过滤高度≥32m(我们口径) → 与我们清单空间匹配 →
距离>阈值视为"我们没有的漏网楼"。报告各城候选数 + 距离敏感度 + 置信度分布。"""
import json, glob, os, math
from openpyxl import load_workbook

SRC = r"analysis/data/other_agent_lists"  # 对方 agent 提供的各城高层清单 xlsx，未入库，需自备
OUT = r"analysis\output"

# 可处理10城: 他们文件英文前缀 -> 我们中文文件
CITY = {
    "Bandung":"万隆","Bekasi":"勿加泗","Tangerang":"唐格朗","Bogor":"茂物",
    "Semarang":"三宝垄","Depok":"德波","Surabaya":"泗水","Makassar":"望加锡",
    "Batam":"巴淡","Palembang":"巨港",
}

def find_header_row(ws, maxscan=6):
    """返回(表头行idx, 行列表)。表头行=含纬度/lat 且含 高度/height 的行。"""
    rows = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows[:maxscan]):
        cells = [str(c).lower() if c is not None else "" for c in r]
        has_lat = any(("纬" in c) or ("lat" in c) for c in cells)
        has_h = any(("高度" in c) or ("height" in c) or (c.strip()=="高度 (m)") for c in cells)
        if has_lat and has_h:
            return i, rows
    return 0, rows

def col_idx(header, kind):
    """按关键字找列；排除"几何中心"。"""
    for i, h in enumerate(header):
        hl = str(h).lower() if h is not None else ""
        if "中心" in hl: continue
        if kind=="lat" and (("纬" in hl) or ("lat" in hl)): return i
        if kind=="lon" and (("经" in hl) or ("lon" in hl)): return i
        if kind=="h" and (("高度" in hl) or ("height" in hl)): return i
        if kind=="conf" and (("置信" in hl) or ("conf" in hl)): return i
    return None

def their_buildings(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    # 取名字含"28"的sheet = 全量(≥28m)
    sheet = next((s for s in wb.sheetnames if "28" in s), None)
    ws = wb[sheet]
    hidx, rows = find_header_row(ws)
    header = [str(c) if c is not None else "" for c in rows[hidx]]
    ci = {k: col_idx(header, k) for k in ("lat","lon","h","conf")}
    out = []
    for r in rows[hidx+1:]:
        try:
            lat = float(r[ci["lat"]]); lon = float(r[ci["lon"]]); h = float(r[ci["h"]])
        except (TypeError, ValueError, IndexError):
            continue
        conf = None
        if ci["conf"] is not None and ci["conf"] < len(r):
            try: conf = float(r[ci["conf"]])
            except (TypeError, ValueError): pass
        out.append((lat, lon, h, conf))
    wb.close()
    return out, sheet

def our_buildings(cn):
    path = os.path.join(OUT, f"{cn}高层建筑清单.xlsx")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]]
    li = next(i for i,h in enumerate(header) if "纬" in h)
    oi = next(i for i,h in enumerate(header) if "经" in h)
    out = []
    for r in rows[1:]:
        try: out.append((float(r[li]), float(r[oi])))
        except (TypeError, ValueError): pass
    wb.close()
    return out

def min_dist_m(lat, lon, ours):
    """到我们最近楼的距离(米)，等距矩形近似。"""
    if not ours: return 1e9
    coslat = math.cos(math.radians(lat))
    best = 1e18
    for (la, lo) in ours:
        dy = (la-lat)*111320.0
        dx = (lo-lon)*111320.0*coslat
        d = dx*dx+dy*dy
        if d < best: best = d
    return math.sqrt(best)

TIERS = [(80,"≥80"),(64,"≥64"),(48,"≥48"),(40,"≥40"),(32,"≥32")]
def tier_of(h):
    for t,_ in TIERS:
        if h>=t: return t
    return None

report = []
for en, cn in CITY.items():
    tp = os.path.join(SRC, f"{en}_高层建筑.xlsx")
    theirs, sheet = their_buildings(tp)
    ours = our_buildings(cn)
    ge32 = [b for b in theirs if b[2] >= 32]
    # 对每栋≥32的算最近距离
    dists = [(b, min_dist_m(b[0], b[1], ours)) for b in ge32]
    def missed(thr): return [b for b,d in dists if d > thr]
    m30 = missed(30); m20 = missed(20); m50 = missed(50)
    # 30m口径下漏网楼的分档分布 & 置信度分布
    tierdist = {}
    confbins = {"≥0.8":0,"0.65-0.8":0,"<0.65":0,"无":0}
    for b in m30:
        t = tier_of(b[2]); tierdist[t] = tierdist.get(t,0)+1
        c = b[3]
        if c is None: confbins["无"]+=1
        elif c>=0.8: confbins["≥0.8"]+=1
        elif c>=0.65: confbins["0.65-0.8"]+=1
        else: confbins["<0.65"]+=1
    report.append({
        "城市": cn, "en": en, "sheet": sheet,
        "我们现有": len(ours), "他们≥32m": len(ge32),
        "漏网@20m": len(m20), "漏网@30m": len(m30), "漏网@50m": len(m50),
        "漏网分档@30m": {f"≥{k}m":v for k,v in sorted(tierdist.items(), reverse=True)},
        "漏网置信度@30m": confbins,
    })

out = r"analysis\data\merge_dryrun.json"
with open(out, "w", encoding="utf-8") as f:
    json.dump(report, f, ensure_ascii=False, indent=1)
# ASCII安全摘要
print(f"{'city':<11}{'ours':>6}{'their32':>8}{'miss20':>7}{'miss30':>7}{'miss50':>7}")
for r in report:
    print(f"{r['en']:<11}{r['我们现有']:>6}{r['他们≥32m']:>8}{r['漏网@20m']:>7}{r['漏网@30m']:>7}{r['漏网@50m']:>7}")
print("WROTE", out)
