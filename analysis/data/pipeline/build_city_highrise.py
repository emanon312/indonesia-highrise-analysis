# -*- coding: utf-8 -*-
"""核心融合脚本：把一个城市做成"高层建筑清单 xlsx"。

数据来源（多源交叉）：
- 足迹+面积+坐标：Google Open Buildings V3（{city}_v3.parquet，由 extract_v3.py 产出）
- 高度：Google 2.5D Temporal（复用 googleob 读取逻辑生成 2m 栅格，对每栋取质心邻域最大高）
- 楼名/用途/真实层数：OSM（{city}_osm.json，由 fetch_osm.py 产出，最近邻匹配）—— 可选
- 地标真高+名：CTBUH（{city}_ctbuh.json）—— 可选

口径：4 米/层，分档 ≥32/40/48/64/80 米 = 8/10/12/16/20 层。
输出：output/{城市中文名}高层建筑清单.xlsx（高层清单 / 地标对照 / 说明 三 sheet）。

用法: python build_city_highrise.py <city>
"""
import sys, os, json, math, time, threading
DATA = r'analysis\data'
OUT = r'analysis\output'
sys.path.insert(0, DATA)
import numpy as np, requests, tifffile, pyproj
from scipy import ndimage
from concurrent.futures import ThreadPoolExecutor
import googleob_city as G
import openpyxl
from openpyxl.styles import Font

CELL = G.CELL; CHUNK = G.CHUNK; POOL = G.POOL


def georef(city):
    """重算该城 2m 栅格的地理参考（与 googleob_city 构建逻辑一致）。"""
    c = G.CITIES[city]; bbox = c['bbox']; epsg = c['epsg']
    tr = pyproj.Transformer.from_crs('EPSG:4326', 'EPSG:%d' % epsg, always_xy=True)
    xs, ys = [], []
    for lon in (bbox[0], bbox[2]):
        for lat in (bbox[1], bbox[3]):
            x, y = tr.transform(lon, lat); xs.append(x); ys.append(y)
    bx0, bx1 = min(xs), max(xs); by0, by1 = min(ys), max(ys)
    GX0 = np.floor(bx0/CELL)*CELL; GY1 = np.ceil(by1/CELL)*CELL
    return tr, GX0, GY1, epsg


def build_height_grid(city):
    """加载或构建该城 2m 高度栅格（复用已验证的 COG 读取逻辑），保存/读取 npy。"""
    npy = os.path.join(DATA, 'googleob_%s_height_2m.npy' % city)
    if os.path.exists(npy):
        return np.load(npy)
    c = G.CITIES[city]; bbox = c['bbox']; epsg = c['epsg']
    (bx0, by0, bx1, by1), tiles = G.find_tiles(bbox, epsg, c['manifests'])
    GX0 = np.floor(bx0/CELL)*CELL; GX1 = np.ceil(bx1/CELL)*CELL
    GY0 = np.floor(by0/CELL)*CELL; GY1 = np.ceil(by1/CELL)*CELL
    W2 = int(round((GX1-GX0)/CELL)); H2 = int(round((GY1-GY0)/CELL))
    master = np.zeros((H2, W2), dtype=np.float32)
    for ti, t in enumerate(tiles):
        a = t['affine']; url = t['url']
        x0 = a['translateX']; y0 = a['translateY']; sx = a['scaleX']; sy = a['scaleY']; tw, th = t['w'], t['h']
        c_lo = max(0, int(np.floor((bx0-x0)/sx))); c_hi = min(tw, int(np.ceil((bx1-x0)/sx)))
        r_lo = max(0, int(np.floor((y0-by1)/(-sy)))); r_hi = min(th, int(np.ceil((y0-by0)/(-sy))))
        c_lo -= c_lo % POOL; r_lo -= r_lo % POOL; c_hi -= c_hi % POOL; r_hi -= r_hi % POOL
        if c_hi <= c_lo or r_hi <= r_lo:
            continue
        f = G.HttpRangeFile(url); p = tifffile.TiffFile(f).pages[0]
        offs = p.dataoffsets; cnts = p.databytecounts
        TGRID = math.ceil(tw/CHUNK); BANDS_PER = TGRID*math.ceil(th/CHUNK)
        H = r_hi-r_lo; Wd = c_hi-c_lo
        sub = np.zeros((H, Wd), dtype=np.float32)
        rb0, rb1 = r_lo//CHUNK, (r_hi-1)//CHUNK; cb0, cb1 = c_lo//CHUNK, (c_hi-1)//CHUNK
        blocks = [(rb, cb) for rb in range(rb0, rb1+1) for cb in range(cb0, cb1+1)]
        sl = threading.local()
        def sess():
            if not hasattr(sl, 's'): sl.s = requests.Session()
            return sl.s
        def fetch(b):
            rb, cb = b; idx = BANDS_PER + rb*TGRID + cb
            r = sess().get(url, headers={'Range': 'bytes=%d-%d' % (offs[idx], offs[idx]+cnts[idx]-1)}, timeout=120)
            rs = rb*CHUNK; re = min((rb+1)*CHUNK, th); cs = cb*CHUNK; ce = min((cb+1)*CHUNK, tw)
            return (rs, re, cs, ce, G.decode_tile(r.content, CHUNK, CHUNK)[:re-rs, :ce-cs])
        with ThreadPoolExecutor(max_workers=G.NTHREADS) as ex:
            for rs, re, cs, ce, data in ex.map(fetch, blocks):
                ar0 = max(rs, r_lo); ar1 = min(re, r_hi); ac0 = max(cs, c_lo); ac1 = min(ce, c_hi)
                if ar1 <= ar0 or ac1 <= ac0: continue
                sub[ar0-r_lo:ar1-r_lo, ac0-c_lo:ac1-c_lo] = data[ar0-rs:ar1-rs, ac0-cs:ac1-cs]
        sp = sub.reshape(H//POOL, POOL, Wd//POOL, POOL).max(axis=(1, 3))
        sub_x0 = x0+c_lo*sx; sub_y0 = y0+r_lo*sy
        col_off = int(round((sub_x0-GX0)/CELL)); row_off = int(round((GY1-sub_y0)/CELL))
        r0 = max(0, row_off); c0 = max(0, col_off); r1 = min(H2, row_off+sp.shape[0]); c1 = min(W2, col_off+sp.shape[1])
        if r1 > r0 and c1 > c0:
            blk = sp[r0-row_off:r0-row_off+(r1-r0), c0-col_off:c0-col_off+(c1-c0)]
            np.maximum(master[r0:r1, c0:c1], blk, out=master[r0:r1, c0:c1])
        print('  height tile %d/%d max=%.1f' % (ti+1, len(tiles), master.max()), flush=True)
    np.save(npy, master)
    return master


def band(h):
    if h >= 80: return '≥80米(约20层)'
    if h >= 64: return '≥64米(约16层)'
    if h >= 48: return '≥48米(约12层)'
    if h >= 40: return '≥40米(约10层)'
    if h >= 32: return '≥32米(约8层)'
    return None


def dedup(lat, lon, h, d=15):
    """就近去重：按高度降序贪心，d 米内合并、保留最高。返回保留的位置索引。"""
    n = len(lat)
    order = np.argsort(-h)
    used = np.zeros(n, dtype=bool)
    keep = []
    for i in order:
        if used[i]:
            continue
        keep.append(i)
        dlat = (lat - lat[i]) * 111000.0
        dlon = (lon - lon[i]) * 111000.0 * math.cos(math.radians(lat[i]))
        used |= (dlat*dlat + dlon*dlon) <= d*d
    return np.array(sorted(keep))


def main(city):
    import duckdb
    c = G.CITIES[city]; cn = c['cn']
    t0 = time.time()
    # 1. V3 足迹
    v3 = os.path.join(DATA, '%s_v3.parquet' % city)
    con = duckdb.connect()
    fp = con.execute("SELECT latitude,longitude,area_in_meters,confidence FROM read_parquet('%s')" % v3).fetchall()
    lat = np.array([r[0] for r in fp]); lon = np.array([r[1] for r in fp])
    area = np.array([r[2] for r in fp]); conf = np.array([r[3] for r in fp])
    print('%s V3 足迹 %d 栋' % (city, len(fp)), flush=True)

    # 2. 高度栅格 + 质心单点采样（保守，避免窗口最大值采到邻栋高楼而虚高）
    master = build_height_grid(city)
    tr, GX0, GY1, epsg = georef(city)
    x, y = tr.transform(lon, lat)
    x = np.asarray(x); y = np.asarray(y)
    col = ((x-GX0)/CELL).astype(int); row = ((GY1-y)/CELL).astype(int)
    H2, W2 = master.shape
    inb = (row >= 0) & (row < H2) & (col >= 0) & (col < W2)
    h = np.zeros(len(fp), dtype=np.float32)
    h[inb] = master[row[inb], col[inb]]

    # 3. 筛高层 ≥32m
    hi = h >= 32
    idx = np.where(hi)[0]
    # 按高度降序
    idx = idx[np.argsort(-h[idx])]
    print('%s 高层(≥32m) %d 栋' % (city, len(idx)), flush=True)

    # 4. 就近去重（V3 同栋楼可能切成多足迹；15m 内只留最高，双子塔间距>30m不会误并）
    keep = dedup(lat[idx], lon[idx], h[idx], d=15)
    idx = idx[keep]
    print('%s 去重后 %d 栋' % (city, len(idx)), flush=True)

    # 5. 组装行
    rows = []
    for i in idx:
        hh = float(round(h[i], 1))
        rows.append({
            '序号': 0, '楼宇名称': '', '名称来源': '无',
            '纬度': round(float(lat[i]), 7), '经度': round(float(lon[i]), 7),
            '高度(米)': hh, '高度来源': 'Google估算',
            '层数': int(round(hh/4)), '层数来源': '估算',
            '高度分档': band(hh), '用途分类': '',
            '建筑面积(㎡)': round(float(area[i]), 1) if area[i] else '',
            '备注': 'Google估算下限',
        })

    # 5. OSM / CTBUH 富化（可选，文件存在才做）—— 接口预留，下一步接入
    osm_p = os.path.join(DATA, '%s_osm.json' % city)
    ctbuh_p = os.path.join(DATA, '%s_ctbuh.json' % city)
    n_osm = enrich_osm(rows, osm_p) if os.path.exists(osm_p) else 0
    landmarks = json.load(open(ctbuh_p, encoding='utf-8')) if os.path.exists(ctbuh_p) else []
    n_ct = enrich_ctbuh(rows, landmarks) if landmarks else 0

    # 6. 写 xlsx
    write_xlsx(city, cn, rows, landmarks, n_osm, n_ct)
    # 分档统计
    from collections import Counter
    cnt = Counter(r['高度分档'] for r in rows)
    print('%s 完成 %.0fs  分档: %s' % (city, time.time()-t0, dict(cnt)), flush=True)
    return rows


def enrich_osm(rows, osm_path):
    """OSM 最近邻匹配补楼名/用途/真实层数（≤40m）。osm_json: [{name,btype,levels,lat,lon}]"""
    osm = json.load(open(osm_path, encoding='utf-8'))
    named = [o for o in osm if o.get('name') or o.get('btype') or o.get('levels')]
    if not named:
        return 0
    olat = np.array([o['lat'] for o in named]); olon = np.array([o['lon'] for o in named])
    n = 0
    for r in rows:
        dlat = (olat - r['纬度']) * 111000.0
        dlon = (olon - r['经度']) * 111000.0 * math.cos(math.radians(r['纬度']))
        d2 = dlat*dlat + dlon*dlon
        j = int(np.argmin(d2))
        if d2[j] <= 40*40:
            o = named[j]
            if o.get('name'):
                r['楼宇名称'] = o['name']; r['名称来源'] = 'OSM'
            if o.get('btype'):
                r['用途分类'] = o['btype']
            if o.get('levels'):
                r['层数'] = o['levels']; r['层数来源'] = 'OSM真值'
            n += 1
    return n


def enrich_ctbuh(rows, landmarks):
    """CTBUH 地标按位置匹配，覆盖楼名 + 用真高替换 Google 估高。landmarks: [{name,height_m,floors,lat,lon}]"""
    n = 0
    rlat = np.array([r['纬度'] for r in rows]); rlon = np.array([r['经度'] for r in rows])
    for lm in landmarks:
        if lm.get('lat') is None or lm.get('lon') is None:
            continue
        dlat = (rlat - lm['lat']) * 111000.0
        dlon = (rlon - lm['lon']) * 111000.0 * math.cos(math.radians(lm['lat']))
        d2 = dlat*dlat + dlon*dlon
        j = int(np.argmin(d2))
        if d2[j] <= 150*150:  # 地标定位较粗,放宽到150m
            r = rows[j]
            r['楼宇名称'] = lm['name']; r['名称来源'] = 'CTBUH地标'
            if lm.get('height_m'):
                r['高度(米)'] = lm['height_m']; r['高度来源'] = 'CTBUH真值'
                r['高度分档'] = band(lm['height_m']) or r['高度分档']
            if lm.get('floors'):
                r['层数'] = lm['floors']; r['层数来源'] = 'CTBUH真值'
            n += 1
    return n


COLS = ['序号', '楼宇名称', '名称来源', '纬度', '经度', '高度(米)', '高度来源',
        '层数', '层数来源', '高度分档', '用途分类', '建筑面积(㎡)', '备注']


def write_xlsx(city, cn, rows, landmarks, n_osm, n_ct):
    # 重新按高度降序并编号
    rows.sort(key=lambda r: -r['高度(米)'])
    for i, r in enumerate(rows, 1):
        r['序号'] = i
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '高层清单'
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r[k] for k in COLS])
    ws.freeze_panes = 'A2'
    # 地标对照 sheet
    ws2 = wb.create_sheet('地标对照')
    ws2.append(['序号', '楼宇名称', '真实高度(米)', '层数', '纬度', '经度', '来源'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for i, lm in enumerate(sorted(landmarks, key=lambda x: -(x.get('height_m') or 0)), 1):
        ws2.append([i, lm.get('name', ''), lm.get('height_m', ''), lm.get('floors', ''),
                    lm.get('lat', ''), lm.get('lon', ''), 'CTBUH/维基'])
    # 说明 sheet
    ws3 = wb.create_sheet('说明')
    from collections import Counter
    cnt = Counter(r['高度分档'] for r in rows)
    notes = [
        '%s 高层建筑清单 — 数据说明' % cn, '',
        '【口径】4 米/层，分档：≥32米(8层)/≥40米(10层)/≥48米(12层)/≥64米(16层)/≥80米(20层)',
        '【高度来源】Google Open Buildings 2.5D Temporal 2023（机器学习估算）；150m+ 地标用 CTBUH/维基真实高度覆盖',
        '【足迹/坐标/面积】Google Open Buildings V3（真实建筑轮廓，坐标为质心）',
        '【楼名/用途/真实层数】OSM（OpenStreetMap）最近邻匹配（≤40米），本城匹配到 %d 栋' % n_osm,
        '【地标】CTBUH/维基收录 %d 栋，已匹配覆盖 %d 栋' % (len(landmarks), n_ct), '',
        '本表高层(≥32米)合计 %d 栋。分档分布：' % len(rows),
    ]
    for b in ['≥80米(约20层)', '≥64米(约16层)', '≥48米(约12层)', '≥40米(约10层)', '≥32米(约8层)']:
        notes.append('  %s : %d 栋' % (b, cnt.get(b, 0)))
    notes += [
        '',
        '【重要局限 — 数量为下限】Google 2.5D 高度系统性偏低：以雅加达官方真值校准，各档只抓到',
        '  8层53% / 10层50% / 12层46% / 16层35% / 20层23%（越高漏越多）。',
        '  即本表数量是"至少这么多"，真实更多；如需量级估计可按上述捕获率反推。',
        '【其它局限】高度约99.5米封顶(本项目最高20层/80米在封顶下,影响有限;150m+地标已用CTBUH真高)；',
        '  多数普通楼无楼名/用途(OSM覆盖有限)；无官方用途数据,未剔除公共设施(医院/学校/政府楼可能混入)；',
        '  坐标为建筑质心点;同建筑群多塔各占一行不去重。',
        '【唯一真值锚点】只有雅加达有官方逐栋真值(DPMPTSP LOD2)；本城为 Google+OSM+CTBUH 交叉估算。',
    ]
    for line in notes:
        ws3.append([line])
    ws3['A1'].font = Font(bold=True, size=14)
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, '%s高层建筑清单.xlsx' % cn)
    wb.save(path)
    print('已写 %s (%d 栋)' % (path, len(rows)), flush=True)


if __name__ == '__main__':
    main(sys.argv[1])
