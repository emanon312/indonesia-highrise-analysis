# -*- coding: utf-8 -*-
"""万隆高层建筑清单 Excel 验收审计脚本。结论以 JSON 输出（ASCII）。"""
import json
from openpyxl import load_workbook

PATH = r"analysis\output\万隆高层建筑清单.xlsx"
W, S, E, N = 107.55, -6.97, 107.73, -6.86

EXPECTED_SHEETS = ["高层清单", "地标对照", "说明"]
EXPECTED_COLS = ["序号", "楼宇名称", "名称来源", "纬度", "经度", "高度(米)",
                 "高度来源", "层数", "层数来源", "高度分档", "用途分类",
                 "建筑面积", "备注"]

result = {"anomalies": []}
wb = load_workbook(PATH, read_only=True, data_only=True)
sheets = wb.sheetnames

# 1 sheets_ok
sheets_ok = sheets == EXPECTED_SHEETS
if not sheets_ok:
    result["anomalies"].append("sheet名称/数量不符: 实际=%s" % sheets)

main = wb[sheets[0]]
rows = list(main.iter_rows(values_only=True))
header = list(rows[0]) if rows else []
data_rows = rows[1:] if len(rows) > 1 else []
# 去掉全空尾行
data_rows = [r for r in data_rows if any(c is not None and str(c).strip() != "" for c in r)]

# 2 columns_ok
header_clean = [str(h).strip() if h is not None else "" for h in header]
columns_ok = (len(header_clean) == 13) and (header_clean == EXPECTED_COLS)
if not columns_ok:
    result["anomalies"].append("列名/列数不符: 实际(%d列)=%s" % (len(header_clean), header_clean))

# 列索引
def cidx(name):
    return EXPECTED_COLS.index(name) if name in EXPECTED_COLS else None

i_name = cidx("楼宇名称")
i_lat = cidx("纬度")
i_lon = cidx("经度")
i_h = cidx("高度(米)")
i_band = cidx("高度分档")

count = len(data_rows)

def to_float(v):
    try:
        if v is None or str(v).strip() == "":
            return None
        return float(str(v).strip())
    except Exception:
        return None

# 4 tier_sum_ok
from collections import Counter
band_counter = Counter()
for r in data_rows:
    b = r[i_band] if i_band is not None and i_band < len(r) else None
    band_counter[str(b).strip() if b is not None else ""] += 1
tier_sum = sum(band_counter.values())
tier_sum_ok = (tier_sum == count)

# 5 band_consistent 口径 4米/层
# >=32->8层, >=40->10, >=48->12, >=64->16, >=80->20
# 期望档位：h>=80 ->"≥80"; 64<=h<80->"≥64"; 48<=h<64->"≥48"; 40<=h<48->"≥40"; 32<=h<40->"≥32"
def expected_band(h):
    if h is None:
        return None
    if h >= 80:
        return "≥80"
    if h >= 64:
        return "≥64"
    if h >= 48:
        return "≥48"
    if h >= 40:
        return "≥40"
    if h >= 32:
        return "≥32"
    return "<32"

def norm_band(b):
    """归一化分档文本，提取阈值数字"""
    if b is None:
        return None
    s = str(b).strip()
    import re
    m = re.search(r"(\d+)", s)
    if m:
        return "≥" + m.group(1)
    return s

inconsistent = 0
inconsistent_examples = []
heights = []
ge32_fail = 0
for r in data_rows:
    h = to_float(r[i_h]) if i_h is not None and i_h < len(r) else None
    heights.append(h)
    b = r[i_band] if i_band is not None and i_band < len(r) else None
    nb = norm_band(b)
    eb = expected_band(h)
    if h is not None:
        if h < 32:
            ge32_fail += 1
        if nb != eb:
            inconsistent += 1
            if len(inconsistent_examples) < 8:
                nm = r[i_name] if i_name is not None and i_name < len(r) else ""
                inconsistent_examples.append("%s h=%s 档=%s 期望=%s" % (str(nm)[:20], h, nb, eb))
    else:
        # 高度缺失也算不自洽
        inconsistent += 1

band_consistent = (inconsistent == 0)
all_height_ge32 = (ge32_fail == 0) and all(h is not None for h in heights)

# 6 报告高度异常大
big = [(r[i_name] if i_name is not None else "", to_float(r[i_h])) for r in data_rows]
big_anom = [(str(n)[:20], h) for n, h in big if h is not None and h > 250]

# 7 coords_in_bbox_pct
in_box = 0
valid_coord = 0
oob_examples = []
for r in data_rows:
    lat = to_float(r[i_lat]) if i_lat is not None and i_lat < len(r) else None
    lon = to_float(r[i_lon]) if i_lon is not None and i_lon < len(r) else None
    if lat is not None and lon is not None:
        valid_coord += 1
        if (S - 0.05) <= lat <= (N + 0.05) and (W - 0.05) <= lon <= (E + 0.05):
            in_box += 1
        else:
            if len(oob_examples) < 6:
                nm = r[i_name] if i_name is not None else ""
                oob_examples.append("%s lat=%s lon=%s" % (str(nm)[:18], lat, lon))
coords_in_bbox_pct = round(100.0 * in_box / count, 1) if count else 0.0

# 8 landmark sheet
landmark_rows = 0
landmark_has_ctbuh = False
if "地标对照" in sheets:
    ls = wb["地标对照"]
    lrows = list(ls.iter_rows(values_only=True))
    lhead = lrows[0] if lrows else []
    ldata = [r for r in lrows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
    landmark_rows = len(ldata)
    # 全表文本搜 CTBUH
    blob = " ".join(str(c) for r in lrows for c in r if c is not None)
    landmark_has_ctbuh = ("CTBUH" in blob.upper())
else:
    result["anomalies"].append("缺少地标对照sheet")

# 9 disclosure_ok
disclosure_ok = False
if "说明" in sheets:
    ds = wb["说明"]
    dblob = " ".join(str(c) for r in ds.iter_rows(values_only=True) for c in r if c is not None)
    disclosure_ok = any(k in dblob for k in ["下限", "估算", "捕获率"])
    if not disclosure_ok:
        result["anomalies"].append("说明sheet缺少诚实标注关键词(下限/估算/捕获率)")
else:
    result["anomalies"].append("缺少说明sheet")

# 10 duplicate_rows 楼名+坐标完全相同
seen = {}
dup = 0
for r in data_rows:
    nm = r[i_name] if i_name is not None and i_name < len(r) else None
    lat = to_float(r[i_lat]) if i_lat is not None else None
    lon = to_float(r[i_lon]) if i_lon is not None else None
    key = (str(nm).strip() if nm is not None else "", lat, lon)
    seen[key] = seen.get(key, 0) + 1
for k, v in seen.items():
    if v > 1:
        dup += (v - 1)

# 11 named_count
named_count = 0
for r in data_rows:
    nm = r[i_name] if i_name is not None and i_name < len(r) else None
    if nm is not None and str(nm).strip() != "":
        named_count += 1

wb.close()

# anomalies 汇总
if coords_in_bbox_pct < 90:
    result["anomalies"].append("坐标越界比例较高: 仅%.1f%%在bbox内; 例: %s" % (coords_in_bbox_pct, oob_examples))
if inconsistent > 0:
    result["anomalies"].append("分档不自洽%d行; 例: %s" % (inconsistent, inconsistent_examples))
if dup > 0:
    result["anomalies"].append("重复行%d条" % dup)
if big_anom:
    result["anomalies"].append("高度>250m可疑(二三线城市): %s" % big_anom)
if ge32_fail > 0:
    result["anomalies"].append("有%d行高度<32米" % ge32_fail)
if landmark_rows == 0:
    result["anomalies"].append("地标对照sheet为空")
if not landmark_has_ctbuh and landmark_rows > 0:
    result["anomalies"].append("地标对照sheet无CTBUH来源")
if named_count < count * 0.3:
    result["anomalies"].append("命名稀疏: 仅%d/%d行有楼名" % (named_count, count))

# pass 规则: 格式类(1,2,4,5,6)全通过且无严重异常
severe = (not sheets_ok) or (not columns_ok) or (not tier_sum_ok) or (not band_consistent) or (not all_height_ge32)
overall_pass = not severe

result["city"] = "万隆(bandung)"
result["pass"] = overall_pass
result["count"] = count
result["checks"] = {
    "sheets_ok": sheets_ok,
    "columns_ok": columns_ok,
    "tier_sum_ok": tier_sum_ok,
    "band_consistent": band_consistent,
    "all_height_ge32": all_height_ge32,
    "coords_in_bbox_pct": coords_in_bbox_pct,
    "landmark_rows": landmark_rows,
    "landmark_has_ctbuh": landmark_has_ctbuh,
    "disclosure_ok": disclosure_ok,
    "duplicate_rows": dup,
    "named_count": named_count,
}
result["_extra"] = {
    "sheets": sheets,
    "header": header_clean,
    "band_counter": dict(band_counter),
    "tier_sum": tier_sum,
    "valid_coord": valid_coord,
    "in_box": in_box,
    "inconsistent_examples": inconsistent_examples,
    "big_anom": big_anom,
}

print(json.dumps(result, ensure_ascii=True))
