# -*- coding: utf-8 -*-
"""
融合脚本 v2：生成与雅加达格式完全对齐的「高层建筑清单 xlsx」。
输出格式：3 sheet / 15 列（与雅加达 build_excel.py 对齐）。

数据来源：
- 足迹+面积+坐标：Google Open Buildings V3
- 高度：Google 2.5D Temporal
- 楼名/类型/地址/公共设施标签：OSM（_osm_full.json，最近邻 ≤40m）
- 地标真高+名：CTBUH（_ctbuh.json）
- 行政区/街道办：GADM Level 3 空间匹配

口径：4米/层，分档 ≥32/40/48/64/80 米 = 8/10/12/16/20 层。
用法: python build_city_highrise_v2.py <city>
输出: output/{城市中文名}高层建筑清单.xlsx
"""
import sys, os, json, math, time, re, collections
# 路径改为相对定位（原 Desktop 绝对路径已失效，2026-07-10 交接时修）。
# 本脚本在 analysis/data/pipeline/：上跳两级 = analysis/data（DATA），三级 = 工作区根（拼 output）。
_BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA = os.path.join(_BASE, 'data')
OUT = os.path.join(_BASE, 'output')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))  # import 同目录 googleob_city
import numpy as np, requests, tifffile, pyproj
from scipy import ndimage
from concurrent.futures import ThreadPoolExecutor
import googleob_city as G
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

CELL = G.CELL; CHUNK = G.CHUNK; POOL = G.POOL

# ===================== 公共设施剔除规则（复用 build_excel.py） =====================

# 基于 OSM building/amenity 标签的公共类别 → 剔除
OSM_PUBLIC_BUILDING = {
    '医院', '学校', '高校', '政府', '公共', '宗教', '交通/停车', '体育设施',
}

OSM_PUBLIC_AMENITY = {
    'school', 'college', 'university', 'kindergarten',  # 学校
    'hospital', 'clinic', 'doctors', 'dentist',         # 医疗
    'place_of_worship', 'mosque', 'church', 'cathedral', # 宗教
    'police', 'fire_station', 'prison', 'courthouse',    # 政府/公共服务
    'townhall', 'embassy', 'public_building',            # 政府
    'library', 'community_centre', 'social_facility',    # 公共
    'bus_station', 'ferry_terminal',                     # 交通
}

# 居住豁免：这些 OSM 类型即使命中公共标签也保留
RESIDENTIAL_TYPES = {'公寓', '住宅', '宿舍'}

# 楼名关键词正则（复用 build_excel.py，加印尼语和中文）
NAME_RULES = [
    ("医院", re.compile(
        r"\b(rumah sakit|rsud|rsup|rsia|rsab|rsu\b|rs\b|klinik|puskesmas|pusat kes|hospital|"
        r"poliklinik|balai pengobatan|medical cent|medikal|medis)\b", re.I)),
    ("学校", re.compile(
        r"\b(sekolah|sdn?\b|smpn?\b|sman?\b|smkn?\b|paud\b|universitas|university|kampus|campus|"
        r"pesantren|madrasah|institut|institute|politeknik|poltek|akademi|college|"
        r"sekolah tinggi|dinas pendidikan|fakultas|gedung kuliah|sd\s|smp\s|sma\s|smk\s"
        r"|中小学|小学|中学|大学|学院|学校)\b", re.I)),
    ("政府/公共服务", re.compile(
        r"\b(kementerian|kementrian|kemenko|pemerintah|pemprov|pemkot|pemkab|pemda|"
        r"gubernur|kejaksaan|kejagung|pengadilan|mahkamah|kepolisian|polda|polres|polsek|polri|"
        r"mabes|kodam|kodim|korem|koramil|lanud|lantamal|dprd|bappeda|bappenas|samsat|imigrasi|"
        r"kanwil|ditjen|balai kota|balaikota|wali kota|walikota|kedutaan|konsulat|embassy|consulate"
        r"|kantor (dinas|pemerintah|wali|gubernur|camat|kelurahan|pajak|pos|imigrasi)|"
        r"direktorat jenderal|\bdinas\b|duta besar|instalasi|markas|komando|resor|resimen"
        r"|政府|警察|法院|检察院|移民|海关|税务|市政)\b", re.I)),
    ("宗教场所", re.compile(
        r"\b(masjid|musholla|mushola|surau|langgar|gereja|vihara|pura|klenteng|kelenteng|"
        r"cathedral|church|mosque|temple|wihara|candi|kuil| chapel\b|"
        r"清真寺|教堂|寺庙|佛寺)\b", re.I)),
    ("军事/国防", re.compile(
        r"\b(militer|tentara|angkatan|komando|batalyon|brigif|y(on|udha|ako|onif)|"
        r"markas besar|mabes\b|\btni\b|\bpolri\b|kesatuan|detasemen|"
        r"军事|部队|军营|国防)\b", re.I)),
    ("交通/停车", re.compile(
        r"\b(terminal\s|stasiun|station|bandara|airport|pelabuhan|port\b|terminal$|"
        r"parkir|parking|gedung parkir|halte|shelter|"
        r"车站|机场|港口|停车场|航站楼)\b", re.I)),
    ("体育设施", re.compile(
        r"\b(stadion|stadium|gelora|gelanggang|sport|gymnasium|velodrome|"
        r"kolam renang|gedung olahraga|gor\b|"
        r"体育场|体育馆|游泳馆|健身)\b", re.I)),
    ("文化设施", re.compile(
        r"\b(museum|galeri|gallery|perpustakaan|library|gedung pertunjukan|"
        r"gedung kesenian|planetarium|"
        r"博物馆|图书馆|美术馆|展览馆)\b", re.I)),
]

# 这些商业属性词说明建筑不是公共设施
COMMERCIAL_HINTS = re.compile(
    r"\b(apartemen|apartment|\brusun\b|rumah susun|asrama|residence|kondominium|condominium|"
    r"hotel|mall|plaza|tower|gedung perkantoran|office|ruko|rukan|pertokoan|"
    r"supermarket|hypermarket|apotek|pharmacy|bank|restoran|cafe|"
    r"公寓|酒店|写字楼|商场|广场|大厦|银行|餐厅)\b", re.I)


def is_residential(btype):
    """判断是否居住类"""
    return btype in RESIDENTIAL_TYPES


def classify_public(rec):
    """对一栋建筑判断是否应剔除为公共设施。
    rec: dict with keys: 楼宇名称, _btype, _amenity, _office, _religion, _building_use
    返回: (keep/drop, 剔除类别, 判定依据, 命中词)"""
    btype = rec.get('_btype') or ''
    amenity = rec.get('_amenity') or ''
    office = rec.get('_office') or ''
    religion_tag = rec.get('_religion') or ''
    building_use = rec.get('_building_use') or ''
    name = rec.get('楼宇名称', '') or ''

    residential = is_residential(btype)

    # 第1层：OSM building 类型是公共设施 → 剔除
    if btype in OSM_PUBLIC_BUILDING:
        if residential:
            return ('保留', '', '居住豁免（OSM类型）', btype)
        return ('剔除', btype, 'OSM building标签', btype)

    # 第2层：OSM amenity 标签是公共设施 → 剔除
    if amenity and amenity in OSM_PUBLIC_AMENITY:
        if residential:
            return ('保留', '', '居住豁免（amenity）', amenity)
        # 但也检查名字中是否有商业暗示
        if COMMERCIAL_HINTS.search(name):
            return ('保留', '', '商业名豁免（amenity）', amenity)
        return ('剔除', amenity_label(amenity), 'OSM amenity标签', amenity)

    # 第2b层：office/religion 标签
    if office and office in ('government', 'educational_institution', 'ngo'):
        return ('剔除', '政府/公共服务', 'OSM office标签', office)
    if religion_tag:
        return ('剔除', '宗教场所', 'OSM religion标签', religion_tag)

    # 第3层：楼名关键词兜底
    for cat, pat in NAME_RULES:
        m = pat.search(name)
        if m:
            # 排除商业楼名（如 "Hotel Hospital" 之类的）
            if COMMERCIAL_HINTS.search(name):
                continue
            return ('剔除', cat, '楼名关键词', m.group(0).strip())

    return ('保留', '', '', '')


def amenity_label(amenity):
    """将 amenity 值映射到中文类别"""
    m = {
        'school': '学校', 'college': '高校', 'university': '高校', 'kindergarten': '学校',
        'hospital': '医院', 'clinic': '医院', 'doctors': '医院', 'dentist': '医院',
        'place_of_worship': '宗教场所', 'mosque': '宗教场所', 'church': '宗教场所',
        'cathedral': '宗教场所', 'temple': '宗教场所',
        'police': '政府/公共服务', 'fire_station': '政府/公共服务',
        'prison': '政府/公共服务', 'courthouse': '政府/公共服务',
        'townhall': '政府/公共服务', 'embassy': '政府/公共服务',
        'public_building': '政府/公共服务',
        'library': '文化设施', 'community_centre': '政府/公共服务',
        'social_facility': '政府/公共服务',
        'bus_station': '交通/停车', 'ferry_terminal': '交通/停车',
        'parking': '交通/停车',
    }
    return m.get(amenity, amenity)


# ===================== 行政分区（OSM Overpass + shapely，无需 geopandas） =====================

def load_admin_boundaries(city):
    """加载 OSM 抓取的 kecamatan 边界，构建 shapely 多边形列表"""
    from shapely.geometry import Polygon, Point as ShpPoint
    path = os.path.join(DATA, '%s_admin_boundaries.json' % city)
    if not os.path.exists(path):
        print('  行政边界文件 %s 不存在，跳过行政区匹配' % path, flush=True)
        return None

    data = json.load(open(path, encoding='utf-8'))
    kotas = data.get('kotas', [])
    kecamatans = data.get('kecamatans', [])

    # 构建 (行政区名, 街道办名, Polygon) 列表
    regions = []

    # 先处理 kota 级别（行政区）
    kota_polys = {}  # name -> [Polygon]
    for kt in kotas:
        name = kt['name']
        polys = []
        for ring in kt['polygons']:
            if len(ring) >= 3:
                try:
                    poly = Polygon(ring)
                    if not poly.is_valid:
                        poly = poly.buffer(0)
                    if poly.is_valid and not poly.is_empty and poly.area > 0:
                        polys.append(poly)
                except Exception:
                    pass
        if polys:
            kota_polys[name] = polys

    # 处理 kecamatan 级别（街道办区）
    for kc in kecamatans:
        kec_name = kc['name']
        polys = []
        for ring in kc['polygons']:
            if len(ring) >= 3:
                try:
                    poly = Polygon(ring)
                    if not poly.is_valid:
                        poly = poly.buffer(0)  # 修复自交等几何问题
                    if poly.is_valid and not poly.is_empty and poly.area > 0:
                        polys.append(poly)
                except Exception:
                    pass
        if polys:
            # 尝试匹配所属 kota
            kota_name = ''
            for kname, kpolys in kota_polys.items():
                # 检查 kecamatan 是否与某个 kota 相交
                for kp in kpolys:
                    try:
                        for p in polys:
                            if p.intersects(kp) or kp.contains(p) or p.contains(kp):
                                kota_name = kname
                                break
                        if kota_name:
                            break
                    except Exception:
                        pass
                if kota_name:
                    break
            regions.append((kota_name, kec_name, polys))

    print('  行政边界: %d 个区域 (%d kota, %d kecamatan)' %
          (len(regions), len(kotas), len(kecamatans)), flush=True)
    return regions


def assign_admin(lats, lons, regions):
    """对坐标数组做 point-in-polygon，返回 (行政区列表, 街道办列表)。
    使用 shapely 手动遍历，不依赖 geopandas。"""
    from shapely.geometry import Point as ShpPoint

    n = len(lats)
    districts = [''] * n
    kecamatans = [''] * n

    if not regions:
        return (districts, kecamatans)

    for i in range(n):
        pt = ShpPoint(lons[i], lats[i])
        best_area = float('inf')
        for kota_name, kec_name, polys in regions:
            for poly in polys:
                try:
                    if poly.contains(pt):
                        # 选面积最小的包含多边形（最精确的行政边界）
                        area = poly.area
                        if area < best_area:
                            best_area = area
                            districts[i] = kota_name
                            kecamatans[i] = kec_name
                        break
                except Exception:
                    continue

    matched = sum(1 for d in districts if d)
    print('  行政匹配: %d/%d' % (matched, n), flush=True)
    return (districts, kecamatans)


# ===================== 高度栅格/地理参考 =====================

def georef(city):
    c = G.CITIES[city]; bbox = c['bbox']; epsg = c['epsg']
    tr = pyproj.Transformer.from_crs('EPSG:4326', 'EPSG:%d' % epsg, always_xy=True)
    xs, ys = [], []
    for lon in (bbox[0], bbox[2]):
        for lat in (bbox[1], bbox[3]):
            x, y = tr.transform(lon, lat); xs.append(x); ys.append(y)
    bx0, bx1 = min(xs), max(xs); by0, by1 = min(ys), max(ys)
    GX0 = np.floor(bx0 / CELL) * CELL; GY1 = np.ceil(by1 / CELL) * CELL
    return tr, GX0, GY1, epsg


def build_height_grid(city):
    npy = os.path.join(DATA, 'raw', 'googleob_%s_height_2m.npy' % city)  # npy 已归档到 data/raw/
    if os.path.exists(npy):
        print('  高度栅格已有, 直接加载', flush=True)
        return np.load(npy)
    # 如果 npy 不存在，这里需要调原逻辑构建（保留但通常已有）
    raise FileNotFoundError('高度栅格 npy 不存在，请先跑原 build_city_highrise.py 生成: %s' % npy)


def band(h):
    if h >= 80: return '≥80米(约20层)'
    if h >= 64: return '≥64米(约16层)'
    if h >= 48: return '≥48米(约12层)'
    if h >= 40: return '≥40米(约10层)'
    if h >= 32: return '≥32米(约8层)'
    return None


def dedup(lat, lon, h, d=15):
    n = len(lat)
    order = np.argsort(-h)
    used = np.zeros(n, dtype=bool)
    keep = []
    for i in order:
        if used[i]:
            continue
        keep.append(i)
        dlat = (lat - lat[i]) * 111000.0
        dlon = (lon - lon[i]) * 111000.0 * math.cos(math.radians(lat[i]))
        used |= (dlat * dlat + dlon * dlon) <= d * d
    return np.array(sorted(keep))


# ===================== 主流程 =====================

def main(city):
    import duckdb
    c = G.CITIES[city]; cn = c['cn']
    t0 = time.time()

    # 1. V3 足迹
    v3_path = os.path.join(DATA, '%s_v3.parquet' % city)
    con = duckdb.connect()
    fp = con.execute(
        "SELECT latitude,longitude,area_in_meters,confidence FROM read_parquet('%s')" % v3_path
    ).fetchall()
    lat = np.array([r[0] for r in fp]); lon = np.array([r[1] for r in fp])
    area = np.array([r[2] for r in fp]); conf = np.array([r[3] for r in fp])
    print('%s V3 足迹 %d 栋' % (city, len(fp)), flush=True)

    # 2. 高度栅格
    master = build_height_grid(city)
    tr, GX0, GY1, epsg = georef(city)
    x, y = tr.transform(lon, lat)
    x = np.asarray(x); y = np.asarray(y)
    col = ((x - GX0) / CELL).astype(int)
    row = ((GY1 - y) / CELL).astype(int)
    H2, W2 = master.shape
    inb = (row >= 0) & (row < H2) & (col >= 0) & (col < W2)
    h = np.zeros(len(fp), dtype=np.float32)
    h[inb] = master[row[inb], col[inb]]

    # 3. 筛高层 ≥32m
    hi = h >= 32
    idx = np.where(hi)[0]
    idx = idx[np.argsort(-h[idx])]
    print('%s 高层(≥32m) %d 栋' % (city, len(idx)), flush=True)

    # 4. 去重
    keep = dedup(lat[idx], lon[idx], h[idx], d=15)
    idx = idx[keep]
    print('%s 去重后 %d 栋' % (city, len(idx)), flush=True)

    # 5. 组装行（初始）
    rows = []
    for i in idx:
        hh = float(round(h[i], 1))
        rows.append({
            '序号': 0,
            '楼宇名称': '',
            '名称来源': '无',
            'OSM匹配距离(米)': '',
            '建筑全名(官方)': '—',
            '纬度': round(float(lat[i]), 7),
            '经度': round(float(lon[i]), 7),
            '高度(米)': hh,
            '层数': int(round(hh / 4)),
            '高度分档': band(hh),
            '用途分类': '',
            '行政区': '',
            '街道办区': '',
            '地址': '',
            '建筑面积(㎡)': round(float(area[i]), 1) if area[i] else '',
            # 内部字段（不输出到 Excel）
            '_conf': float(conf[i]) if conf[i] else 0,
            '_lat': float(lat[i]),
            '_lon': float(lon[i]),
            '_btype': '',
            '_amenity': '',
            '_office': '',
            '_religion': '',
            '_building_use': '',
        })

    # 6. OSM 富化（用 _osm_full.json）
    osm_path = os.path.join(DATA, '%s_osm_full.json' % city)
    if not os.path.exists(osm_path):
        osm_path = os.path.join(DATA, '%s_osm.json' % city)  # 降级
    n_osm = enrich_osm_v2(rows, osm_path)

    # 7. CTBUH 富化
    ctbuh_path = os.path.join(DATA, '%s_ctbuh.json' % city)
    landmarks = json.load(open(ctbuh_path, encoding='utf-8')) if os.path.exists(ctbuh_path) else []
    n_ct = enrich_ctbuh(rows, landmarks) if landmarks else 0

    # 8. 行政分区
    regions = load_admin_boundaries(city)
    if regions:
        rlats = np.array([r['_lat'] for r in rows])
        rlons = np.array([r['_lon'] for r in rows])
        districts, kecamatans = assign_admin(rlats, rlons, regions)
        for r, d, k in zip(rows, districts, kecamatans):
            if d:
                r['行政区'] = d
            if k:
                r['街道办区'] = k

    # 9. 公共设施筛选
    keep_rows, drop_rows = [], []
    drop_cat = collections.Counter()
    for r in rows:
        status, cat, basis, hit = classify_public(r)
        r['_drop_cat'] = cat
        r['_drop_basis'] = basis
        r['_drop_hit'] = hit
        if status == '剔除':
            drop_rows.append(r)
            drop_cat[cat] += 1
        else:
            keep_rows.append(r)

    print('  保留 %d / 剔除 %d' % (len(keep_rows), len(drop_rows)), flush=True)
    print('  剔除明细: %s' % dict(drop_cat.most_common()), flush=True)

    # 10. 重编号 + 写 Excel
    keep_rows.sort(key=lambda r: -r['高度(米)'])
    for i, r in enumerate(keep_rows, 1):
        r['序号'] = i

    drop_rows.sort(key=lambda r: -r['高度(米)'])
    for i, r in enumerate(drop_rows, 1):
        r['序号'] = i

    write_xlsx_v2(city, cn, keep_rows, drop_rows, landmarks, n_osm, n_ct)

    cnt = collections.Counter(r['高度分档'] for r in keep_rows)
    elapsed = time.time() - t0
    print('%s 完成 %.0fs  保留%d栋 分档: %s' % (city, elapsed, len(keep_rows), dict(cnt)), flush=True)
    return keep_rows, drop_rows


# ===================== OSM 富化（v2，扩展标签） =====================

def enrich_osm_v2(rows, osm_path):
    """OSM 最近邻匹配 ≤40m：补楼名/类型/层数/地址/amenity，并记录匹配距离"""
    osm = json.load(open(osm_path, encoding='utf-8'))
    named = [o for o in osm if o.get('name') or o.get('btype') or o.get('levels')
             or o.get('amenity') or o.get('addr_full') or o.get('addr_street')]
    if not named:
        print('  OSM 无可富化记录', flush=True)
        return 0
    olat = np.array([o['lat'] for o in named])
    olon = np.array([o['lon'] for o in named])
    n = 0
    for r in rows:
        dlat = (olat - r['_lat']) * 111000.0
        dlon = (olon - r['_lon']) * 111000.0 * math.cos(math.radians(r['_lat']))
        d2 = dlat * dlat + dlon * dlon
        j = int(np.argmin(d2))
        if d2[j] <= 40 * 40:
            o = named[j]
            dist = round(float(math.sqrt(d2[j])), 1)
            r['OSM匹配距离(米)'] = dist

            if o.get('name'):
                if not r['楼宇名称'] or r['名称来源'] == '无':
                    r['楼宇名称'] = o['name']
                    r['名称来源'] = 'OSM补充'

            if o.get('btype') and not r['用途分类']:
                r['用途分类'] = o['btype']

            if o.get('levels'):
                r['层数'] = o['levels']

            # 补充内部字段（用于公共设施筛选）
            r['_btype'] = o.get('btype') or ''
            r['_amenity'] = o.get('amenity') or ''
            r['_office'] = o.get('office') or ''
            r['_religion'] = o.get('religion') or ''
            r['_building_use'] = o.get('building_use') or ''

            # 地址
            addr = o.get('addr_full') or ''
            if not addr:
                parts = []
                if o.get('addr_street'):
                    parts.append(o['addr_street'])
                if o.get('addr_city'):
                    parts.append(o['addr_city'])
                addr = ', '.join(parts)
            if addr and not r['地址']:
                r['地址'] = addr

            n += 1
    print('  OSM 匹配: %d/%d 栋' % (n, len(rows)), flush=True)
    return n


# ===================== CTBUH 富化 =====================

def enrich_ctbuh(rows, landmarks):
    n = 0
    rlat = np.array([r['_lat'] for r in rows])
    rlon = np.array([r['_lon'] for r in rows])
    for lm in landmarks:
        if lm.get('lat') is None or lm.get('lon') is None:
            continue
        dlat = (rlat - lm['lat']) * 111000.0
        dlon = (rlon - lm['lon']) * 111000.0 * math.cos(math.radians(lm['lat']))
        d2 = dlat * dlat + dlon * dlon
        j = int(np.argmin(d2))
        if d2[j] <= 150 * 150:
            r = rows[j]
            r['楼宇名称'] = lm['name']
            r['名称来源'] = 'CTBUH地标'
            if lm.get('height_m'):
                r['高度(米)'] = lm['height_m']
                r['高度分档'] = band(lm['height_m']) or r['高度分档']
            if lm.get('floors'):
                r['层数'] = lm['floors']
            n += 1
    print('  CTBUH 匹配: %d' % n, flush=True)
    return n


# ===================== Excel 输出（雅加达格式） =====================

# 与雅加达 build_excel.py 完全对齐的 15 列
COLS_KEEP = [
    '序号', '楼宇名称', '名称来源', 'OSM匹配距离(米)', '建筑全名(官方)',
    '纬度', '经度', '高度(米)', '层数', '高度分档', '用途分类',
    '行政区', '街道办区', '地址', '建筑面积(㎡)',
]

COLS_DROP = [
    '序号', '楼宇名称', '名称来源', '建筑全名(官方)',
    '纬度', '经度', '高度(米)', '层数',
    '剔除类别', '判定依据', '命中关键词', 'OSM标签',
    '行政区', '街道办区', '地址',
]


def write_xlsx_v2(city, cn, keep_rows, drop_rows, landmarks, n_osm, n_ct):
    wb = openpyxl.Workbook()
    HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
    HDR_FILL = PatternFill('solid', fgColor='2F5496')
    CEN = Alignment(horizontal='center', vertical='center')

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CEN
        ws.freeze_panes = 'A2'

    # Sheet1: 商业高层(纯净)
    ws1 = wb.active
    ws1.title = '商业高层(纯净)'
    ws1.append(COLS_KEEP)
    for r in keep_rows:
        ws1.append([r.get(k, '') for k in COLS_KEEP])
    style_header(ws1, len(COLS_KEEP))

    # Sheet2: 已剔除-公共设施
    ws2 = wb.create_sheet('已剔除-公共设施')
    ws2.append(COLS_DROP)
    for r in drop_rows:
        row_data = [
            r['序号'], r.get('楼宇名称', ''), r.get('名称来源', ''),
            r.get('建筑全名(官方)', '—'),
            r.get('纬度', ''), r.get('经度', ''), r.get('高度(米)', ''),
            r.get('层数', ''),
            r.get('_drop_cat', ''), r.get('_drop_basis', ''),
            r.get('_drop_hit', ''),
            '%s | %s' % (r.get('_btype', ''), r.get('_amenity', '')),
            r.get('行政区', ''), r.get('街道办区', ''), r.get('地址', ''),
        ]
        ws2.append(row_data)
    style_header(ws2, len(COLS_DROP))

    # 列宽
    W1 = {'A': 6, 'B': 32, 'C': 13, 'D': 14, 'E': 30, 'F': 12, 'G': 12,
          'H': 9, 'I': 7, 'J': 15, 'K': 11, 'L': 10, 'M': 16, 'N': 38, 'O': 13}
    for k, v in W1.items():
        ws1.column_dimensions[k].width = v
    W2 = {'A': 6, 'B': 30, 'C': 13, 'D': 30, 'E': 12, 'F': 12, 'G': 9,
          'H': 7, 'I': 13, 'J': 18, 'K': 22, 'L': 22, 'M': 10, 'N': 16, 'O': 38}
    for k, v in W2.items():
        ws2.column_dimensions[k].width = v

    # Sheet3: 说明
    ws3 = wb.create_sheet('说明')
    cnt = collections.Counter(r['高度分档'] for r in keep_rows)
    src = collections.Counter(r['名称来源'] for r in keep_rows)
    drop_cnt = collections.Counter(r['_drop_cat'] for r in drop_rows)

    notes = [
        ('%s 高层建筑清单 — 数据说明' % cn, True),
        ('', False),
        ('数据来源：Google Open Buildings V3（建筑足迹）+ Google 2.5D Temporal（高度估算）+ '
         'OpenStreetMap（楼名/用途/地址）+ CTBUH/维基（地标真高）', False),
        ('数据性质：机器学习估算下限（非官方真值）。雅加达官方 LOD2 校准：各档捕获率 53→23%%，越高漏越多。', False),
        ('高层口径：4米/层，≥32米(8层)/≥40米(10层)/≥48米(12层)/≥64米(16层)/≥80米(20层)', False),
        ('', False),
        ('【公共设施剔除规则】已剔除医院/学校/政府/宗教/军事/交通/体育/文化等公共设施。', True),
        ('  判定依据: OSM building 标签 + OSM amenity 标签 + 楼名关键词兜底。', False),
        ('  与雅加达的区别: 雅加达有官方 jns_bgn 细类(更准确), 本城仅用 OSM+名称关键词, 精度次于雅加达。', False),
        ('  居住豁免: OSM 标记为公寓/住宅/宿舍者不剔除。', False),
        ('  全部被剔记录见『已剔除-公共设施』sheet，可逐条复核。', False),
        ('', False),
        ('【统计】', True),
        ('  高层建筑(≥32米)总数: %d 栋（剔除公共设施前）' % (len(keep_rows) + len(drop_rows)), False),
        ('  剔除公共设施: %d 栋  商业高层(本清单): %d 栋' % (len(drop_rows), len(keep_rows)), False),
        ('', False),
        ('  各档数量:  ≥80米 %d  |  ≥64米 %d  |  ≥48米 %d  |  ≥40米 %d  |  ≥32米 %d' % (
            cnt.get('≥80米(约20层)', 0), cnt.get('≥64米(约16层)', 0),
            cnt.get('≥48米(约12层)', 0), cnt.get('≥40米(约10层)', 0),
            cnt.get('≥32米(约8层)', 0)), False),
        ('', False),
        ('  剔除明细: ' + ', '.join('%s %d' % (k, v) for k, v in drop_cnt.most_common()), False),
        ('', False),
        ('【楼名来源】CTBUH地标 %d | OSM补充 %d | 无 %d' % (
            src.get('CTBUH地标', 0), src.get('OSM补充', 0), src.get('无', 0)), False),
        ('  OSM匹配距离 ≤40m 才采纳，距离越小越可信。', False),
        ('', False),
        ('【重要局限】', True),
        ('  1. 高度为 Google ML 估算(约99.5m封顶)，系统偏低。雅加达实测：≥80m 仅捕获23%%。', False),
        ('  2. 每次 ML 估算抽风不同 → 同一栋楼不同跑法可能差 2-5 米；各城之间纵向对比应谨慎。', False),
        ('  3. 大部分楼无楼名/用途(OSM覆盖有限)；坐标为建筑质心点。', False),
        ('  4. 层数默认=高度÷4(估算)，仅少数有 OSM/CTBUH 真实层数。', False),
        ('  5. 公共设施剔除基于 OSM 标签+楼名关键词，不含官方 jns_bgn，精度次于雅加达。', False),
        ('  6. 行政区来自 OSM 行政边界空间匹配(可能有小幅偏移)，匹配率见统计。', False),
        ('  7. 地址来自 OSM addr:* 标签，覆盖率有限。', False),
        ('  8. 本表数量为保守下限，真实更多。', False),
        ('', False),
        ('【地标对照（CTBUH/维基收录）】', True),
    ]
    for i, lm in enumerate(sorted(landmarks, key=lambda x: -(x.get('height_m') or 0)), 1):
        notes.append(('  %d. %s — %.1f米 %s层' % (
            i, str(lm.get('name', '')), float(lm.get('height_m') or 0),
            str(lm.get('floors') or '?')), False))

    # 确保每个 note 是 2 元组
    for idx, note in enumerate(notes):
        if len(note) != 2:
            print('WARNING: notes[%d] 长度=%d: %s' % (idx, len(note), str(note)[:100]), flush=True)
            notes[idx] = (str(note), False)

    for i, (txt, bold) in enumerate(notes, 1):
        c = ws3.cell(i, 1, txt)
        if bold:
            c.font = Font(bold=True, size=11)
    ws3.column_dimensions['A'].width = 120

    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, '%s高层建筑清单.xlsx' % cn)
    wb.save(path)
    print('已写 %s (%d 栋保留, %d 剔除)' % (path, len(keep_rows), len(drop_rows)), flush=True)


if __name__ == '__main__':
    main(sys.argv[1])
