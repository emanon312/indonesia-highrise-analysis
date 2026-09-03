# -*- coding: utf-8 -*-
"""巴淡高层建筑清单 Excel 验收脚本。所有结论以 JSON 输出，避免中文 print 乱码。"""
import json
import openpyxl

PATH = r"analysis\output\巴淡高层建筑清单.xlsx"
# bbox: W,S,E,N
W, S, E, N = 103.98, 1.03, 104.15, 1.18
TOL = 0.05

EXPECTED_SHEETS = ["高层清单", "地标对照", "说明"]
EXPECTED_COLS = ["序号", "楼宇名称", "名称来源", "纬度", "经度", "高度(米)",
                 "高度来源", "层数", "层数来源", "高度分档", "用途分类", "建筑面积", "备注"]

result = {"anomalies": []}
anom = result["anomalies"]

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
sheets = wb.sheetnames

# 1 sheets_ok
sheets_ok = sheets == EXPECTED_SHEETS
if not sheets_ok:
    anom.append("sheet 名称/数量不符，实际为: " + str(sheets))

main = wb[sheets[0]]
rows = list(main.iter_rows(values_only=True))
header = list(rows[0]) if rows else []
data = rows[1:] if len(rows) > 1 else []

# 2 columns_ok
columns_ok = (len(header) == 13) and (list(header) == EXPECTED_COLS)
if not columns_ok:
    anom.append("主sheet列名/列数不符，实际表头: " + str(header))

# 列索引（按预期定位，若不符则尽量按位置）
def idx(name):
    return EXPECTED_COLS.index(name)

i_name = idx("楼宇名称")
i_lat = idx("纬度")
i_lon = idx("经度")
i_h = idx("高度(米)")
i_band = idx("高度分档")

# 3 count
count = len(data)

# 4 tier_sum_ok
from collections import Counter
band_counter = Counter()
for r in data:
    band_counter[r[i_band]] += 1
tier_sum_ok = sum(band_counter.values()) == count

# 辅助：解析高度数值
def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("米", "").replace("m", "").replace("M", "")
    try:
        return float(s)
    except Exception:
        return None

# 期望分档：≥32=8层,≥40=10,≥48=12,≥64=16,≥80=20
def expected_band(h):
    if h is None:
        return None
    if h >= 80:
        return 80
    if h >= 64:
        return 64
    if h >= 48:
        return 48
    if h >= 40:
        return 40
    if h >= 32:
        return 32
    return 0  # <32

# 从分档单元格提取阈值数字
import re
def band_threshold(v):
    if v is None:
        return None
    nums = re.findall(r"\d+", str(v))
    if not nums:
        return None
    # 取最大的那个数(通常分档文本里阈值是其中数字)，优先匹配常见阈值
    cand = [int(x) for x in nums]
    for t in (80, 64, 48, 40, 32):
        if t in cand:
            return t
    return cand[0] if cand else None

# 5 band_consistent + 6 all_height_ge32
inconsistent = 0
inconsistent_examples = []
below32 = 0
below32_examples = []
heights = []
for r in data:
    h = to_float(r[i_h])
    heights.append(h)
    if h is not None and h < 32:
        below32 += 1
        if len(below32_examples) < 5:
            below32_examples.append((r[i_name], h))
    exp = expected_band(h)
    got = band_threshold(r[i_band])
    if exp is not None and got is not None:
        if exp != got:
            inconsistent += 1
            if len(inconsistent_examples) < 8:
                inconsistent_examples.append((r[i_name], h, r[i_band], "应≥%s" % exp))
    elif exp is None or got is None:
        # 无法判定也算疑点
        inconsistent += 1
        if len(inconsistent_examples) < 8:
            inconsistent_examples.append((r[i_name], r[i_h], r[i_band], "无法解析"))

band_consistent = (inconsistent == 0)
all_height_ge32 = (below32 == 0)
if not band_consistent:
    anom.append("高度分档与高度不自洽行数=%d，示例: %s" % (inconsistent, inconsistent_examples))
if not all_height_ge32:
    anom.append("存在高度<32米行数=%d，示例: %s" % (below32, below32_examples))

# 7 coords_in_bbox_pct
in_box = 0
valid_coord = 0
for r in data:
    lat = to_float(r[i_lat])
    lon = to_float(r[i_lon])
    if lat is None or lon is None:
        continue
    valid_coord += 1
    if (S - TOL) <= lat <= (N + TOL) and (W - TOL) <= lon <= (E + TOL):
        in_box += 1
coords_in_bbox_pct = round(100.0 * in_box / count, 1) if count else 0.0
if coords_in_bbox_pct < 80:
    anom.append("坐标落在bbox内占比偏低: %.1f%% (in_box=%d/%d)" % (coords_in_bbox_pct, in_box, count))

# 高度异常大 (>250m 对二三线城市可疑)
big = [(r[i_name], to_float(r[i_h])) for r in data if (to_float(r[i_h]) or 0) > 250]
if big:
    anom.append("高度>250米的可疑行: " + str(big))

# 8 landmark
lm = wb[sheets[1]] if len(sheets) > 1 else None
landmark_rows = 0
landmark_has_ctbuh = False
if lm is not None:
    lm_rows = list(lm.iter_rows(values_only=True))
    landmark_rows = max(0, len(lm_rows) - 1)
    for r in lm_rows:
        for c in r:
            if c and "CTBUH" in str(c).upper():
                landmark_has_ctbuh = True
                break
        if landmark_has_ctbuh:
            break
if landmark_rows == 0:
    anom.append("地标对照sheet为空")
if not landmark_has_ctbuh:
    anom.append("地标对照sheet未含CTBUH来源")

# 9 disclosure_ok
disc = wb[sheets[2]] if len(sheets) > 2 else None
disclosure_ok = False
disc_text = ""
if disc is not None:
    parts = []
    for r in disc.iter_rows(values_only=True):
        for c in r:
            if c is not None:
                parts.append(str(c))
    disc_text = "\n".join(parts)
    for kw in ("下限", "估算", "捕获率"):
        if kw in disc_text:
            disclosure_ok = True
            break
if not disclosure_ok:
    anom.append("说明sheet缺少诚实标注关键词(下限/估算/捕获率)")

# 10 duplicate_rows
seen = {}
dup = 0
for r in data:
    key = (r[i_name], r[i_lat], r[i_lon])
    seen[key] = seen.get(key, 0) + 1
for k, v in seen.items():
    if v > 1:
        dup += (v - 1)
if dup > 0:
    anom.append("楼名+坐标完全重复行数=%d" % dup)

# 11 named_count
named_count = sum(1 for r in data if r[i_name] is not None and str(r[i_name]).strip() != "")

wb.close()

# pass 判定：格式类(1,2,4,5,6)全通过 且 无严重异常
format_ok = sheets_ok and columns_ok and tier_sum_ok and band_consistent and all_height_ge32
serious = (not sheets_ok) or (not columns_ok) or (not tier_sum_ok) or (not band_consistent) or (not all_height_ge32) or (landmark_rows == 0) or (dup > 0)
passed = bool(format_ok and not serious)

result.update({
    "city": "巴淡(batam)",
    "pass": passed,
    "count": count,
    "checks": {
        "sheets_ok": bool(sheets_ok),
        "columns_ok": bool(columns_ok),
        "tier_sum_ok": bool(tier_sum_ok),
        "band_consistent": bool(band_consistent),
        "all_height_ge32": bool(all_height_ge32),
        "coords_in_bbox_pct": coords_in_bbox_pct,
        "landmark_rows": landmark_rows,
        "landmark_has_ctbuh": bool(landmark_has_ctbuh),
        "disclosure_ok": bool(disclosure_ok),
        "duplicate_rows": dup,
        "named_count": named_count,
    },
    "_debug": {
        "sheets": sheets,
        "header": list(header),
        "band_counter": {str(k): v for k, v in band_counter.items()},
        "inconsistent": inconsistent,
        "inconsistent_examples": [list(map(str, x)) for x in inconsistent_examples],
        "below32": below32,
        "in_box": in_box,
        "valid_coord": valid_coord,
        "big_height": [list(map(str, x)) for x in big],
        "height_max": max([h for h in heights if h is not None], default=None),
        "height_min": min([h for h in heights if h is not None], default=None),
    }
})

print(json.dumps(result, ensure_ascii=True))
