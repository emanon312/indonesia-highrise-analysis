# -*- coding: utf-8 -*-
"""通用验收审计脚本 v2 — 适配新格式（3sheet/15列，对齐雅加达）。
用法: python audit_v2.py <city_en>
输出: JSON 格式审计结果
"""
import sys, json, os, re
from collections import Counter
from openpyxl import load_workbook

# 城市 bbox 和中文名
CITY_BBOX = {
    'surabaya':  (112.62, -7.36, 112.84, -7.19),
    'bandung':   (107.55, -6.97, 107.73, -6.86),
    'bekasi':    (106.94, -6.30, 107.06, -6.18),
    'tangerang': (106.58, -6.27, 106.74, -6.12),
    'bogor':     (106.76, -6.64, 106.85, -6.54),
    'medan':     (98.62,  3.52,  98.74,  3.66),
    'semarang':  (110.36, -7.05, 110.49, -6.95),
    'makassar':  (119.39, -5.19, 119.52, -5.07),
    'batam':     (103.98,  1.03, 104.15,  1.18),
    'palembang': (104.68, -3.04, 104.82, -2.93),
    'depok':     (106.76, -6.47, 106.86, -6.35),
}
CITY_CN = {
    'surabaya': '泗水', 'bandung': '万隆', 'bekasi': '勿加泗', 'tangerang': '唐格朗',
    'bogor': '茂物', 'medan': '棉兰', 'semarang': '三宝垄', 'makassar': '望加锡',
    'batam': '巴淡', 'palembang': '巨港', 'depok': '德波',
}

city = sys.argv[1]
cn = CITY_CN[city]
W, S, E, N = CITY_BBOX[city]
PATH = os.path.join(r'analysis\output', '%s高层建筑清单.xlsx' % cn)

# 新格式预期
EXPECTED_SHEETS = ['商业高层(纯净)', '已剔除-公共设施', '说明']
EXPECTED_COLS_KEEP = [
    '序号', '楼宇名称', '名称来源', 'OSM匹配距离(米)', '建筑全名(官方)',
    '纬度', '经度', '高度(米)', '层数', '高度分档', '用途分类',
    '行政区', '街道办区', '地址', '建筑面积(㎡)',
]
EXPECTED_COLS_DROP = [
    '序号', '楼宇名称', '名称来源', '建筑全名(官方)',
    '纬度', '经度', '高度(米)', '层数',
    '剔除类别', '判定依据', '命中关键词', 'OSM标签',
    '行政区', '街道办区', '地址',
]

result = {'anomalies': []}
wb = load_workbook(PATH, read_only=True, data_only=True)
sheets = wb.sheetnames

# 1. sheets_ok
sheets_ok = (sheets[:2] == EXPECTED_SHEETS[:2])  # 前两个必须对
if not sheets_ok:
    result['anomalies'].append('sheet名称不符: 实际=%s 期望=%s' % (sheets[:2], EXPECTED_SHEETS[:2]))

main = wb[sheets[0]]
rows = list(main.iter_rows(values_only=True))
header = list(rows[0]) if rows else []
data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() != '' for c in r)]

header_clean = [str(h).strip() if h is not None else '' for h in header]

# 2. columns_ok
columns_ok = (len(header_clean) == 15) and (header_clean == EXPECTED_COLS_KEEP)
if not columns_ok:
    result['anomalies'].append('列名不符: 实际(%d列)=%s' % (len(header_clean), header_clean))

count = len(data_rows)

def to_float(v):
    try:
        if v is None or str(v).strip() == '':
            return None
        return float(str(v).strip())
    except Exception:
        return None

def cidx(name):
    return EXPECTED_COLS_KEEP.index(name) if name in EXPECTED_COLS_KEEP else None

i_name = cidx('楼宇名称')
i_lat = cidx('纬度')
i_lon = cidx('经度')
i_h = cidx('高度(米)')
i_band = cidx('高度分档')
i_admin = cidx('行政区')
i_kec = cidx('街道办区')
i_addr = cidx('地址')
i_area = cidx('建筑面积(㎡)')
i_osmdist = cidx('OSM匹配距离(米)')

# 3. tier_sum_ok
band_counter = Counter()
for r in data_rows:
    b = r[i_band] if i_band is not None and i_band < len(r) else None
    band_counter[str(b).strip() if b is not None else ''] += 1
tier_sum = sum(band_counter.values())
tier_sum_ok = (tier_sum == count)

# 4. band_consistent
def expected_band(h):
    if h is None: return None
    if h >= 80: return '≥80米(约20层)'
    if h >= 64: return '≥64米(约16层)'
    if h >= 48: return '≥48米(约12层)'
    if h >= 40: return '≥40米(约10层)'
    if h >= 32: return '≥32米(约8层)'
    return '<32米'

inconsistent = 0
inconsistent_examples = []
ge32_fail = 0
height_null = 0
for r in data_rows:
    h = to_float(r[i_h]) if i_h is not None and i_h < len(r) else None
    b = str(r[i_band]).strip() if i_band is not None and i_band < len(r) else ''
    eb = expected_band(h)
    if h is None:
        height_null += 1
        inconsistent += 1
        continue
    if h < 32:
        ge32_fail += 1
    if b != eb:
        inconsistent += 1
        if len(inconsistent_examples) < 8:
            nm = r[i_name] if i_name is not None else ''
            inconsistent_examples.append('%s h=%s 档=%s 期望=%s' % (str(nm)[:20], h, b, eb))

band_consistent = (inconsistent == 0)
all_height_ge32 = (ge32_fail == 0) and (height_null == 0)

# 5. coords_in_bbox_pct
in_box = 0
oob_examples = []
for r in data_rows:
    lat = to_float(r[i_lat]) if i_lat is not None and i_lat < len(r) else None
    lon = to_float(r[i_lon]) if i_lon is not None and i_lon < len(r) else None
    if lat is not None and lon is not None:
        if (S - 0.05) <= lat <= (N + 0.05) and (W - 0.05) <= lon <= (E + 0.05):
            in_box += 1
        else:
            if len(oob_examples) < 6:
                nm = r[i_name] if i_name is not None else ''
                oob_examples.append('%s (%s, %s)' % (str(nm)[:18], lat, lon))
coords_in_bbox_pct = round(100.0 * in_box / count, 1) if count else 0.0

# 6. admin fill rate
admin_filled = sum(1 for r in data_rows if r[i_admin] and str(r[i_admin]).strip()) if i_admin is not None else 0
admin_pct = round(100.0 * admin_filled / count, 1) if count else 0.0

# 7. 已剔除-公共设施 sheet
drop_sheet_ok = False
drop_count = 0
drop_columns_ok = False
if len(sheets) >= 2 and sheets[1] == '已剔除-公共设施':
    ds = wb[sheets[1]]
    drows = list(ds.iter_rows(values_only=True))
    dheader = [str(c).strip() if c is not None else '' for c in drows[0]] if drows else []
    drop_columns_ok = (len(dheader) == 15) and (dheader == EXPECTED_COLS_DROP)
    ddata = [r for r in drows[1:] if any(c is not None and str(c).strip() != '' for c in r)]
    drop_count = len(ddata)
    drop_sheet_ok = (drop_count > 0) and drop_columns_ok
    if not drop_columns_ok:
        result['anomalies'].append('已剔除sheet列名不符: %s' % dheader)

# 8. disclosure_ok
disclosure_ok = False
if '说明' in sheets:
    ds = wb['说明']
    dblob = ' '.join(str(c) for r in ds.iter_rows(values_only=True) for c in r if c is not None)
    disclosure_ok = any(k in dblob for k in ['下限', '估算', 'ML'])
    if not disclosure_ok:
        result['anomalies'].append('说明sheet缺少诚实标注')

# 9. duplicate
seen = {}
dup = 0
for r in data_rows:
    lat = to_float(r[i_lat]) if i_lat is not None else None
    lon = to_float(r[i_lon]) if i_lon is not None else None
    key = (round(lat, 6) if lat else None, round(lon, 6) if lon else None)
    seen[key] = seen.get(key, 0) + 1
for k, v in seen.items():
    if v > 1:
        dup += (v - 1)

# 10. named_count
named_count = sum(1 for r in data_rows if r[i_name] and str(r[i_name]).strip())

# 11. 建筑面积 fill
area_filled = sum(1 for r in data_rows if i_area is not None and i_area < len(r) and r[i_area] is not None and str(r[i_area]).strip() != '')

wb.close()

# 汇总 anomalies
if coords_in_bbox_pct < 90:
    result['anomalies'].append('坐标越界: %.1f%% 例=%s' % (coords_in_bbox_pct, oob_examples))
if inconsistent > 0:
    result['anomalies'].append('分档不自洽 %d行 例=%s' % (inconsistent, inconsistent_examples))
if dup > 0:
    result['anomalies'].append('重复行 %d条' % dup)
if ge32_fail > 0:
    result['anomalies'].append('<32m误入 %d行' % ge32_fail)
if height_null > 0:
    result['anomalies'].append('高度缺失 %d行' % height_null)
if named_count < count * 0.2:
    result['anomalies'].append('命名率低: %d/%d (%.1f%%)' % (named_count, count, 100*named_count/count if count else 0))

severe = (not sheets_ok) or (not columns_ok) or (not tier_sum_ok) or (not band_consistent) or (not all_height_ge32) or (coords_in_bbox_pct < 90)
overall_pass = not severe

result['city'] = '%s(%s)' % (cn, city)
result['pass'] = overall_pass
result['count'] = count
result['checks'] = {
    'sheets_ok': sheets_ok,
    'columns_ok': columns_ok,
    'tier_sum_ok': tier_sum_ok,
    'band_consistent': band_consistent,
    'all_height_ge32': all_height_ge32,
    'coords_in_bbox_pct': coords_in_bbox_pct,
    'admin_fill_pct': admin_pct,
    'drop_sheet_ok': drop_sheet_ok,
    'drop_count': drop_count,
    'disclosure_ok': disclosure_ok,
    'duplicate_rows': dup,
    'named_count': named_count,
    'named_pct': round(100 * named_count / count, 1) if count else 0,
    'area_filled': area_filled,
}
result['_extra'] = {
    'sheets': sheets,
    'header': header_clean,
    'band_counter': dict(band_counter),
    'inconsistent_examples': inconsistent_examples[:5],
    'oob_examples': oob_examples,
}

print(json.dumps(result, ensure_ascii=True, indent=2))
