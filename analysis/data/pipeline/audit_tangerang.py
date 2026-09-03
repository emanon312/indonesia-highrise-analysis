# -*- coding: utf-8 -*-
"""唐格朗高层建筑清单 Excel 验收脚本，结论以 JSON 输出"""
import json
import openpyxl

PATH = r"analysis\output\唐格朗高层建筑清单.xlsx"
# bbox W,S,E,N
W, S, E, N = 106.58, -6.27, 106.74, -6.12
TOL = 0.05

result = {"anomalies": []}
an = result["anomalies"]

wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)
sheets = wb.sheetnames

# 1. sheets_ok
expect_sheets = ["高层清单", "地标对照", "说明"]
sheets_ok = sheets == expect_sheets
result["sheets_ok"] = sheets_ok
if not sheets_ok:
    an.append("sheet名/顺序不符: 实际=%s 期望=%s" % (sheets, expect_sheets))

# 主sheet
ws = wb[sheets[0]]
rows = list(ws.iter_rows(values_only=True))
if not rows:
    rows = [[]]
header = list(rows[0])
data = rows[1:]
# 去掉全空行
data = [r for r in data if any(c is not None and str(c).strip() != "" for c in r)]

# 2. columns_ok
expect_cols = ["序号", "楼宇名称", "名称来源", "纬度", "经度", "高度(米)",
               "高度来源", "层数", "层数来源", "高度分档", "用途分类", "建筑面积", "备注"]
header_clean = [str(h).strip() if h is not None else "" for h in header]
columns_ok = len(header_clean) == 13 and header_clean == expect_cols
result["columns_ok"] = columns_ok
if not columns_ok:
    an.append("列名/列数不符: 实际(%d列)=%s" % (len(header_clean), header_clean))

# 建立列索引
def col_idx(name):
    try:
        return header_clean.index(name)
    except ValueError:
        return None

i_name = col_idx("楼宇名称")
i_lat = col_idx("纬度")
i_lon = col_idx("经度")
i_h = col_idx("高度(米)")
i_band = col_idx("高度分档")

# 3. count
count = len(data)
result["count"] = count

# helper
def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

# 4. tier_sum_ok
band_counts = {}
for r in data:
    b = r[i_band] if i_band is not None and i_band < len(r) else None
    b = str(b).strip() if b is not None else ""
    band_counts[b] = band_counts.get(b, 0) + 1
tier_sum = sum(band_counts.values())
tier_sum_ok = (tier_sum == count)
result["tier_sum_ok"] = tier_sum_ok
if not tier_sum_ok:
    an.append("分档计数和(%d)!=总行数(%d)" % (tier_sum, count))

# 5. band_consistent  口径: h>=80->"≥80", 64<=h<80->"≥64", 48<=h<64->"≥48", 40<=h<48->"≥40", 32<=h<40->"≥32"
def expected_band(h):
    if h is None:
        return None
    if h >= 80:
        return "≥80"
    if h >= 64:
        return "≥64"
    if h >= 48:
        return "≥48"
    if h >= 40:
        return "≥40"
    if h >= 32:
        return "≥32"
    return "<32"

import re
def norm_band(b):
    # 从分档文本中提取阈值数字，归一化为 "≥NN"
    if b is None:
        return ""
    s = str(b)
    m = re.search(r"(\d+)", s)
    if m:
        return "≥" + m.group(1)
    return s.strip()

band_bad = 0
band_samples = []
for r in data:
    h = to_float(r[i_h]) if i_h is not None and i_h < len(r) else None
    b = r[i_band] if i_band is not None and i_band < len(r) else None
    exp = expected_band(h)
    got = norm_band(b)
    if exp is None:
        band_bad += 1
        if len(band_samples) < 5:
            band_samples.append("行无高度: 分档=%s" % b)
        continue
    if got != exp:
        band_bad += 1
        if len(band_samples) < 5:
            band_samples.append("高度%.1f→应%s 实际分档=%s(归一%s)" % (h, exp, b, got))
band_consistent = (band_bad == 0)
result["band_consistent"] = band_consistent
if not band_consistent:
    an.append("分档不自洽行数=%d 示例:%s" % (band_bad, "; ".join(band_samples)))

# 6. all_height_ge32
h_below = 0
h_none = 0
h_max = None
for r in data:
    h = to_float(r[i_h]) if i_h is not None and i_h < len(r) else None
    if h is None:
        h_none += 1
        continue
    if h < 32:
        h_below += 1
    if h_max is None or h > h_max:
        h_max = h
all_height_ge32 = (h_below == 0 and h_none == 0)
result["all_height_ge32"] = all_height_ge32
if h_below > 0:
    an.append("高度<32米行数=%d" % h_below)
if h_none > 0:
    an.append("高度为空行数=%d" % h_none)
if h_max is not None and h_max > 250:
    an.append("出现>250米高度(二三线城市可疑): max=%.1f" % h_max)

# 7. coords_in_bbox_pct
in_box = 0
n_coord = 0
for r in data:
    lat = to_float(r[i_lat]) if i_lat is not None and i_lat < len(r) else None
    lon = to_float(r[i_lon]) if i_lon is not None and i_lon < len(r) else None
    if lat is None or lon is None:
        continue
    n_coord += 1
    if (S - TOL) <= lat <= (N + TOL) and (W - TOL) <= lon <= (E + TOL):
        in_box += 1
pct = round(in_box / count * 100, 1) if count else 0.0
result["coords_in_bbox_pct"] = pct
if pct < 80:
    an.append("坐标在bbox内占比偏低=%.1f%% (有效坐标行%d/%d)" % (pct, n_coord, count))

# 8. 地标对照
ws2 = wb[sheets[1]] if len(sheets) > 1 else None
landmark_rows = 0
landmark_has_ctbuh = False
if ws2 is not None:
    lrows = list(ws2.iter_rows(values_only=True))
    if lrows:
        lhdr = lrows[0]
        ldata = [r for r in lrows[1:] if any(c is not None and str(c).strip() != "" for c in r)]
        landmark_rows = len(ldata)
        # 搜索整sheet是否含CTBUH
        for r in lrows:
            for c in r:
                if c is not None and "CTBUH" in str(c).upper():
                    landmark_has_ctbuh = True
                    break
            if landmark_has_ctbuh:
                break
result["landmark_rows"] = landmark_rows
result["landmark_has_ctbuh"] = landmark_has_ctbuh
if landmark_rows == 0:
    an.append("地标对照sheet为空")
if not landmark_has_ctbuh:
    an.append("地标对照sheet未含CTBUH来源")

# 9. disclosure_ok
ws3 = wb[sheets[2]] if len(sheets) > 2 else None
disclosure_ok = False
discl_text = ""
if ws3 is not None:
    parts = []
    for r in ws3.iter_rows(values_only=True):
        for c in r:
            if c is not None:
                parts.append(str(c))
    discl_text = "\n".join(parts)
    for kw in ["下限", "估算", "捕获率"]:
        if kw in discl_text:
            disclosure_ok = True
            break
result["disclosure_ok"] = disclosure_ok
if not disclosure_ok:
    an.append("说明sheet缺诚实标注关键词(下限/估算/捕获率)")

# 10. duplicate_rows  楼宇名称+纬度+经度
seen = {}
dup = 0
for r in data:
    nm = r[i_name] if i_name is not None and i_name < len(r) else None
    lat = r[i_lat] if i_lat is not None and i_lat < len(r) else None
    lon = r[i_lon] if i_lon is not None and i_lon < len(r) else None
    key = (str(nm).strip() if nm is not None else "",
           str(lat).strip() if lat is not None else "",
           str(lon).strip() if lon is not None else "")
    seen[key] = seen.get(key, 0) + 1
for k, v in seen.items():
    if v > 1:
        dup += (v - 1)
result["duplicate_rows"] = dup
if dup > 0:
    an.append("楼名+坐标完全重复行数=%d" % dup)

# 11. named_count
named = 0
for r in data:
    nm = r[i_name] if i_name is not None and i_name < len(r) else None
    if nm is not None and str(nm).strip() != "":
        named += 1
result["named_count"] = named

# pass: 格式类(1,2,4,5,6)全通过且无严重异常
format_ok = sheets_ok and columns_ok and tier_sum_ok and band_consistent and all_height_ge32
# 严重异常: 高度>250, 坐标越界比例过高(此处坐标越界少量不算fail), 地标空, 缺诚实标注
severe = []
if h_max is not None and h_max > 250:
    severe.append("高度>250")
if landmark_rows == 0:
    severe.append("地标空")
if not disclosure_ok:
    severe.append("缺诚实标注")
overall_pass = format_ok and len(severe) == 0
result["pass"] = overall_pass
result["city"] = "唐格朗"

result["_extra"] = {
    "band_counts": band_counts,
    "h_max": h_max,
    "n_coord": n_coord,
    "in_box": in_box,
    "format_ok": format_ok,
    "severe": severe,
    "sheets": sheets,
}

summary = "通过" if overall_pass else "未通过"
result["summary"] = "唐格朗清单%d行, 格式检查%s, 坐标命中%.1f%%, 地标%d行%s, %s" % (
    count, "全过" if format_ok else "有问题", pct, landmark_rows,
    "含CTBUH" if landmark_has_ctbuh else "无CTBUH", summary)

print(json.dumps(result, ensure_ascii=True))
wb.close()
