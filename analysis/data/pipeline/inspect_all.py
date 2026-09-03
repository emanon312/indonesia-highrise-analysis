# -*- coding: utf-8 -*-
"""扫描 output/ 下全部 12 城 xlsx，输出结构化验收摘要（JSON）。
大文件不进对话：本脚本只打印紧凑摘要供 AI 验收。"""
import json, glob, os
from openpyxl import load_workbook

OUT = r"analysis\output"

def summarize(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    info = {"文件": os.path.basename(path), "sheet数": len(sheets), "sheet名": sheets}
    # 主清单 sheet（第一个）
    ws = wb[sheets[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        info["错误"] = "主sheet为空"
        return info
    header = [str(c) if c is not None else "" for c in rows[0]]
    data = rows[1:]
    info["列数"] = len(header)
    info["列名"] = header
    info["数据行数"] = len(data)
    # 找关键列下标
    def idx(name):
        for i, h in enumerate(header):
            if name in h:
                return i
        return None
    i_band = idx("分档")
    i_name = idx("楼宇名称") or idx("名称") or idx("建筑名称")
    i_nsrc = idx("名称来源")
    i_h = idx("高度")
    # 分档分布
    if i_band is not None:
        dist = {}
        for r in data:
            v = r[i_band] if i_band < len(r) else None
            v = str(v) if v is not None else "(空)"
            dist[v] = dist.get(v, 0) + 1
        info["分档分布"] = dist
    # 命名情况
    if i_name is not None:
        named = sum(1 for r in data if i_name < len(r) and r[i_name] and str(r[i_name]).strip())
        info["有名称行数"] = named
        info["楼名样例"] = [str(r[i_name]) for r in data if i_name < len(r) and r[i_name] and str(r[i_name]).strip()][:5]
    if i_nsrc is not None:
        srcdist = {}
        for r in data:
            v = r[i_nsrc] if i_nsrc < len(r) else None
            v = str(v) if v is not None else "(空)"
            srcdist[v] = srcdist.get(v, 0) + 1
        info["名称来源分布"] = srcdist
    # 高度范围
    if i_h is not None:
        hs = [r[i_h] for r in data if i_h < len(r) and isinstance(r[i_h], (int, float))]
        if hs:
            info["高度范围"] = [round(min(hs), 1), round(max(hs), 1)]
    # 其余 sheet 行数
    other = {}
    for s in sheets[1:]:
        wss = wb[s]
        n = sum(1 for _ in wss.iter_rows(values_only=True))
        other[s] = max(0, n - 1)  # 减表头
    info["其余sheet数据行"] = other
    wb.close()
    return info

results = []
for p in sorted(glob.glob(os.path.join(OUT, "*.xlsx"))):
    try:
        results.append(summarize(p))
    except Exception as e:
        results.append({"文件": os.path.basename(p), "错误": repr(e)})

print(json.dumps(results, ensure_ascii=False, indent=1))
