# -*- coding: utf-8 -*-
"""给异名近距簇打「近距簇」Tag（疑似建筑群/相邻楼）。<30m连通分量聚类，多点簇标簇ID，独立楼留空。
新增列「近距簇」(末列)。同时输出两口径统计。dry-run默认，--run写回。
注意：近距簇含真综合体+无关相邻楼，是"疑似"标记，供统计筛选/人工判断，不做合并。"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict

HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
CODE={'雅加达':'JKT','泗水':'SBY','万隆':'BDG','唐格朗':'TNG','望加锡':'MKS','棉兰':'MDN','巴淡':'BTM','三宝垄':'SMG','勿加泗':'BKS','德波':'DPK','茂物':'BGR','巨港':'PLM'}
R=30.0; SHEET='商业高层(纯净)'; COL='近距簇'
def dist(a,b):return math.hypot((a[0]-b[0])*111000,(a[1]-b[1])*111000*math.cos(math.radians(a[0])))

def process(city, write):
    path=os.path.join(OUT,f'{city}高层建筑清单.xlsx')
    wb=load_workbook(path); ws=wb[SHEET]
    header=[str(c.value).strip() if c.value else '' for c in ws[1]]
    li=header.index('纬度');oi=header.index('经度')
    rows=[[c.value for c in row] for row in ws.iter_rows(min_row=2) if any(c.value is not None for c in row)]
    pts=[(float(r[li]),float(r[oi])) for r in rows]
    # <30m 并查集连通分量
    parent=list(range(len(pts)))
    def find(x):
        while parent[x]!=x: parent[x]=parent[parent[x]];x=parent[x]
        return x
    cell=0.0004;grid=defaultdict(list)
    for i,p in enumerate(pts):grid[(int(p[0]/cell),int(p[1]/cell))].append(i)
    for i,p in enumerate(pts):
        ck=(int(p[0]/cell),int(p[1]/cell))
        for dx in(-1,0,1):
            for dy in(-1,0,1):
                for j in grid.get((ck[0]+dx,ck[1]+dy),[]):
                    if j>i and dist(p,pts[j])<R:
                        ra,rb=find(i),find(j)
                        if ra!=rb:parent[ra]=rb
    comp=defaultdict(list)
    for i in range(len(pts)):comp[find(i)].append(i)
    multi=[c for c in comp.values() if len(c)>1]
    # 簇ID分配(按簇内最小行号排序)
    multi_sorted=sorted(multi,key=lambda c:min(c))
    tag_of={}
    for k,c in enumerate(multi_sorted,1):
        cid='%s-%03d'%(CODE[city],k)
        for i in c: tag_of[i]=cid
    n_single=len(pts); n_grouped=len(comp)  # 每栋单算 / 每簇算1
    if write:
        import shutil
        bak=os.path.join(OUT,'..','backup_加Tag前'); os.makedirs(bak,exist_ok=True)
        shutil.copy2(path,os.path.join(bak,f'{city}高层建筑清单.xlsx'))
        if COL in header:
            newcol=header.index(COL)+1  # 复用已有列，重算时不再新增（避免重复列）
        else:
            newcol=len(header)+1
            hc=ws.cell(row=1,column=newcol,value=COL)
            # 套用表头样式(参照第1列表头)
            t=ws.cell(row=1,column=1)
            hc.font=Font(name=t.font.name,bold=True,color=t.font.color.rgb if t.font.color else None,size=t.font.size)
            if t.fill and t.fill.patternType: hc.fill=PatternFill('solid',fgColor=t.fill.fgColor.rgb)
        for ridx in range(len(rows)):
            ws.cell(row=ridx+2,column=newcol,value=tag_of.get(ridx,''))
        ws.column_dimensions[ws.cell(row=1,column=newcol).column_letter].width=10
        wb.save(path)
    wb.close()
    return n_single, n_grouped, len(multi_sorted)

def main():
    write='--run' in sys.argv
    print('='*56);print('  近距簇Tag(<30m,异名疑似建筑群) - '+('执行' if write else 'dry-run'));print('='*56)
    ts=tg=tc=0
    for c in CITIES:
        s,g,nc=process(c,write);ts+=s;tg+=g;tc+=nc
        print('  %-5s 单栋%-5d 合算%-5d 近距簇%d个'%(c,s,g,nc))
    print('-'*56)
    print('  【每栋单算】12城 %d 栋'%ts)
    print('  【建筑群合算】12城 %d 个(每近距簇算1，独立楼各算1)'%tg)
    print('  近距簇共 %d 个'%tc)

main()
