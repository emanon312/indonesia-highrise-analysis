# -*- coding: utf-8 -*-
"""从12城Excel生成汇总Excel：Sheet1 总表，每城一行，含分档累计+数据来源"""
import os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
CITIES = ["雅加达","泗水","万隆","唐格朗","望加锡","棉兰","巴淡","三宝垄","勿加泗","德波","茂物","巨港"]

def count_bands(xlsx_path):
    """读第一个sheet，统计高度分档累计"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    bi = next((i for i, h in enumerate(header) if "分档" in h), None)
    if bi is None:
        wb.close()
        return {}
    buckets = {32:0, 40:0, 48:0, 64:0, 80:0}
    for r in rows[1:]:
        if r[bi] is None:
            continue
        m = re.search(r"(80|64|48|40|32)", str(r[bi]))
        if m:
            thr = int(m.group(1))
            if thr in buckets:
                buckets[thr] += 1
    wb.close()
    cum = {t: sum(v for k,v in buckets.items() if k>=t) for t in (32,40,48,64,80)}
    return cum

def count_calibers(xlsx_path):
    """读第一个sheet，返回(总行数, 独立楼数, 近距簇个数)。
    近距簇列有值(如'JKT-085')表示该行属于某个近距簇；同值行合为一栋建筑群。
    每栋单算 = 总行数；建筑群合算 = 独立楼数 + 近距簇个数。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return 0, 0, 0
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    ci = next((i for i, h in enumerate(header) if "近距簇" in h), None)
    total = 0
    indep = 0
    clusters = set()
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        total += 1
        val = r[ci] if ci is not None and ci < len(r) else None
        if val is None or str(val).strip() == "":
            indep += 1
        else:
            clusters.add(str(val).strip())
    wb.close()
    return total, indep, len(clusters)

def main():
    wb = Workbook()

    # ---- 样式 ----
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name="微软雅黑", bold=True, size=14)
    data_font = Font(name="微软雅黑", size=11)
    jakarta_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 绿色标记官方数据
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # ==================== Sheet1: 汇总总表 ====================
    ws = wb.active
    ws.title = "12城汇总"

    # 标题行
    ws.merge_cells("A1:G1")
    ws["A1"] = "印尼12城高层建筑汇总"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = f"口径：4米/层。≥32m=8层 | ≥40m=10层 | ≥48m=12层 | ≥64m=16层 | ≥80m=20层。雅加达为官方真值，其余为Google ML估算下限。"
    ws["A2"].font = Font(name="微软雅黑", size=9, color="666666")
    ws["A2"].alignment = left_align

    ws.merge_cells("A3:G3")
    ws["A3"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(name="微软雅黑", size=9, color="999999")
    ws["A3"].alignment = left_align

    # 表头 (row 5)
    headers = ["城市", "≥32m\n(8层)", "≥40m\n(10层)", "≥48m\n(12层)", "≥64m\n(16层)", "≥80m\n(20层)", "数据来源"]
    col_widths = [12, 12, 12, 12, 12, 12, 24]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 数据行
    rows_data = []
    for city in CITIES:
        xlsx = os.path.join(OUT, f"{city}高层建筑清单.xlsx")
        cum = count_bands(xlsx) if os.path.exists(xlsx) else {}
        src = "官方真值" if city == "雅加达" else "Google ML估算(下限)"
        rows_data.append({"name": city, **cum, "src": src})

    rows_data.sort(key=lambda x: x.get(32, 0), reverse=True)
    totals = {32:0, 40:0, 48:0, 64:0, 80:0}

    for ri, r in enumerate(rows_data, 6):
        values = [r["name"], r.get(32,0), r.get(40,0), r.get(48,0), r.get(64,0), r.get(80,0), r["src"]]
        is_jakarta = r["name"] == "雅加达"
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.alignment = center if ci > 1 else left_align
            cell.border = thin_border
            if is_jakarta:
                cell.fill = jakarta_fill
        for k in totals:
            totals[k] += r.get(k, 0)

    # 合计行
    total_row = 6 + len(rows_data)
    total_values = ["合计", totals[32], totals[40], totals[48], totals[64], totals[80], "—"]
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    for ci, v in enumerate(total_values, 1):
        cell = ws.cell(row=total_row, column=ci, value=v)
        cell.font = Font(name="微软雅黑", bold=True, size=11)
        cell.fill = total_fill
        cell.alignment = center if ci > 1 else left_align
        cell.border = thin_border

    # ==================== Sheet2: 分城明细 ====================
    ws2 = wb.create_sheet("分城明细")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "各城分档明细（逐档独立计数，非累计）"
    ws2["A1"].font = Font(name="微软雅黑", bold=True, size=12)
    ws2["A1"].alignment = left_align

    detail_headers = ["城市", "32m档\n(8-9层)", "40m档\n(10-11层)", "48m档\n(12-15层)", "64m档\n(16-19层)", "80m档\n(20层+)", "数据来源"]
    for ci, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]

    for ri, r in enumerate(rows_data, 4):
        # 逐档独立计数 = cum[t] - cum[next_higher]
        c32 = r.get(32,0)
        c40 = r.get(40,0)
        c48 = r.get(48,0)
        c64 = r.get(64,0)
        c80 = r.get(80,0)
        b32 = c32 - c40  # only 32-39m
        b40 = c40 - c48  # only 40-47m
        b48 = c48 - c64  # only 48-63m
        b64 = c64 - c80  # only 64-79m
        values = [r["name"], b32, b40, b48, b64, c80, r["src"]]
        is_jakarta = r["name"] == "雅加达"
        for ci, v in enumerate(values, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.alignment = center if ci > 1 else left_align
            cell.border = thin_border
            if is_jakarta:
                cell.fill = jakarta_fill

    # 合计
    t32 = totals[32] - totals[40]
    t40 = totals[40] - totals[48]
    t48 = totals[48] - totals[64]
    t64 = totals[64] - totals[80]
    dt_values = ["合计", t32, t40, t48, t64, totals[80], "—"]
    for ci, v in enumerate(dt_values, 1):
        cell = ws2.cell(row=4+len(rows_data), column=ci, value=v)
        cell.font = Font(name="微软雅黑", bold=True, size=11)
        cell.fill = total_fill
        cell.alignment = center if ci > 1 else left_align
        cell.border = thin_border

    # ==================== Sheet3: 说明 ====================
    ws3 = wb.create_sheet("说明")
    jkt32 = next((r[32] for r in rows_data if r['name'] == '雅加达'), 0)

    # 两口径统计（动态读近距簇列计算）
    cal_total = 0      # 每栋单算 = 全部主表行数之和
    cal_indep = 0      # 独立楼数
    cal_clusters = 0   # 近距簇个数
    for city in CITIES:
        xlsx = os.path.join(OUT, f"{city}高层建筑清单.xlsx")
        if os.path.exists(xlsx):
            t, ind, nc = count_calibers(xlsx)
            cal_total += t
            cal_indep += ind
            cal_clusters += nc
    cal_group = cal_indep + cal_clusters  # 建筑群合算

    notes = [
        ["口径说明", "4米/层折算。≥32m≈8层，≥40m≈10层，≥48m≈12层，≥64m≈16层，≥80m≈20层。"],
        ["两种计数口径", f"①每栋单算：主表每行算1栋，12城合计 {cal_total} 栋（本汇总表各档统计均按此口径）。"],
        ["", f"②建筑群合算：把'近距簇'内多栋楼视为1个建筑群 = 独立楼 {cal_indep} 栋 + 近距簇 {cal_clusters} 个 = {cal_group} 栋。"],
        ["", "近距簇=同址相邻的多栋楼（主表末列'近距簇'同值即同一群）；两口径差值即簇内被合并的重复栋数。"],
        ["数据来源", "雅加达：DPMPTSP官方建筑登记数据（2026-06-22抓取），经公共设施剔除、医院并入商业、多轮去重。"],
        ["", "其余11城：Google ML建筑足迹高度估算（Open Buildings 2024），取bbox内≥32m建筑。"],
        ["", "ML估算高度为下限值——实际高度只高不低；官方面积越大标注覆盖越全。"],
        ["去重处理", f"雅加达：官方登记多轮去重（精确坐标+高度、同名近距、25米空间去重），去除同一栋楼的重复登记，最终保留 {jkt32} 栋。"],
        ["", "跨城去重：唐格朗删8栋、勿加泗删3栋确认重复建筑。"],
        ["公共设施剔除", "雅加达三层剔除：官方商业细类→楼名关键词（政府/学校/宗教等）→fungsi兜底。医院不剔除，并入商业高层。"],
        ["", f"本表为剔除公共设施后的商业高层（含医院），12城合计 {totals[32]} 栋。"],
        ["汇总表说明", "Sheet1\"12城汇总\"为累计阈值统计（≥32m包含所有≥32m的建筑）。"],
        ["", "Sheet2\"分城明细\"为逐档独立计数（32m档仅含32-39m，不重复计算）。"],
    ]
    for ri, (label, note) in enumerate(notes, 1):
        ws3.cell(row=ri, column=1, value=label).font = Font(name="微软雅黑", bold=True, size=10)
        ws3.cell(row=ri, column=2, value=note).font = Font(name="微软雅黑", size=10)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 80

    # 保存
    out_path = os.path.join(OUT, "12城高层建筑汇总.xlsx")
    wb.save(out_path)
    print(f"已生成: {out_path}")
    print(f"  Sheet1: 12城汇总 (累计阈值)")
    print(f"  Sheet2: 分城明细 (逐档独立)")
    print(f"  Sheet3: 说明")
    print(f"\n合计: ≥32m={totals[32]} ≥40m={totals[40]} ≥48m={totals[48]} ≥64m={totals[64]} ≥80m={totals[80]}")
    print(f"两口径: 每栋单算={cal_total} | 建筑群合算={cal_group} (独立楼{cal_indep}+近距簇{cal_clusters})")

if __name__ == "__main__":
    main()
