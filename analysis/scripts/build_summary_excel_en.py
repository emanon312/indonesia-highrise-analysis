# -*- coding: utf-8 -*-
"""英文版全国总结 Excel：从12城中文清单读实际数据算，输出 output/en/Indonesia_HighRise_Summary.xlsx。
参照 build_summary_excel.py 结构，3个sheet：National Summary / By-Tier Detail / Notes。
数字全部读实际数据动态计算，不写死。"""
import os, re, sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)  # 便于 import 同目录的 i18n_map
from i18n_map import CITY_EN

OUT = os.path.join(HERE, "..", "output")
OUT_EN = os.path.join(OUT, "en")
CITIES = ["雅加达", "泗水", "万隆", "唐格朗", "望加锡", "棉兰", "巴淡", "三宝垄", "勿加泗", "德波", "茂物", "巨港"]


def count_bands(xlsx_path):
    """读主sheet，统计高度分档累计。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    bi = next((i for i, h in enumerate(header) if "分档" in h), None)
    if bi is None:
        wb.close()
        return {}
    buckets = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
    for r in rows[1:]:
        if r[bi] is None:
            continue
        m = re.search(r"(80|64|48|40|32)", str(r[bi]))
        if m:
            thr = int(m.group(1))
            if thr in buckets:
                buckets[thr] += 1
    wb.close()
    cum = {t: sum(v for k, v in buckets.items() if k >= t) for t in (32, 40, 48, 64, 80)}
    return cum


def count_calibers(xlsx_path):
    """读主sheet，返回(总行数, 独立楼数, 近距簇个数)。
    每栋单算=总行数；建筑群合算=独立楼数+近距簇个数。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return 0, 0, 0
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    ci = next((i for i, h in enumerate(header) if "近距簇" in h), None)
    total = 0
    indep = 0
    clusters = set()
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        total += 1
        val = r[ci] if ci is not None and ci < len(r) else None
        if val is None or str(val).strip() == "":
            indep += 1
        else:
            clusters.add(str(val).strip())
    wb.close()
    return total, indep, len(clusters)


def main():
    wb = Workbook()

    # ---- 样式（英文用 Calibri）----
    FONT = "Calibri"
    header_font = Font(name=FONT, bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(name=FONT, bold=True, size=14)
    data_font = Font(name=FONT, size=11)
    jakarta_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 浅绿=官方
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    SRC_JKT = "Official 3D DB (DPMPTSP)"
    SRC_OTHER = "Google ML estimate (lower bound)"

    # ---- 读全部城市数据 ----
    rows_data = []
    for city in CITIES:
        xlsx = os.path.join(OUT, f"{city}高层建筑清单.xlsx")
        cum = count_bands(xlsx) if os.path.exists(xlsx) else {}
        src = SRC_JKT if city == "雅加达" else SRC_OTHER
        rows_data.append({"name": city, "en": CITY_EN[city], **cum, "src": src})
    rows_data.sort(key=lambda x: x.get(32, 0), reverse=True)

    totals = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
    for r in rows_data:
        for k in totals:
            totals[k] += r.get(k, 0)

    # ==================== Sheet1: National Summary (累计阈值) ====================
    ws = wb.active
    ws.title = "National Summary"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Indonesia 12-City High-Rise Summary"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = ("Cumulative counts (a >=32m column includes every building >=32m). "
                "~4m/floor: >=32m=8F | >=40m=10F | >=48m=12F | >=64m=16F | >=80m=20F. "
                "Jakarta = official ground truth; other cities = Google ML estimate (conservative lower bound).")
    ws["A2"].font = Font(name=FONT, size=9, color="666666")
    ws["A2"].alignment = left_align

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(name=FONT, size=9, color="999999")
    ws["A3"].alignment = left_align

    headers = ["City", "≥32m (8F)", "≥40m (10F)", "≥48m (12F)", "≥64m (16F)", "≥80m (20F)", "Data Source"]
    col_widths = [14, 12, 12, 12, 12, 12, 30]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, r in enumerate(rows_data, 6):
        values = [r["en"], r.get(32, 0), r.get(40, 0), r.get(48, 0), r.get(64, 0), r.get(80, 0), r["src"]]
        is_jakarta = r["name"] == "雅加达"
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.alignment = center if ci > 1 else left_align
            cell.border = thin_border
            if is_jakarta:
                cell.fill = jakarta_fill

    total_row = 6 + len(rows_data)
    total_values = ["Total", totals[32], totals[40], totals[48], totals[64], totals[80], "—"]
    for ci, v in enumerate(total_values, 1):
        cell = ws.cell(row=total_row, column=ci, value=v)
        cell.font = Font(name=FONT, bold=True, size=11)
        cell.fill = total_fill
        cell.alignment = center if ci > 1 else left_align
        cell.border = thin_border

    # ==================== Sheet2: By-Tier Detail (逐档独立) ====================
    ws2 = wb.create_sheet("By-Tier Detail")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Per-Tier Detail (independent counts, non-cumulative)"
    ws2["A1"].font = Font(name=FONT, bold=True, size=12)
    ws2["A1"].alignment = left_align

    detail_headers = ["City", "32m band", "40m band", "48m band", "64m band", "80m+ band", "Data Source"]
    for ci, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]

    for ri, r in enumerate(rows_data, 4):
        c32, c40, c48, c64, c80 = (r.get(t, 0) for t in (32, 40, 48, 64, 80))
        # 逐档独立 = 相邻累计相减
        values = [r["en"], c32 - c40, c40 - c48, c48 - c64, c64 - c80, c80, r["src"]]
        is_jakarta = r["name"] == "雅加达"
        for ci, v in enumerate(values, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.alignment = center if ci > 1 else left_align
            cell.border = thin_border
            if is_jakarta:
                cell.fill = jakarta_fill

    dt_values = ["Total", totals[32] - totals[40], totals[40] - totals[48],
                 totals[48] - totals[64], totals[64] - totals[80], totals[80], "—"]
    for ci, v in enumerate(dt_values, 1):
        cell = ws2.cell(row=4 + len(rows_data), column=ci, value=v)
        cell.font = Font(name=FONT, bold=True, size=11)
        cell.fill = total_fill
        cell.alignment = center if ci > 1 else left_align
        cell.border = thin_border

    # ==================== Sheet3: Notes ====================
    ws3 = wb.create_sheet("Notes")
    jkt32 = next((r.get(32, 0) for r in rows_data if r["name"] == "雅加达"), 0)

    # 两口径统计（动态读近距簇列）
    cal_total = cal_indep = cal_clusters = 0
    for city in CITIES:
        xlsx = os.path.join(OUT, f"{city}高层建筑清单.xlsx")
        if os.path.exists(xlsx):
            t, ind, nc = count_calibers(xlsx)
            cal_total += t
            cal_indep += ind
            cal_clusters += nc
    cal_group = cal_indep + cal_clusters

    notes = [
        ["Scope", "High-rise buildings >=32m (approx. 8 floors and above) across 12 major Indonesian cities."],
        ["Height tiers", "Converted at ~4m per floor. >=32m~=8F, >=40m~=10F, >=48m~=12F, >=64m~=16F, >=80m~=20F."],
        ["Counting caliber 1", f"Per-building: each row on the main sheet is one building. 12-city total = {cal_total} buildings (all counts in this summary use this caliber)."],
        ["Counting caliber 2", f"Per building-group: buildings inside one proximity cluster count as a single complex = {cal_indep} standalone + {cal_clusters} clusters = {cal_group} groups."],
        ["", "A proximity cluster = several adjacent buildings at the same site (same value in the 'Cluster' column of each city list). The gap between the two calibers equals the duplicated towers merged inside clusters."],
        ["Data source", f"Jakarta ({jkt32} buildings): DPMPTSP official 3D building database (Jakarta Satu, LOD2 modeling + government validation, scraped 2026-06-22). Ground truth for height, coordinates and floor area."],
        ["", "Other 11 cities: Google Open Buildings V3 footprints + 2.5D height estimation (machine learning). Estimated heights are a conservative lower bound - actual heights are only higher, never lower. A few landmark towers use CTBUH measured heights."],
        ["Deduplication", f"Jakarta: multi-pass dedup on official registry (exact coordinate+height, same-name proximity, 25m spatial dedup) removes repeated records of the same tower, keeping {jkt32}. Cross-city: confirmed duplicates removed in Tangerang and Bekasi."],
        ["Public facilities", f"Government/school/religious and similar public buildings were filtered out; hospitals are retained and merged into the commercial high-rise set. This summary reflects the cleaned set (hospitals included), 12-city total {totals[32]}."],
        ["Sheets", "'National Summary' = cumulative thresholds. 'By-Tier Detail' = independent per-tier counts (the 32m band holds only 32-39m, no double counting)."],
        ["Disclaimer", "Jakarta figures are official deduplicated ground truth; the other 11 cities are Google ML conservative lower bounds - real counts are only higher. All 12 cities share the same spatial dedup caliber to avoid over-counting a single tower split into multiple points."],
    ]
    for ri, (label, note) in enumerate(notes, 1):
        ws3.cell(row=ri, column=1, value=label).font = Font(name=FONT, bold=True, size=10)
        c = ws3.cell(row=ri, column=2, value=note)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 110

    # ---- 保存 ----
    os.makedirs(OUT_EN, exist_ok=True)
    out_path = os.path.join(OUT_EN, "Indonesia_HighRise_Summary.xlsx")
    wb.save(out_path)
    print(f"Generated: {out_path}")
    print("  Sheet1: National Summary (cumulative)")
    print("  Sheet2: By-Tier Detail (independent)")
    print("  Sheet3: Notes")
    print(f"Total: >=32m={totals[32]} >=40m={totals[40]} >=48m={totals[48]} >=64m={totals[64]} >=80m={totals[80]}")
    print(f"Calibers: per-building={cal_total} | per-group={cal_group} (standalone {cal_indep} + clusters {cal_clusters})")


if __name__ == "__main__":
    main()
