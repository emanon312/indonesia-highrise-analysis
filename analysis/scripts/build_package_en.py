# -*- coding: utf-8 -*-
"""英文交付包：读实际数据算关键数字→动态README_EN.md→组织英文目录→zip。
用法：python build_package_en.py
依赖：需先跑过 build_excel_en.py 和 build_summary_excel_en.py（output/en/ 有产物）。
统计数字读中文源清单 output/*.xlsx（表头已知、数值一致），仅复制英文产物入包。"""
import os, shutil, zipfile
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = os.path.join(HERE, "..")
OUT = os.path.join(BASE, "output")
OUT_EN = os.path.join(OUT, "en")
CITIES = ["雅加达", "泗水", "万隆", "唐格朗", "望加锡", "棉兰", "巴淡", "三宝垄", "勿加泗", "德波", "茂物", "巨港"]
CITY_EN_FILE = {
    "雅加达": "Jakarta", "泗水": "Surabaya", "万隆": "Bandung", "唐格朗": "Tangerang",
    "望加锡": "Makassar", "棉兰": "Medan", "巴淡": "Batam", "三宝垄": "Semarang",
    "勿加泗": "Bekasi", "德波": "Depok", "茂物": "Bogor", "巨港": "Palembang",
}


def stats():
    """读中文源清单主sheet算：总栋数、雅加达栋数、雅加达医院数、≥80m数、
    两口径(每栋单算 / 建筑群合算=独立楼+近距簇，簇个数按城内去重)。"""
    total = jkt = ge80 = hosp = 0
    indep = clusters = 0
    for c in CITIES:
        wb = load_workbook(os.path.join(OUT, f"{c}高层建筑清单.xlsx"), read_only=True, data_only=True)
        ws = wb["商业高层(纯净)"]
        h = [str(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        hi = h.index("高度(米)")
        ui = h.index("用途分类")
        ci = h.index("近距簇")
        n = 0
        seen = set()  # 本城近距簇编号去重
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            n += 1
            total += 1
            hh = float(r[hi]) if r[hi] else 0
            if hh >= 80:
                ge80 += 1
            if c == "雅加达" and str(r[ui]) == "医院":
                hosp += 1
            val = r[ci] if ci < len(r) else None
            if val is None or str(val).strip() == "":
                indep += 1
            else:
                seen.add(str(val).strip())
        clusters += len(seen)
        if c == "雅加达":
            jkt = n
        wb.close()
    group = indep + clusters  # 建筑群合算
    return {
        "total": total, "jkt": jkt, "ge80": ge80, "hosp": hosp,
        "group": group, "indep": indep, "clusters": clusters,
    }


def readme(s):
    other = s["total"] - s["jkt"]
    return f"""# Indonesia 12-City High-Rise Buildings - Delivery Guide

## 1. Overview

This package contains the high-rise building analysis (**>=32m, approx. 8 floors and above**) for **12 major Indonesian cities**, totaling **{s['total']} buildings** (per-building caliber).

## 2. Data source & reliability

| Scope | Source | Nature | Buildings |
|-------|--------|--------|-----------|
| Jakarta | Indonesia DPMPTSP official 3D building database (Jakarta Satu) | **Official ground truth** (LOD2 modeling + government validation) | {s['jkt']} |
| Other 11 cities | Google Open Buildings V3 footprints + 2.5D height estimation | Machine-learning estimate (conservative lower bound) | {other} |

- **Jakarta is the official ground-truth benchmark** - height, coordinates and floor area come from government registry.
- **The other 11 cities are ML estimates**; heights are systematically low (a conservative lower bound), so actual heights are only higher. A few landmark towers use CTBUH measured heights (see the "Data Source" column in each Excel).
- DPMPTSP = Dinas Penanaman Modal dan Pelayanan Terpadu Satu Pintu (Investment and One-Stop Integrated Services Agency), DKI Jakarta Provincial Government. Platform: Jakarta Satu. Data nature: LOD2 3D modeling + government validation (Validasi), scraped 2026-06-22.

## 3. Height tiers

Five tiers by building height (~4m per floor): >=32m (8F) - >=40m (10F) - >=48m (12F) - >=64m (16F) - >=80m (20F).

## 4. Counting calibers

- **Per-building**: each row is one building = **{s['total']}** across 12 cities (used by all summary counts).
- **Per building-group**: buildings inside one proximity cluster count as a single complex = {s['indep']} standalone + {s['clusters']} clusters = **{s['group']}**.
- A proximity cluster = several adjacent buildings at the same site (same value in the "Cluster" column). The gap between the two calibers equals the duplicated towers merged inside clusters.

## 5. Deduplication (important)

**A single tower is often counted multiple times**; spatial deduplication was applied to avoid over-counting:
- **Jakarta**: the official registry holds multiple tax/registration records per tower -> exact dedup + same-name proximity + 25m spatial dedup.
- **Other 11 cities**: Google ML may split one tower into several footprint points -> 25m spatial dedup (within 25m radius and <25% height difference merged; adjacent buildings of different heights kept).
- All 12 cities share the same dedup caliber. Hospitals are retained and merged into the commercial high-rise set (not filtered as public facilities).

## 6. File navigation

### Building_Lists_12Cities/
12 Excel files, one per city. Each has 3 sheets: main commercial high-rise list (core data, hospitals included) / removed public facilities / notes.
17 columns per building: building name, name source, data source, coordinates, height, floors, category, district, address, floor area, cluster, etc.
Building names, addresses and sub-districts keep the original Indonesian text (needed for on-site survey and navigation).

### Summary/
- **Indonesia_HighRise_Summary.xlsx** - tier statistics (cumulative thresholds + per-tier detail + notes).

### Report/
- **Indonesia_HighRise_Report.html** - fully English interactive report (overview + map + searchable list + export). Open by double-click; forward directly to external parties.

## 7. Usage notes

- The HTML report **needs internet** to load the map basemap on first open; **building data is embedded (a fixed snapshot) and is never altered by the network**.
- The map library is inlined, so recipients can just double-click to use it.
- Building names and addresses keep the original Indonesian text.

## 8. Key figures

| Metric | Value |
|--------|-------|
| 12-city total (>=32m, per-building) | **{s['total']}** |
| Per building-group caliber | **{s['group']}** ({s['indep']} standalone + {s['clusters']} clusters) |
| >=80m super-tall | {s['ge80']} |
| Jakarta (official ground truth) | {s['jkt']} (incl. {s['hosp']} hospitals) |
| Data snapshot date | 2026-06-22 (Jakarta official scrape) |

---

*Jakarta figures are official deduplicated ground truth; the other 11 cities are Google conservative lower bounds - real counts are only higher. All 12 cities have spatial dedup applied to avoid over-counting a single tower split into multiple points.*
"""


def main():
    s = stats()
    PKG = os.path.join(BASE, "Indonesia_HighRise_EN_Package")
    if os.path.exists(PKG):
        shutil.rmtree(PKG)
    d_lists = os.path.join(PKG, "Building_Lists_12Cities")
    d_sum = os.path.join(PKG, "Summary")
    d_report = os.path.join(PKG, "Report")
    for d in (d_lists, d_sum, d_report):
        os.makedirs(d)

    with open(os.path.join(PKG, "README_EN.md"), "w", encoding="utf-8") as f:
        f.write(readme(s))

    # 12城英文清单
    for c in CITIES:
        src = os.path.join(OUT_EN, f"{CITY_EN_FILE[c]}_HighRise_Buildings.xlsx")
        shutil.copy(src, d_lists)
    # 英文总结
    shutil.copy(os.path.join(OUT_EN, "Indonesia_HighRise_Summary.xlsx"), d_sum)
    # 英文报告
    shutil.copy(os.path.join(OUT, "Indonesia_HighRise_Report.html"), d_report)

    # zip 文件名用英文，转发给外方时整包无中文
    zipname = os.path.join(BASE, "Indonesia_HighRise_EN_Package.zip")
    if os.path.exists(zipname):
        os.remove(zipname)
    file_count = 0
    with zipfile.ZipFile(zipname, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(PKG):
            for f in files:
                fp = os.path.join(root, f)
                z.write(fp, os.path.relpath(fp, BASE))
                file_count += 1
    shutil.rmtree(PKG)

    print(f"EN package generated: {zipname}")
    print(f"  size: {os.path.getsize(zipname) / 1024 / 1024:.1f} MB | files: {file_count}")
    print(f"  README figures: total {s['total']} | Jakarta {s['jkt']}(hosp {s['hosp']}) | "
          f">=80m {s['ge80']} | calibers {s['total']}/{s['group']}")


if __name__ == "__main__":
    main()
