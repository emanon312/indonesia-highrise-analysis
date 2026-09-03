# -*- coding: utf-8 -*-
"""读现成的 12 城中文 Excel 清单，逐 sheet 翻译成全英文另存到 output/en/。

只读不改：不重跑任何数据管线，不改动任何现有文件，只新建英文 xlsx。
翻译资产复用 i18n_map.py；官方标签正式化两处映射写在本脚本内常量。
说明(Notes)页完全重写成友好导读式英文，所有数字实时读该城主 sheet 计算。
"""
import os
from collections import Counter, OrderedDict

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import i18n_map as M

# ── 路径约定 ───────────────────────────────────────────────
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # analysis/
CN_DIR = os.path.join(BASE, 'output')
EN_DIR = os.path.join(BASE, 'output', 'en')

# 中文城市名列表（对应 output/{城市}高层建筑清单.xlsx）
CITIES_CN = ['雅加达', '泗水', '万隆', '唐格朗', '望加锡', '棉兰',
             '巴淡', '三宝垄', '勿加泗', '德波', '茂物', '巨港']

# ── sheet 名翻译 ──────────────────────────────────────────
SHEET_NAME_EN = {
    '商业高层(纯净)': 'Commercial High-Rise (Clean)',
    '已剔除-公共设施': 'Excluded — Public Facilities',
    '说明': 'Notes',
}

# ── 官方标签正式化（仅英文 Excel，用户要求；不改 i18n_map）──
# 数据来源列：雅加达官方 → 数据平台正式名（比 i18n_map 的 HTML 简称更正式）
OFFICIAL_DATA_SRC = {'官方(DPMPTSP)': 'Jakarta Satu 3D Building Database (DPMPTSP)'}
# 名称来源列：官方 → 正式注册来源（覆盖 i18n_map 的 'Official'）
OFFICIAL_NAME_SRC = {'官方': 'Official registry (DPMPTSP)'}

# ── 已剔除 sheet 独有列表头（不在 COLS_EN 中，本脚本补充）──
EXCL_COLS_EN = {
    '剔除类别': 'Excluded Category',
    '判定依据': 'Filter Basis',
    '命中关键词': 'Matched Keyword',
    'OSM标签': 'OSM Tag',
    '官方细类(jns_bgn)': 'Official Subtype (jns_bgn)',
}

# 已剔除 sheet 分类值翻译（剔除类别列）
EXCL_CAT_EN = {
    '交通/停车': 'Transport/Parking', '体育设施': 'Sports Facility', '公共': 'Public',
    '军事/国防': 'Military/Defense', '学校': 'School', '宗教场所': 'Religious Site',
    '政府/公共服务': 'Government/Public Service', '文化设施': 'Cultural Facility',
    '高校': 'University/College',
}

# 已剔除 sheet 判定依据值翻译（判定依据列）
EXCL_BASIS_EN = {
    'OSM amenity标签': 'OSM amenity tag', 'OSM building标签': 'OSM building tag',
    'OSM office标签': 'OSM office tag', 'OSM religion标签': 'OSM religion tag',
    'fungsi兜底': 'fungsi fallback', '官方细类': 'Official subtype',
    '楼名关键词': 'Building-name keyword',
    '楼名关键词(OSM补名)': 'Building-name keyword (OSM-supplied name)',
    '坐标连带剔除(楼名关键词)': 'Removed by proximity (building-name keyword)',
    '坐标连带剔除(楼名关键词(OSM补名))': 'Removed by proximity (OSM-supplied name keyword)',
}

# ── 表头翻译（中文表头 → 英文），合并 COLS_EN 与已剔除专属列 ──
HEADER_EN = {**M.COLS_EN, **EXCL_COLS_EN}

# ── 逐列值翻译规则：中文表头 → 值翻译字典 ─────────────────
# 数据来源/名称来源先套官方正式标签，找不到再套 i18n_map 技术表述。
VALUE_MAPS = {
    '用途分类': M.CATEGORY_EN,
    '数据来源': {**M.DATA_SRC_EN, **OFFICIAL_DATA_SRC},
    '名称来源': {**M.NAME_SRC_EN, **OFFICIAL_NAME_SRC},
    '高度分档': M.TIER_EN,
    '行政区': M.DISTRICT_EN,           # 命中则译（仅雅加达 5 区），否则原样（11 城印尼语保留）
    '剔除类别': EXCL_CAT_EN,
    '判定依据': EXCL_BASIS_EN,
}


# 所有枚举英文值合集（用于兜底：清理源数据错位导致的中文枚举泄漏，如个别行 层数 列
# 被填入剔除类别文本。楼名/地址/街道办区为印尼语，不在此集合，不受影响）。
ALL_ENUM_EN = {}
for _d in (M.CATEGORY_EN, {**M.DATA_SRC_EN, **OFFICIAL_DATA_SRC},
           {**M.NAME_SRC_EN, **OFFICIAL_NAME_SRC}, M.TIER_EN, M.DISTRICT_EN,
           EXCL_CAT_EN, EXCL_BASIS_EN):
    ALL_ENUM_EN.update(_d)


def _tr_value(header_cn, value):
    """按列翻译单个值；找不到映射则原样返回（不丢数据）。楼名/地址/坐标等无映射列直接原样。"""
    if value is None:
        return None
    vm = VALUE_MAPS.get(header_cn)
    if vm is not None:
        return vm.get(value, value)
    # 无专属列映射：仅当值恰为已知中文枚举（源数据错位泄漏）时兜底翻译，其余原样
    if isinstance(value, str) and value in ALL_ENUM_EN:
        return ALL_ENUM_EN[value]
    return value


def _col_widths(headers_en, rows):
    """按表头与数据内容估算列宽，上限 60。"""
    widths = []
    for ci, h in enumerate(headers_en):
        w = len(str(h))
        for r in rows:
            v = r[ci]
            if v is not None:
                w = max(w, len(str(v)))
        widths.append(min(max(w + 2, 8), 60))
    return widths


def translate_data_sheet(ws_src, ws_dst):
    """翻译主 sheet / 已剔除 sheet：表头翻译 + 逐列值翻译，保留加粗表头与列宽。"""
    all_rows = list(ws_src.iter_rows(values_only=True))
    headers_cn = list(all_rows[0])
    data_rows = all_rows[1:]

    headers_en = [HEADER_EN.get(h, h) for h in headers_cn]
    ws_dst.append(headers_en)

    for r in data_rows:
        ws_dst.append([_tr_value(headers_cn[ci], v) for ci, v in enumerate(r)])

    # 表头加粗
    for cell in ws_dst[1]:
        cell.font = Font(bold=True)
    # 列宽
    widths = _col_widths(headers_en, data_rows)
    for ci, w in enumerate(widths, start=1):
        ws_dst.column_dimensions[get_column_letter(ci)].width = w


# ── 主 sheet 统计（说明页实时数字全部来自这里）──────────────
def compute_stats(ws_main):
    rows = list(ws_main.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {h: i for i, h in enumerate(hdr)}
    data = rows[1:]

    total = len(data)

    # 用途分类 → 医院数
    hospital = sum(1 for r in data if r[idx['用途分类']] == '医院')

    # 高度分档：各档计数 + 累计（≥某高度的总数）
    tier_order = ['≥32米(约8层)', '≥40米(约10层)', '≥48米(约12层)',
                  '≥64米(约16层)', '≥80米(约20层)']
    tier_cnt = Counter(r[idx['高度分档']] for r in data)
    # 累计：从最高档往下累加
    tiers = []
    cum = 0
    for t in reversed(tier_order):          # ≥80 → ≥32
        cum += tier_cnt.get(t, 0)
        tiers.append((M.TIER_EN.get(t, t), tier_cnt.get(t, 0), cum))
    tiers.reverse()                          # 展示时 ≥32 → ≥80

    # 名称来源分布
    name_src = Counter(r[idx['名称来源']] for r in data)
    name_src_en = OrderedDict()
    for k, v in name_src.most_common():
        label = OFFICIAL_NAME_SRC.get(k) or M.NAME_SRC_EN.get(k, k)
        name_src_en[label] = v

    # 近距簇两口径
    clusters = set()
    independent = 0
    for r in data:
        c = r[idx['近距簇']]
        if c is None or str(c).strip() in ('', '—'):
            independent += 1
        else:
            clusters.add(str(c).strip())
    grouped = independent + len(clusters)

    return {
        'total': total, 'hospital': hospital, 'tiers': tiers,
        'name_src': name_src_en, 'independent': independent,
        'n_clusters': len(clusters), 'grouped': grouped,
    }


def _write_notes(ws, city_cn, city_en, stats, is_jakarta, excluded_count):
    """完全重写说明页为友好导读式英文，数字来自 stats。"""
    lines = []  # (text, style)  style ∈ {'title','h','p','b'(bullet)}

    def H(t): lines.append((t, 'h'))
    def P(t): lines.append((t, 'p'))
    def B(t): lines.append((t, 'b'))
    def BL(): lines.append(('', 'p'))

    s = stats
    lines.append((f'{city_en} — High-Rise Buildings (≥32 m) — Data Guide', 'title'))
    BL()

    # 1) What this is —— 大白话导读
    H('What this is')
    if is_jakarta:
        src_line = ("The data comes from the Jakarta provincial government's official 3D building "
                    "database (Jakarta Satu / DPMPTSP), captured on 2026-06-22, so the heights are "
                    "government-validated true values.")
    else:
        src_line = ("The data comes from Google Open Buildings V3 (building footprints plus "
                    "machine-learning height estimates); a few landmarks use CTBUH published heights, "
                    "and building names are matched from OpenStreetMap. Heights are conservative "
                    "estimates — a lower bound, so real buildings are only taller, never shorter.")
    P(f"This spreadsheet lists the high-rise buildings in {city_en}, Indonesia — everything "
      f"about 8 floors and up (32 m or taller). The main sheet, "
      f"“Commercial High-Rise (Clean)”, holds {s['total']} buildings; a second sheet, "
      f"“Excluded — Public Facilities”, lists {excluded_count} tall structures we "
      f"filtered out (government offices, schools, places of worship, transport hubs and the like). "
      f"{src_line}")
    P("You can use it as a quick inventory for market screening, site scouting or urban research. "
      "When reading the table: each row is one building shown as a single representative point (not a "
      "street-door-accurate location); building names, addresses and sub-district names are kept in "
      "the original Indonesian; and coordinates/heights/floors/areas are raw numeric values.")
    BL()

    # 2) Data source & reliability
    H('Data source & reliability')
    if is_jakarta:
        P("Source: Jakarta Satu 3D Building Database, published by DPMPTSP. This is an official "
          "LOD2 three-dimensional model with government validation (Validasi), i.e. verified true "
          "values, not estimates. Snapshot date: 2026-06-22.")
        P("Agency full name: DPMPTSP = Dinas Penanaman Modal dan Pelayanan Terpadu Satu Pintu "
          "(Investment and One-Stop Integrated Services Agency), DKI Jakarta Provincial Government. "
          "Data platform: Jakarta Satu. Data nature: LOD2 3D modelling + government validation.")
    else:
        P("Source: Google Open Buildings V3 — building footprints plus a 2.5D machine-learning "
          "height estimate. These heights are a conservative lower bound: actual heights are only "
          "higher. Well-known landmarks instead use CTBUH published true heights. Building names are "
          "matched from OpenStreetMap where available.")
    BL()

    # 3) Height tiers
    H('Height tiers')
    P("Buildings are grouped into five height bands (roughly 4 m per floor): "
      "≥32 m (~8F), ≥40 m (~10F), ≥48 m (~12F), ≥64 m (~16F), ≥80 m (~20F). "
      "Each building sits in the single highest band it reaches.")
    P("Count in this city (band count / cumulative at-or-above):")
    for label, cnt, cum in s['tiers']:
        B(f"{label}: {cnt} in this band  |  {cum} at or above this height")
    BL()

    # 4) Building counts (two calibers)
    H('Building counts (two calibers)')
    P(f"Per-building count: {s['total']}. This equals the number of rows in the main sheet — "
      f"every individual building tower counts as one.")
    P(f"Grouped (building-complex) count: {s['grouped']}. Here, several towers that sit very close "
      f"together (a “near cluster”, shown in the Cluster column) are counted as one "
      f"integrated complex. This city has {s['independent']} standalone buildings plus "
      f"{s['n_clusters']} clusters, giving {s['independent']} + {s['n_clusters']} = {s['grouped']}.")
    BL()

    # 5) How duplicates were handled
    H('How duplicates were handled')
    if is_jakarta:
        P("The official registry can hold several records for the same physical tower (multiple "
          "permit registrations). We de-duplicated in several passes: exact coordinate + height "
          "match, then same-name near matches, then a 25 m spatial merge.")
    else:
        P("De-duplication was already applied when the list was generated, using a 15 m spatial "
          "merge to collapse overlapping footprints of the same building.")
    BL()

    # 6) Public-facility filtering
    H('Public-facility filtering')
    P(f"We keep commercial high-rises — offices, apartments, hotels, malls — plus "
      f"hospitals ({s['hospital']} hospitals kept in this city). We remove public facilities: "
      f"government, schools/universities, religious sites, military, transport and similar. "
      f"All {excluded_count} removed structures are listed, with the reason, in the "
      f"“Excluded — Public Facilities” sheet.")
    BL()

    # 7) Name vs Data source columns
    H('Name source vs Data source columns')
    P("These two columns answer different questions. Name Source = where the building's name came "
      "from (official registry, OpenStreetMap, CTBUH landmark, etc.). Data Source = where the "
      "building's own data — coordinates, height and so on — came from.")
    P("Name-source breakdown in this city:")
    for label, cnt in s['name_src'].items():
        B(f"{label}: {cnt}")
    BL()

    # 8) Limitations
    H('Limitations')
    B("The point shown is a representative location, not a precise street address / door number.")
    B("Some buildings lack a floor count; for those the height band is used as a proxy for floors.")
    B("Names added from OpenStreetMap can be uncertain; such cases are judged by match distance "
      "(see the OSM Match Dist. column and the “OSM (uncertain)” name source).")
    if not is_jakarta:
        B("Heights for this city are ML estimates and represent a lower bound — treat them as "
          "“at least this tall”.")
    BL()

    # 9) Disclaimer
    H('Disclaimer')
    P("For preliminary screening only; verify on-site before formal decisions.")

    # ── 写入并排版 ──
    ws.column_dimensions['A'].width = 118
    title_font = Font(bold=True, size=14)
    h_font = Font(bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical='top')
    for text, style in lines:
        ws.append([text])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.alignment = wrap
        if style == 'title':
            cell.font = title_font
        elif style == 'h':
            cell.font = h_font
        elif style == 'b':
            cell.value = '  • ' + text


def build_city(city_cn):
    src_path = os.path.join(CN_DIR, f'{city_cn}高层建筑清单.xlsx')
    city_en = M.CITY_EN[city_cn]
    dst_path = os.path.join(EN_DIR, f'{city_en}_HighRise_Buildings.xlsx')
    is_jakarta = (city_cn == '雅加达')

    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    wb_dst = openpyxl.Workbook()
    wb_dst.remove(wb_dst.active)

    main_ws_src = wb_src['商业高层(纯净)']
    excl_ws_src = wb_src['已剔除-公共设施']

    stats = compute_stats(main_ws_src)
    excluded_count = excl_ws_src.max_row - 1  # 减表头

    # 主 sheet
    ws = wb_dst.create_sheet(SHEET_NAME_EN['商业高层(纯净)'])
    translate_data_sheet(main_ws_src, ws)
    # 已剔除 sheet
    ws2 = wb_dst.create_sheet(SHEET_NAME_EN['已剔除-公共设施'])
    translate_data_sheet(excl_ws_src, ws2)
    # 说明 sheet（完全重写）
    ws3 = wb_dst.create_sheet(SHEET_NAME_EN['说明'])
    _write_notes(ws3, city_cn, city_en, stats, is_jakarta, excluded_count)

    wb_src.close()
    wb_dst.save(dst_path)
    return dst_path, stats['total']


def main():
    os.makedirs(EN_DIR, exist_ok=True)
    print(f'输出目录: {EN_DIR}')
    total_all = 0
    for city_cn in CITIES_CN:
        path, n = build_city(city_cn)
        total_all += n
        print(f'  {M.CITY_EN[city_cn]:12s}  行数(主sheet)={n:5d}  -> {path}')
    print(f'合计每栋单算(主sheet行数): {total_all}')


if __name__ == '__main__':
    main()
