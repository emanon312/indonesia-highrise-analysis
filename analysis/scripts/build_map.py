# -*- coding: utf-8 -*-
"""12城高层建筑地理分布地图 —— 读各城Excel坐标，生成单文件Leaflet交互地图 + 附带CSV。
点按高度分档5色着色，聚类，弹窗显示详情。用法：python build_map.py"""
import os, sys, json, csv
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

HERE=os.path.dirname(os.path.abspath(__file__))
OUT=os.path.join(HERE,'..','output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
SHEET='商业高层(纯净)'

# 高度分档 → 颜色（浅底图可见，点带白描边）
def band_color(h):
    if h>=80: return '#d73027','≥80米'   # 红
    if h>=64: return '#fc8d59','≥64米'   # 橙
    if h>=48: return '#f4a300','≥48米'   # 金
    if h>=40: return '#66bd63','≥40米'   # 绿
    return '#4575b4','≥32米'             # 蓝

def col(header, name):
    return next((i for i,h in enumerate(header) if h and name in str(h)), None)

def read_all():
    pts=[]
    stats={}
    for city in CITIES:
        path=os.path.join(OUT,f'{city}高层建筑清单.xlsx')
        wb=load_workbook(path, read_only=True, data_only=True)
        ws=wb[SHEET]
        rows=list(ws.iter_rows(values_only=True))
        h=[str(c).strip() if c is not None else '' for c in rows[0]]
        ci={k:col(h,k) for k in ['楼宇名称','纬度','经度','高度(米)','层数','用途分类','数据来源','地址','近距簇']}
        n=0
        for r in rows[1:]:
            if all(v is None for v in r): continue
            lat=r[ci['纬度']]; lon=r[ci['经度']]
            if lat is None or lon is None: continue
            hh=float(r[ci['高度(米)']]) if r[ci['高度(米)']] is not None else 0
            def g(k):
                i=ci[k]; return r[i] if i is not None and r[i] is not None else ''
            pts.append({
                'c':city,'n':str(g('楼宇名称')),'lat':round(float(lat),7),'lon':round(float(lon),7),
                'h':round(hh,1),'fl':g('层数'),'u':str(g('用途分类')),'s':str(g('数据来源')),'a':str(g('地址')),
                'cl':str(g('近距簇')),
            })
            n+=1
        stats[city]=n
        wb.close()
        print(f'  {city}: {n} 点')
    return pts, stats

def build_html(pts, stats):
    total=len(pts)
    # 分档统计
    bands={'≥80米':0,'≥64米':0,'≥48米':0,'≥40米':0,'≥32米':0}
    for p in pts:
        _,lbl=band_color(p['h']); bands[lbl]+=1
    # 近距簇统计：cl有值=簇内点；不同cl值=不同簇
    clustered=sum(1 for p in pts if p['cl'])
    n_clusters=len({p['cl'] for p in pts if p['cl']})
    data_json=json.dumps(pts, ensure_ascii=False)
    city_json=json.dumps(CITIES, ensure_ascii=False)
    legend_rows=''.join(
        f'<div><i style="background:{band_color(80 if l=="≥80米" else 64 if l=="≥64米" else 48 if l=="≥48米" else 40 if l=="≥40米" else 32)[0]}"></i>{l} ({bands[l]})</div>'
        for l in ['≥80米','≥64米','≥48米','≥40米','≥32米'])
    html=f'''<!DOCTYPE html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>印尼12城高层建筑分布图</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<style>
  html,body{{margin:0;height:100%}}#map{{height:100%}}
  .info{{position:absolute;z-index:1000;background:rgba(255,255,255,.94);padding:10px 12px;
    border-radius:6px;box-shadow:0 1px 6px rgba(0,0,0,.3);font:13px/1.5 "Microsoft YaHei",sans-serif}}
  #legend{{right:12px;top:12px}}#title{{left:52px;top:12px;font-weight:bold}}
  #legend i{{display:inline-block;width:12px;height:12px;margin-right:6px;border:1px solid #fff;
    border-radius:50%;box-shadow:0 0 2px #666;vertical-align:middle}}
  #legend b{{display:block;margin-bottom:4px}}
  .pop b{{font-size:14px}}.pop table{{border-collapse:collapse;margin-top:4px}}
  .pop td{{padding:1px 6px 1px 0;font-size:12px}}.pop td:first-child{{color:#888}}
  .cltip{{background:#111;color:#ffd400;border:none;font:bold 11px "Microsoft YaHei",sans-serif;
    padding:1px 5px;border-radius:3px;box-shadow:0 1px 3px rgba(0,0,0,.4)}}
  .cltip::before{{border-top-color:#111}}
</style></head><body>
<div id="map"></div>
<div id="title" class="info">印尼12城高层建筑（≥32米）· 共 {total} 栋</div>
<div id="legend" class="info"><b>高度分档（点击图层控制筛城市）</b>{legend_rows}
<div style="margin-top:6px;padding-top:6px;border-top:1px solid #ddd">
<i style="background:#ffd400;border:1.5px solid #111!important;box-shadow:none!important"></i>建筑群黄圈+建筑群名（放大浮现）{clustered} 点 / {n_clusters} 群</div></div>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<script>
var PTS={data_json}, CITIES={city_json};
var map=L.map('map').setView([-2.5,118],5);
L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',
  {{maxZoom:19,attribution:'&copy; OpenStreetMap'}}).addTo(map);
function color(h){{return h>=80?'#d73027':h>=64?'#fc8d59':h>=48?'#f4a300':h>=40?'#66bd63':'#4575b4';}}
function esc(s){{return String(s).replace(/[&<>]/g,function(c){{return{{'&':'&amp;','<':'&lt;','>':'&gt;'}}[c];}});}}
// 每城一个聚类图层，便于图层控制显隐
var layers={{}}, overlays={{}};
CITIES.forEach(function(c){{
  layers[c]=L.markerClusterGroup({{maxClusterRadius:45,chunkedLoading:true}});
}});
var bounds=[];
PTS.forEach(function(p){{
  var cl=p.cl&&p.cl.length;
  var m=L.circleMarker([p.lat,p.lon],cl?
    {{radius:6,color:'#111',weight:2,fillColor:color(p.h),fillOpacity:.9}}:
    {{radius:5,color:'#fff',weight:1,fillColor:color(p.h),fillOpacity:.85}});
  var pop='<div class="pop"><b>'+esc(p.n||'(无名)')+'</b><table>'+
    '<tr><td>城市</td><td>'+esc(p.c)+'</td></tr>'+
    '<tr><td>高度</td><td>'+p.h+' 米</td></tr>'+
    '<tr><td>层数</td><td>'+esc(p.fl)+'</td></tr>'+
    '<tr><td>用途</td><td>'+esc(p.u)+'</td></tr>'+
    '<tr><td>数据来源</td><td>'+esc(p.s)+'</td></tr>'+
    (cl?'<tr><td>近距簇</td><td><b>'+esc(p.cl)+'</b>（疑似建筑群）</td></tr>':'')+
    '<tr><td>地址</td><td>'+esc(p.a)+'</td></tr></table></div>';
  m.bindPopup(pop);
  layers[p.c].addLayer(m);
  bounds.push([p.lat,p.lon]);
}});
CITIES.forEach(function(c){{ layers[c].addTo(map); overlays[c+' ('+layers[c].getLayers().length+')']=layers[c]; }});
L.control.layers(null,overlays,{{collapsed:true,position:'bottomright'}}).addTo(map);
if(bounds.length) map.fitBounds(bounds,{{padding:[30,30]}});

// ===== 建筑群包围圈 + 簇ID标签 =====
// 每个近距簇算外接圆（成员均值为心，到最远成员距离+padding为半径），画黄圈+永久ID标签。
// 放大到城市级(zoom>=11)自动浮现，缩小自动隐藏（否则全国视图上百个标签会糊成一团）。
var groups={{}};
PTS.forEach(function(p){{ if(p.cl&&p.cl.length){{ (groups[p.cl]=groups[p.cl]||[]).push(p); }} }});
var clusterLayer=L.layerGroup();
Object.keys(groups).forEach(function(cid){{
  var ms=groups[cid], lat=0, lon=0;
  ms.forEach(function(p){{ lat+=p.lat; lon+=p.lon; }}); lat/=ms.length; lon/=ms.length;
  var r=0; ms.forEach(function(p){{ var d=map.distance([lat,lon],[p.lat,p.lon]); if(d>r) r=d; }});
  r=Math.max(r+22, 35);  // 至少35m，框得住
  // 群名：取最高成员楼名；若最高成员无名，退化为"城市-簇序号"
  var m0=ms[0]; ms.forEach(function(p){{ if(p.h>m0.h) m0=p; }});
  var noname={{'':1,'(无名)':1,'（无名）':1,'None':1}};
  var gname=noname[m0.n]?(m0.c+'-'+cid.split('-')[1]):m0.n;
  var circ=L.circle([lat,lon],{{radius:r,color:'#111',weight:1.5,opacity:.75,
    fillColor:'#ffd400',fillOpacity:.15}});
  circ.bindTooltip(gname+'（'+ms.length+'栋）',{{permanent:true,direction:'top',className:'cltip'}});
  circ.bindPopup('<b>'+gname+'</b> 疑似建筑群<br>'+ms.length+' 栋：'+
    ms.map(function(p){{return esc(p.n||'(无名)')+' '+p.h+'m';}}).join('<br>'));
  clusterLayer.addLayer(circ);
}});
function toggleClusters(){{
  if(map.getZoom()>=11){{ if(!map.hasLayer(clusterLayer)) clusterLayer.addTo(map); }}
  else {{ if(map.hasLayer(clusterLayer)) map.removeLayer(clusterLayer); }}
}}
map.on('zoomend',toggleClusters); toggleClusters();
</script></body></html>'''
    p=os.path.join(OUT,'建筑分布地图.html')
    open(p,'w',encoding='utf-8').write(html)
    print(f'\n已生成: {p}')
    print(f'  分档: '+' | '.join(f'{k}:{v}' for k,v in bands.items()))
    print(f'  近距簇: {clustered} 点 / {n_clusters} 群（深色粗描边标记）')
    return bands

def build_csv(pts):
    p=os.path.join(OUT,'建筑坐标.csv')
    with open(p,'w',encoding='utf-8-sig',newline='') as f:
        w=csv.writer(f)
        w.writerow(['城市','楼宇名称','纬度','经度','高度(米)','高度分档','用途分类','数据来源','近距簇','地址'])
        for x in pts:
            _,lbl=band_color(x['h'])
            w.writerow([x['c'],x['n'],x['lat'],x['lon'],x['h'],lbl,x['u'],x['s'],x['cl'],x['a']])
    print(f'已生成: {p} ({len(pts)} 行)')

def main():
    print('读取12城坐标...')
    pts,stats=read_all()
    print(f'\n合计 {len(pts)} 点')
    build_html(pts,stats)
    build_csv(pts)

main()
