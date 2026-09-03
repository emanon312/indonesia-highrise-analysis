# -*- coding: utf-8 -*-
"""英文一体化交付物生成器 —— 读12城Excel，翻译成英文，生成单文件HTML。
三视图：Overview(含城市天际线剖面) / Map(Leaflet) / Directory(可搜索清单)。
用法：python build_deliverable.py"""
import os, sys, json, html
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import load_workbook
from i18n_map import (CITY_EN, NAME_SRC_EN, DATA_SRC_EN, TIER_EN, DISTRICT_EN, CATEGORY_EN)

HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
# 高度分档色阶（贯穿地图/天际线/表格）
def tier_color(h):
    if h>=80: return '#b3202e'
    if h>=64: return '#e0662c'
    if h>=48: return '#d9a520'
    if h>=40: return '#4e9e5a'
    return '#3a6ea5'
def tier_key(h):
    return '≥80m' if h>=80 else '≥64m' if h>=64 else '≥48m' if h>=48 else '≥40m' if h>=40 else '≥32m'

def col(header,name): return next((i for i,x in enumerate(header) if x and name in str(x)),None)

def read_data():
    pts=[]; cities=[]
    for city in CITIES:
        wb=load_workbook(os.path.join(OUT,f'{city}高层建筑清单.xlsx'),read_only=True,data_only=True)
        ws=wb['商业高层(纯净)']
        rows=list(ws.iter_rows(values_only=True))
        h=[str(c).strip() if c is not None else '' for c in rows[0]]
        ix={k:col(h,k) for k in ['楼宇名称','名称来源','数据来源','纬度','经度','高度(米)','层数','用途分类','行政区','街道办区','地址','近距簇']}
        heights=[]
        cen=CITY_EN[city]
        for r in rows[1:]:
            if all(v is None for v in r): continue
            lat=r[ix['纬度']]; lon=r[ix['经度']]
            if lat is None or lon is None: continue
            hh=round(float(r[ix['高度(米)']]),1) if r[ix['高度(米)']] is not None else 0
            def g(k):
                i=ix[k]; return r[i] if i is not None and r[i] is not None else ''
            district=str(g('行政区'))
            nm=str(g('楼宇名称'))
            if nm in ('（无名）','(无名)','','None'): nm='(unnamed)'
            pts.append({
                'c':cen,
                'n':nm,
                'cat':CATEGORY_EN.get(str(g('用途分类')),str(g('用途分类'))),
                's':DATA_SRC_EN.get(str(g('数据来源')),str(g('数据来源'))),
                'ns':NAME_SRC_EN.get(str(g('名称来源')),str(g('名称来源'))),
                'h':hh,'fl':g('层数'),
                'd':DISTRICT_EN.get(district,district),        # 中文行政区译英，印尼语保留
                'sd':str(g('街道办区')),                        # 印尼语原文
                'a':str(g('地址')),                            # 印尼语原文
                'cl':str(g('近距簇')),                          # 近距簇标签，空串=独立楼
                'lat':round(float(lat),6),'lon':round(float(lon),6),
            })
            heights.append(hh)
        wb.close()
        heights.sort(reverse=True)
        cities.append({'name':cen,'count':len(heights),'max':heights[0] if heights else 0,'heights':heights})
    return pts, cities

def skyline_svg(heights, W=150, H=54, maxH=340):
    """城市天际线剖面：建筑按高度降序竖条，色阶着色。采样至≤300栋保持轮廓。"""
    n=len(heights)
    if n==0: return ''
    step=max(1,n//300)
    sampled=heights[::step]
    m=len(sampled)
    rects=[]
    for i,hh in enumerate(sampled):
        bh=max(1.0, hh/maxH*H)
        rects.append(f'<rect x="{i}" y="{H-bh:.1f}" width="1" height="{bh:.1f}" fill="{tier_color(hh)}"/>')
    return (f'<svg viewBox="0 0 {m} {H}" width="{W}" height="{H}" preserveAspectRatio="none" '
            f'class="sky">{"".join(rects)}</svg>')

def build():
    pts, cities = read_data()
    total=len(pts)
    # 两口径：单栋=每栋单算；grouped=近距簇按1个建筑群折算
    cluster_members=sum(1 for p in pts if p['cl'])
    n_clusters=len({p['cl'] for p in pts if p['cl']})
    grouped=total-(cluster_members-n_clusters)
    caliber=f'{total:,} individual buildings / {grouped:,} grouped ({n_clusters} clusters)'
    # 统计
    tiers={'≥80m':0,'≥64m':0,'≥48m':0,'≥40m':0,'≥32m':0}
    sources={}
    for p in pts:
        tiers[tier_key(p['h'])]+=1
        sources[p['s']]=sources.get(p['s'],0)+1
    cities_sorted=sorted(cities,key=lambda x:-x['count'])
    # 天际线卡片HTML
    maxH=max(c['max'] for c in cities)
    sky_cards=[]
    for c in cities_sorted:
        svg=skyline_svg(c['heights'],maxH=maxH)
        sky_cards.append(
            f'<div class="skycard"><div class="skywrap">{svg}</div>'
            f'<div class="skymeta"><span class="skyname">{html.escape(c["name"])}</span>'
            f'<span class="skynum">{c["count"]}</span></div>'
            f'<div class="skymax">max {c["max"]:.0f}m</div></div>')
    # tier图例数据 & 分布条
    tier_order=['≥80m','≥64m','≥48m','≥40m','≥32m']
    tier_colors={'≥80m':'#b3202e','≥64m':'#e0662c','≥48m':'#d9a520','≥40m':'#4e9e5a','≥32m':'#3a6ea5'}
    tier_labels={'≥80m':'≥80m (~20F)','≥64m':'≥64m (~16F)','≥48m':'≥48m (~12F)','≥40m':'≥40m (~10F)','≥32m':'≥32m (~8F)'}
    tier_bar=''.join(f'<div class="tseg" style="flex:{tiers[t]};background:{tier_colors[t]}" title="{tier_labels[t]}: {tiers[t]}"></div>' for t in tier_order)
    tier_legend=''.join(f'<div class="lgi"><i style="background:{tier_colors[t]}"></i>{tier_labels[t]}<b>{tiers[t]}</b></div>' for t in tier_order)
    # 数据来源分布
    src_items=''.join(f'<li><span>{html.escape(k)}</span><b>{v}</b></li>' for k,v in sorted(sources.items(),key=lambda x:-x[1]))

    data_json=json.dumps(pts,ensure_ascii=False,separators=(',',':'))
    kpi_high=tiers['≥80m']
    official=sources.get('Official (DPMPTSP)',0)

    tmpl=open(os.path.join(HERE,'deliverable_template.html'),encoding='utf-8').read()
    # 内联地图库（离线自包含，只剩底图瓦片需联网）
    vdir=os.path.join(HERE,'vendor')
    vcss='\n'.join(open(os.path.join(vdir,f),encoding='utf-8').read() for f in ['leaflet.css','MarkerCluster.css','MarkerCluster.Default.css'])
    vjs=open(os.path.join(vdir,'leaflet.js'),encoding='utf-8').read()+'\n'+open(os.path.join(vdir,'markercluster.js'),encoding='utf-8').read()
    out=(tmpl
        .replace('__VENDOR_CSS__',vcss)
        .replace('__VENDOR_JS__',vjs)
        .replace('__CALIBER__',caliber)
        .replace('__TOTAL__',f'{total:,}')
        .replace('__CITIES__',str(len(cities)))
        .replace('__HIGH__',f'{kpi_high:,}')
        .replace('__OFFICIAL__',f'{official:,}')
        .replace('__SKYCARDS__',''.join(sky_cards))
        .replace('__TIERBAR__',tier_bar)
        .replace('__TIERLEGEND__',tier_legend)
        .replace('__SRCITEMS__',src_items)
        .replace('__DATA__',data_json))
    p=os.path.join(OUT,'Indonesia_HighRise_Report.html')
    open(p,'w',encoding='utf-8').write(out)
    print(f'已生成: {p} ({os.path.getsize(p)/1024/1024:.1f} MB)')
    print(f'  总计 {total} 栋 | 城市 {len(cities)} | ≥80m {kpi_high} | 官方 {official}')
    print(f'  两口径: 单栋 {total} | grouped {grouped} | 近距簇 {n_clusters} 个(成员 {cluster_members} 栋)')
    print(f'  天际线卡片 {len(sky_cards)} 城')

build()
