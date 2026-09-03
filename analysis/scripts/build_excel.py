# -*- coding: utf-8 -*-
"""
雅加达高层建筑清单 → Excel
- 读 jakarta_official_geom_raw.json (7220栋, 含 gmaps 坐标)
- 三层剔除公共设施: ①信任官方商业细类 ②居住豁免 ③官方公共细类/楼名关键词
- 无名楼用 OSM 坐标匹配补真名(新增 名称来源/匹配距离 列, 官方真名与补名严格区分)
- 按高度米数分 5 档 (32/40/48/64/80m = 8/10/12/16/20层)
- 输出: Sheet1 商业高层(纯净) / Sheet2 已剔除-公共设施 / Sheet3 说明
"""
import json, os, re, sys, math, collections
sys.stdout.reconfigure(encoding="utf-8")
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
RAW = os.path.join(HERE, "..", "data", "jakarta_official_geom_raw.json")
OSM_RAW = os.path.join(HERE, "..", "data", "jakarta_buildings_raw.json")
OUT = os.path.join(HERE, "..", "output", "雅加达高层建筑清单.xlsx")
FETCH_DATE = "2026-06-22"
TRUST_M = 50    # ≤此距离视为可信补名(也用于二次剔除)
MAX_M = 100     # ≤此距离才采纳(50~100标"存疑")


# ---------------- 工具 ----------------
def clean(v):
    if v is None: return ""
    v = str(v).strip()
    return "" if v in ("<Null>", "-", ".", "null", "NULL") else v

def normj(v):
    """归一化 jns_bgn: 小写, 下划线/标点/多空格 → 单空格"""
    return re.sub(r"[\s_,/().]+", " ", clean(v).lower()).strip()

def official(rec):
    return clean(rec.get("nama_tower")) or clean(rec.get("nama_bgn"))


# ---------------- 分类规则 ----------------
# 第一层: 信任的官方商业/居住细类 (直接保留, 不被楼名兜底推翻); 顺序敏感, 综合体优先
KEEP_TRUST = [
    ("综合体",     ["kantor apartemen", "kantor apt", "apartemen dan fasilitas", "apt fasilitas",
                    "campuran", "multifungsi", "mix", "pertokoan kantor", "pertokoan hotel",
                    "hotel hunian", "kantor hunian", "kantor hotel"]),
    ("商场",       ["pusat perbelanjaan", "perbelanjaan"]),
    ("酒店",       ["perhotelan"]),
    ("公寓",       ["rumah tinggal susun", "rumah flat", "rumah kost"]),
    ("商业/店铺",  ["perdagangan", "pertokoan", "rumah toko", "ruko"]),
    ("住宅",       ["rumah tinggal"]),
    ("办公住宅",   ["rumah kantor"]),
    ("工业",       ["perindustrian", "industri"]),
    ("仓库",       ["gudang", "penyimpanan"]),
    ("旅游休闲",   ["wisata", "rekreasi"]),
]

# 第三层(a): 官方公共细类 → 剔除
PUBLIC_JNS = [
    ("医院",          ["kesehatan"]),
    ("学校",          ["pendidikan"]),
    ("政府/公共服务", ["pelayanan umum"]),
    ("军事/国防",     ["pertahanan", "keamanan"]),
    ("宗教场所",      ["gereja", "vihara", "masjid", "musollah", "mushola"]),
    ("交通/停车",     ["transportasi", "parkir"]),
    ("体育设施",      ["olahraga"]),
    ("文化设施",      ["kebudayaan"]),
]

# 第三层(b): 楼名关键词兜底 (针对 jns_bgn 是纯写字楼/Lainnya/空 但名字明显是公共机构)
NAME_RULES = [
    ("医院", re.compile(r"\b(rumah sakit|rsud|rsup|rsia|rsab|rsu|rs|klinik|puskesmas|hospital)\b|medical cent", re.I)),
    ("学校", re.compile(r"\b(sekolah|sdn?|smpn?|sman?|smkn?|paud|universitas|university|kampus|campus|"
                        r"pesantren|madrasah|institut|institute|politeknik|poltek|akademi|college)\b|"
                        r"sekolah tinggi|dinas pendidikan", re.I)),
    ("政府/公共服务", re.compile(r"\b(kementerian|kementrian|kemenko|pemerintah|pemprov|pemkot|pemkab|pemda|"
                        r"gubernur|kejaksaan|kejagung|pengadilan|mahkamah|kepolisian|polda|polres|polsek|polri|"
                        r"mabes|kodam|kodim|korem|koramil|lanud|lantamal|dprd|bappeda|bappenas|samsat|imigrasi|"
                        r"kanwil|ditjen|balai kota|balaikota|wali kota|walikota|kedutaan|konsulat|embassy|consulate|"
                        r"\bBKN\b|\bbkn\b|badan pengawas|badan pemeriksa|badan informasi|badan pusat statistik|"
                        r"\bbps\b|\bbpk\b|\bbpn\b|\bbappenas\b|\bbasarnas\b|\bbmkg\b|\bbppt\b|\blankam\b)\b|"
                        r"kantor (dinas|pemerintah|wali|gubernur|camat|kelurahan|pajak|pos|imigrasi|badan)|"
                        r"direktorat jenderal|\bdinas |duta besar|\bbadan \b", re.I)),
]

RESI_NAME = re.compile(r"apartemen|apartment|\brusun\b|rumah susun|asrama|residence|kondominium|condominium|\bkost\b", re.I)
RESI_JNS = ["rumah tinggal susun", "rumah flat", "rumah kost", "rumah tinggal"]


def fungsi_fallback(fungsi, name=""):
    f = clean(fungsi).lower()
    if "keagamaan" in f:                       return ("剔除", "宗教场所", "fungsi兜底", "")
    if "sosial budaya" in f or "sosbud" in f:  return ("剔除", "文化设施", "fungsi兜底", "")
    if f == "hunian":                          return ("保留", "住宅类", "", "")
    if "usaha" in f:                           return ("保留", "商业类", "", "")
    if "campuran" in f:                        return ("保留", "综合体", "", "")
    if "nonhunian" in f or "non hunian" in f:  return ("保留", "非住宅", "", "")
    if "khusus" in f:
        # 特殊用途: 再查一次楼名关键词（防止医院/政府机构 jns_bgn=None 时漏网）
        for cat, pat in NAME_RULES:
            m = pat.search(name)
            if m:
                return ("剔除", cat, "fungsi兜底+楼名关键词", m.group(0).strip())
        return ("保留", "特殊(待核)", "", "")
    return ("保留", "未分类", "", "")


def is_resi(jns, name, fungsi):
    if any(k in jns for k in RESI_JNS):     return True
    if RESI_NAME.search(name):              return True
    if clean(fungsi).lower() == "hunian":   return True
    return False


def classify(rec):
    """返回 (保留/剔除, 用途或剔除类别, 判定依据, 命中词); 名称用 rec['_kwname'](官方名+可信OSM补名)"""
    jns = normj(rec.get("jns_bgn"))
    name = rec.get("_kwname", "")
    fungsi = rec.get("fungsi")
    resi = is_resi(jns, name, fungsi)

    # 1) 官方明确商业/居住保留类 → 信任, 直接保留 (不被楼名兜底推翻)
    for cat, kws in KEEP_TRUST:
        if any(kw in jns for kw in kws):
            return ("保留", cat, "", "")

    # 2) 官方公共细类 → 剔除 (居住豁免: 公租房/宿舍当公寓保留)
    for cat, kws in PUBLIC_JNS:
        if any(kw in jns for kw in kws):
            if resi:
                return ("保留", "公寓", "居住豁免", "")
            return ("剔除", cat, "官方细类", clean(rec.get("jns_bgn")))

    # 3) 楼名关键词兜底 → 揪政府/学校/医院（先于居住豁免，避免"DINAS KEBAKARAN"因fungsi=Hunian漏网）
    for cat, pat in NAME_RULES:
        m = pat.search(name)
        if m:
            return ("剔除", cat, "楼名关键词", m.group(0).strip())

    # 4) jns 为 写字楼/Lainnya/空: 居住豁免 (公家宿舍/公租房 → 公寓)
    if resi:
        cat = "公寓" if (RESI_NAME.search(name) or any(k in jns for k in ["rumah tinggal susun", "rumah flat", "rumah kost"])) else "住宅类"
        return ("保留", cat, "", "")

    # 5) 剩余
    if "perkantoran" in jns:
        return ("保留", "写字楼", "", "")
    if jns:
        return ("保留", "其他", "", "")
    keep, cat, basis, hit = fungsi_fallback(rec.get("fungsi"), name)
    return (keep, cat, basis, hit)


def tier(h):
    if h >= 80: return "≥80米(约20层)"
    if h >= 64: return "≥64米(约16层)"
    if h >= 48: return "≥48米(约12层)"
    if h >= 40: return "≥40米(约10层)"
    return "≥32米(约8层)"


# ---------------- OSM 补名 ----------------
def load_osm():
    els = json.load(open(OSM_RAW, encoding="utf-8")).get("elements", [])
    names, la, lo = [], [], []
    for e in els:
        nm = e.get("tags", {}).get("name"); c = e.get("center")
        if nm and c:
            names.append(nm); la.append(c["lat"]); lo.append(c["lon"])
    return names, np.array(la), np.array(lo)

def nearest_osm(lat, lon, names, ola, olo):
    dy = (ola - lat) * 110574.0
    dx = (olo - lon) * 111320.0 * math.cos(math.radians(lat))
    d = np.sqrt(dx * dx + dy * dy)
    i = int(np.argmin(d))
    return names[i], float(d[i])


# ---------------- 主流程 ----------------
def main():
    data = json.load(open(RAW, encoding="utf-8"))
    names, ola, olo = load_osm()
    print(f"OSM 带名要素: {len(names)}")

    # OSM 补名 (仅无官方名且有坐标者)
    n_fill = n_trust = 0
    for rec in data:
        off = official(rec)
        rec["_osm"] = ""; rec["_mdist"] = None
        if not off and rec.get("_lat") is not None:
            nm, dist = nearest_osm(rec["_lat"], rec["_lon"], names, ola, olo)
            if dist <= MAX_M:
                rec["_osm"] = nm; rec["_mdist"] = round(dist, 1); n_fill += 1
                if dist <= TRUST_M: n_trust += 1
        # 关键词判定用名: 合并官方 tower名+building名(两者都取，避免"GEDUNG 2"淹没"BKN")，否则可信OSM补名
        tower = clean(rec.get("nama_tower"))
        bgn = clean(rec.get("nama_bgn"))
        if tower and bgn and tower != bgn:
            rec["_kwname"] = tower + " " + bgn
        elif off:
            rec["_kwname"] = off
        elif rec["_osm"] and rec["_mdist"] is not None and rec["_mdist"] <= TRUST_M:
            rec["_kwname"] = rec["_osm"]
        else:
            rec["_kwname"] = ""
        # 最终展示名 + 来源
        if off:
            rec["_final"], rec["_src"] = off, "官方"
        elif rec["_osm"] and rec["_mdist"] <= TRUST_M:
            rec["_final"], rec["_src"] = rec["_osm"], "OSM补充"
        elif rec["_osm"]:
            rec["_final"], rec["_src"] = rec["_osm"], "OSM补充(存疑)"
        else:
            rec["_final"], rec["_src"] = "（无名）", "无"
    print(f"OSM 补名采纳: {n_fill} 栋 (其中可信≤{TRUST_M}m: {n_trust})")

    keep, drop = [], []
    drop_cat = collections.Counter()
    osm_drop = []  # 因OSM补名而被识别剔除的
    for rec in data:
        rec["_h"] = round(rec.get("bldgheight") or 0, 1)
        status, cat, basis, hit = classify(rec)
        rec["_status"] = status  # 记录分类结果，供坐标级后处理使用
        if status == "剔除" and basis == "楼名关键词" and not official(rec):
            basis = "楼名关键词(OSM补名)"
            osm_drop.append((rec["_final"], cat, hit, rec["_mdist"], rec["_h"]))
        rec["_cat"] = cat; rec["_basis"] = basis; rec["_hit"] = hit; rec["_tier"] = tier(rec["_h"])
        (drop if status == "剔除" else keep).append(rec)
        if status == "剔除":
            drop_cat[cat] += 1

    # ---------------- 坐标级后处理：同坐标同高度连带剔除 ----------------
    # 同一物理建筑可能有多条官方记录(jns_bgn不同)，若其中一条被剔除，
    # 则同坐标(round8)+同高度(round1)的其他记录也连带剔除，防止Sheet交叉
    coord_groups = {}
    for rec in data:
        lat = rec.get("_lat")
        lon = rec.get("_lon")
        h = rec.get("_h")
        if lat is not None and lon is not None and h is not None:
            key = (round(lat, 8), round(lon, 8), round(h, 1))
            if key not in coord_groups:
                coord_groups[key] = []
            coord_groups[key].append(rec)

    cross_drop = 0
    for key, recs in coord_groups.items():
        dropped = [r for r in recs if r.get("_status") == "剔除"]
        if not dropped:
            continue
        template = dropped[0]  # 沿用第一条被剔除记录的类别和判定依据
        for r in recs:
            if r.get("_status") != "剔除":
                orig_cat = r.get("_cat", "")
                r["_status"] = "剔除"
                r["_cat"] = template["_cat"]
                r["_basis"] = "坐标连带剔除(" + template["_basis"] + ")"
                r["_hit"] = template.get("_hit", "")
                cross_drop += 1
                print(f"  坐标连带剔除: [{template['_cat']}] {r.get('_final','?')} h={r['_h']}m "
                      f"({r.get('_lat')},{r.get('_lon')}) 原分类={orig_cat} → {template['_cat']}")

    if cross_drop:
        print(f"\n坐标连带剔除: {cross_drop} 条记录因同坐标同高度有被剔除记录而被连带剔除")

    # 重建保留/剔除列表（反映坐标级后处理结果）
    keep = [r for r in data if r.get("_status") != "剔除"]
    drop = [r for r in data if r.get("_status") == "剔除"]
    drop_cat = collections.Counter(r["_cat"] for r in drop)

    keep.sort(key=lambda r: -r["_h"])
    drop.sort(key=lambda r: -r["_h"])

    # 统计
    src_cnt = collections.Counter(r["_src"] for r in keep)
    print(f"\n总计 {len(data)}  →  保留 {len(keep)}  剔除 {len(drop)}")
    print(f"\nSheet1 名称来源: 官方 {src_cnt['官方']} | OSM补充 {src_cnt['OSM补充']} | "
          f"OSM补充(存疑) {src_cnt['OSM补充(存疑)']} | 无名 {src_cnt['无']}")
    print("\n剔除明细:")
    for c, n in drop_cat.most_common():
        print(f"  {n:4d}  {c}")
    print(f"\n因OSM补名而被剔除(二次剔除) {len(osm_drop)} 条:")
    for nm, cat, hit, dist, h in osm_drop[:20]:
        print(f"  [{cat}] {nm}  (匹配{dist}m, 命中'{hit}', {h}m)")
    print("\n保留各档 (商业高层):")
    for thr, lbl in [(32,"≥8层"),(40,"≥10层"),(48,"≥12层"),(64,"≥16层"),(80,"≥20层")]:
        print(f"  ≥{thr}m {lbl}: {sum(1 for r in keep if r['_h'] >= thr)}")

    write_excel(data, keep, drop)


# ---------------- 写 Excel ----------------
def bname(rec):
    return rec.get("_final") or official(rec) or "（无名）"

def write_excel(data, keep, drop):
    HDR = Font(bold=True, color="FFFFFF", size=10)
    HDR_FILL = PatternFill("solid", fgColor="2F5496")
    CEN = Alignment(horizontal="center", vertical="center")

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c); cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = CEN
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}1"

    def mdist_cell(r):
        return r["_mdist"] if str(r["_src"]).startswith("OSM") and r["_mdist"] is not None else "—"

    wb = Workbook()

    # Sheet1
    ws1 = wb.active; ws1.title = "商业高层(纯净)"
    cols1 = ["序号", "楼宇名称", "名称来源", "OSM匹配距离(米)", "建筑全名(官方)", "纬度", "经度",
             "高度(米)", "层数", "高度分档", "用途分类", "行政区", "街道办区", "地址", "建筑面积(㎡)"]
    ws1.append(cols1)
    for i, r in enumerate(keep, 1):
        lap = r.get("jml_lapis")
        ws1.append([
            i, r["_final"], r["_src"], mdist_cell(r),
            clean(r.get("nama_bgn")) or "—", r.get("_lat"), r.get("_lon"), r["_h"],
            int(lap) if lap and lap > 0 else "", r["_tier"], r["_cat"], clean(r.get("_dist")),
            clean(r.get("wadmkc")) or "—", clean(r.get("alamat")) or "—",
            round(r.get("luas_bgn"), 1) if r.get("luas_bgn") else "",
        ])
    style_header(ws1, len(cols1))

    # Sheet2
    ws2 = wb.create_sheet("已剔除-公共设施")
    cols2 = ["序号", "楼宇名称", "名称来源", "建筑全名(官方)", "纬度", "经度", "高度(米)", "层数",
             "剔除类别", "判定依据", "命中关键词", "官方细类(jns_bgn)", "行政区", "街道办区", "地址"]
    ws2.append(cols2)
    for i, r in enumerate(drop, 1):
        lap = r.get("jml_lapis")
        ws2.append([
            i, r["_final"], r["_src"], clean(r.get("nama_bgn")) or "—",
            r.get("_lat"), r.get("_lon"), r["_h"],
            int(lap) if lap and lap > 0 else "", r["_cat"], r["_basis"],
            clean(r.get("_hit")) or "—", clean(r.get("jns_bgn")) or "—", clean(r.get("_dist")),
            clean(r.get("wadmkc")) or "—", clean(r.get("alamat")) or "—",
        ])
    style_header(ws2, len(cols2))

    # 列宽
    W1 = {"A":6,"B":32,"C":13,"D":14,"E":30,"F":12,"G":12,"H":9,"I":7,"J":15,"K":11,"L":10,"M":16,"N":38,"O":13}
    for k,v in W1.items(): ws1.column_dimensions[k].width = v
    W2 = {"A":6,"B":30,"C":13,"D":30,"E":12,"F":12,"G":9,"H":7,"I":13,"J":18,"K":22,"L":22,"M":10,"N":16,"O":38}
    for k,v in W2.items(): ws2.column_dimensions[k].width = v
    for col in ("F", "G"):
        for row in range(2, ws1.max_row + 1): ws1[f"{col}{row}"].number_format = "0.000000"
    for col in ("E", "F"):
        for row in range(2, ws2.max_row + 1): ws2[f"{col}{row}"].number_format = "0.000000"

    # Sheet3 说明
    ws3 = wb.create_sheet("说明")
    def kge(thr): return sum(1 for r in keep if r["_h"] >= thr)
    def age(thr): return sum(1 for r in data if r["_h"] >= thr)
    src = collections.Counter(r["_src"] for r in keep)
    lines = [
        ("雅加达高层建筑清单 — 数据说明", True),
        ("", False),
        ("数据来源：雅加达省政府 DPMPTSP 三维建筑库 (Jakarta Satu, gis-dpmptsp.jakarta.go.id ArcGIS REST)", False),
        ("数据性质：真实数据 — 三维(LOD2)建模反算高度 + 政府验证(Validasi)，非 AI 二维估算", False),
        (f"抓取日期：{FETCH_DATE}    覆盖：南/中/北/西/东 五区全市", False),
        ("坐标来源：官方记录 gmaps 字段内嵌经纬度(楼栋点)，坐标系 WGS84；纬度为负(南半球)", False),
        ("", False),
        ("【口径定义】高层分 5 档，按建筑高度(米)，约 4 米/层近似换算楼层：", True),
        ("    ≥32米(约8层) · ≥40米(约10层) · ≥48米(约12层) · ≥64米(约16层) · ≥80米(约20层)", False),
        ("", False),
        ("【公共设施剔除规则】只保留写字楼/公寓/酒店/商场等商业+居住楼；剔除政府/学校/医院/宗教/军事/交通/体育/文化等公共设施。", True),
        ("    三层判定：① 信任官方商业细类(综合体/酒店/商场/公寓直接保留)；② 居住豁免(公租房/宿舍/机构公寓当公寓保留)；", False),
        ("    ③ 官方用途细类 jns_bgn 命中公共类目 或 楼名关键词兜底(捞出标为写字楼但实为部委/厅局/医院/学校的)。", False),
        ("    全部被剔记录见『已剔除-公共设施』表，判定依据/命中关键词列已标明，可逐条复核。", False),
        ("", False),
        ("【无名楼补名】官方无名楼用 OpenStreetMap 坐标最近邻匹配真名(OSM=众包真实地图，非估算)。", True),
        (f"    名称来源列标明：官方 / OSM补充(≤{TRUST_M}米,可信) / OSM补充(存疑,{TRUST_M}~{MAX_M}米) / 无；并给出 OSM匹配距离。", False),
        ("    官方真名绝不被覆盖；补名仅用于原本无名的楼。距离越小越可信，可按『名称来源』『匹配距离』列自行取舍。", False),
        (f"    本次：官方名 {src['官方']} 栋；OSM补充(可信) {src['OSM补充']} 栋；OSM补充(存疑) {src['OSM补充(存疑)']} 栋；仍无名 {src['无']} 栋。", False),
        ("", False),
        ("【统计】", True),
        (f"    官方全量高层(≥32米)：{len(data)} 栋    剔除公共设施：{len(drop)} 栋    商业高层(本清单)：{len(keep)} 栋", False),
        ("", False),
        ("    各档数量（栋）：    官方全量  →  剔除公共设施后(商业高层)", False),
        (f"      ≥32米(约8层) ：  {age(32):5d}  →  {kge(32)}", False),
        (f"      ≥40米(约10层)：  {age(40):5d}  →  {kge(40)}", False),
        (f"      ≥48米(约12层)：  {age(48):5d}  →  {kge(48)}", False),
        (f"      ≥64米(约16层)：  {age(64):5d}  →  {kge(64)}", False),
        (f"      ≥80米(约20层)：  {age(80):5d}  →  {kge(80)}", False),
        ("", False),
        ("【数据局限】", True),
        ("    1. 坐标为楼栋代表点(非门牌精确点)。", False),
        ("    2. 约 24% 楼官方无层数，故分档统一按高度(米)，层数列仅供参考。", False),
        ("    3. OSM 补名为跨源匹配，存疑项(>50米)可能匹配到邻近建筑，请按距离判断；仍无名者为本身无正式楼名的普通楼。", False),
        ("    4. 按官方记录口径，同一建筑群的多座塔楼/裙楼可能各占一行(与官方 7220 口径一致，未强制去重)。", False),
        ("    5. 用途/楼名均无线索者默认保留并标『未分类/特殊(待核)』，秉持『存疑不删』原则。", False),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        c = ws3.cell(i, 1, txt)
        if bold: c.font = Font(bold=True, size=11)
    ws3.column_dimensions["A"].width = 120

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"\n已生成: {os.path.abspath(OUT)}")


if __name__ == "__main__":
    main()
