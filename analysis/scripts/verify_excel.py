# -*- coding: utf-8 -*-
"""读回生成的 xlsx, 端到端验证 (适配带OSM补名的新列结构)"""
import os, sys, collections
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

XLSX = os.path.join(os.path.dirname(__file__), "..", "output", "雅加达高层建筑清单.xlsx")
wb = load_workbook(XLSX)
print("Sheet 列表:", wb.sheetnames)
s1 = wb["商业高层(纯净)"]; s2 = wb["已剔除-公共设施"]
n1, n2 = s1.max_row - 1, s2.max_row - 1
print(f"\nSheet1: {n1} 行   Sheet2: {n2} 行   合计 {n1+n2} (应=7220)")
print("\n表头1:", [c.value for c in s1[1]])

# Sheet1 列: 0序号1名称2来源3距离4全名5纬6经7高8层9档10用途11区12街道13址14面积
print("\n--- Sheet1 Top12 (高度|用途|来源|名称|坐标) ---")
for r in s1.iter_rows(min_row=2, max_row=13, values_only=True):
    print(f"  {r[7]:>6}m {str(r[10]):<6} [{str(r[2]):<10}] {str(r[1])[:30]:<30} ({r[5]},{r[6]})")

print("\n--- 名称来源分布 ---")
src = collections.Counter(r[2] for r in s1.iter_rows(min_row=2, values_only=True))
for k, v in src.items(): print(f"  {k}: {v}")

print("\n--- OSM补名样例(来源含OSM, 前10) ---")
cnt = 0
for r in s1.iter_rows(min_row=2, values_only=True):
    if str(r[2]).startswith("OSM"):
        print(f"  {str(r[2]):<14} 距{r[3]}m | {str(r[1])[:34]} | {r[7]}m [{r[10]}]")
        cnt += 1
        if cnt >= 10: break

# 坐标 sanity
bad = 0
for r in s1.iter_rows(min_row=2, values_only=True):
    lat, lon = r[5], r[6]
    if lat is None or lon is None or not (-6.5 < lat < -5.9 and 106.5 < lon < 107.2): bad += 1
print(f"\n坐标越界/缺失: {bad}")

print("\n--- Sheet1 各档(高度米数) ---")
rows1 = list(s1.iter_rows(min_row=2, values_only=True))
for thr, lbl in [(32,"≥8层"),(40,"≥10层"),(48,"≥12层"),(64,"≥16层"),(80,"≥20层")]:
    print(f"  ≥{thr}m {lbl}: {sum(1 for r in rows1 if (r[7] or 0) >= thr)}")

# Sheet2 Top: 列 8剔除类别 9判定依据
print("\n--- Sheet2 已剔除 Top10 (高度|类别|依据|名称) ---")
for r in s2.iter_rows(min_row=2, max_row=11, values_only=True):
    print(f"  {r[6]:>6}m [{str(r[8]):<10}] {str(r[9]):<16} {str(r[1])[:34]}")

found = [r for r in rows1 if r[1] and "Autograph" in str(r[1])]
print(f"\nAutograph 在 Sheet1: {'是' if found else '否'}")
