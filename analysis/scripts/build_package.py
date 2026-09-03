# -*- coding: utf-8 -*-
"""生成交付包：读实际数据算统计→动态README→组织目录→zip。用法：python build_package.py"""
import os, shutil, zipfile
from openpyxl import load_workbook

HERE=os.path.dirname(os.path.abspath(__file__)); BASE=os.path.join(HERE,'..')
OUT=os.path.join(BASE,'output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']

def stats():
    total=0; jkt=0; ge80=0; hosp=0
    for c in CITIES:
        wb=load_workbook(os.path.join(OUT,f'{c}高层建筑清单.xlsx'),read_only=True,data_only=True)
        ws=wb['商业高层(纯净)']
        h=[str(x) for x in next(ws.iter_rows(min_row=1,max_row=1,values_only=True))]
        hi=h.index('高度(米)'); ui=h.index('用途分类')
        n=0
        for r in ws.iter_rows(min_row=2,values_only=True):
            if all(v is None for v in r):continue
            n+=1; total+=1
            hh=float(r[hi]) if r[hi] else 0
            if hh>=80: ge80+=1
            if c=='雅加达' and str(r[ui])=='医院': hosp+=1
        if c=='雅加达': jkt=n
        wb.close()
    return total, jkt, ge80, hosp

def readme(total, jkt, ge80, hosp):
    other=total-jkt
    return f"""# 印尼12城高层建筑分析 · 交付说明

## 一、项目概况

本交付包为印尼 12 座主要城市**高层建筑（≥32 米，约 8 层及以上）**的分析成果，共 **{total} 栋**。

## 二、数据来源与可信度

| 范围 | 来源 | 性质 | 栋数 |
|------|------|------|------|
| 雅加达 | 印尼 DPMPTSP 官方三维建筑库（Jakarta Satu） | **官方真值**（LOD2 建模 + 政府验证） | {jkt} |
| 其余 11 城 | Google Open Buildings V3 足迹 + 2.5D 高度估算 | 机器学习估算（保守下限） | {other} |

- **雅加达是"官方真值标杆"**——高度、坐标、面积均来自政府登记。
- **其余 11 城为 ML 估算**，高度系统性偏低（保守下限），实际只高不低；少数地标楼采用 CTBUH 实测真高（见 Excel「数据来源」列）。

## 三、高度口径

按建筑高度分 5 档（约 4 米/层折算）：≥32 米(8层) · ≥40 米(10层) · ≥48 米(12层) · ≥64 米(16层) · ≥80 米(20层)

## 四、去重处理（重要）

**同一栋楼常被重复计入**，本分析已做空间去重避免高估：
- **雅加达**：官方登记同一栋楼有多条税务/登记记录 → 精确去重 + 同名近距 + 25米空间去重。
- **其余 11 城**：Google ML 会把一栋楼拆成多个足迹点 → 25米空间去重（半径25m+高度差<25%，同楼拆点合并、相邻不同高楼保留）。
- 全 12 城统一去重口径，医院并入商业高层（不作公共设施剔除）。

## 五、文件导航

### 📁 建筑清单_12城/
12 个城市各一个 Excel，每个含 3 个 sheet：商业高层(纯净)（核心主数据，含医院）/ 已剔除-公共设施 / 说明。
每栋含 17 列：楼名、名称来源、数据来源、坐标、高度、层数、用途、行政区、地址、面积等。

### 📁 汇总统计/
- **12城高层建筑汇总.xlsx**：分档统计总表（累计阈值 + 逐档明细 + 说明）
- **12城高层建筑汇总.md**：同内容 Markdown 版

### 📁 可视化/
- **Indonesia_HighRise_Report.html** — 英文一体化交互报告（概览+地图+可搜索清单+导出）。**全英文，可直接转发外方**，双击打开。
- **建筑分布地图.html** — 中文交互地图，按高度分档着色，可缩放点击。
- **建筑分布地图.kml** — Google Earth 用，卫星 3D 查看。
- **建筑坐标.csv** — 全部建筑坐标明细，可导入其他 GIS 工具。

## 六、使用提醒

- HTML 地图/报告**需联网**加载地图底图（首次打开需网络）；**建筑数据已内嵌文件，是固定快照，不会被网络改动**。
- 英文报告已内联地图库，转发后对方双击即用。
- 楼名与地址保留印尼语原文（当地实地调研、导航需用原名）。

## 七、关键数字

| 指标 | 数值 |
|------|------|
| 12 城合计（≥32 米） | **{total} 栋** |
| 其中 ≥80 米超高层 | {ge80} 栋 |
| 雅加达（官方真值） | {jkt} 栋（含医院 {hosp} 栋）|
| 数据生成日期 | 2026-07-07 |

---

*雅加达为官方登记去重后真值；其余 11 城为 Google 估算保守下限，实际数量只多不少。全 12 城已做空间去重，避免同楼多点高估。*
"""

def main():
    total, jkt, ge80, hosp = stats()
    PKG=os.path.join(BASE,'印尼高层建筑分析_交付包')
    if os.path.exists(PKG): shutil.rmtree(PKG)
    d_data=os.path.join(PKG,'建筑清单_12城'); d_sum=os.path.join(PKG,'汇总统计'); d_vis=os.path.join(PKG,'可视化')
    for d in (d_data,d_sum,d_vis): os.makedirs(d)
    with open(os.path.join(PKG,'README.md'),'w',encoding='utf-8') as f:
        f.write(readme(total,jkt,ge80,hosp))
    for c in CITIES: shutil.copy(os.path.join(OUT,f'{c}高层建筑清单.xlsx'), d_data)
    for f in ['12城高层建筑汇总.xlsx','12城高层建筑汇总.md']: shutil.copy(os.path.join(OUT,f), d_sum)
    for f in ['Indonesia_HighRise_Report.html','建筑分布地图.html','建筑分布地图.kml','建筑坐标.csv']:
        shutil.copy(os.path.join(OUT,f), d_vis)
    zipname=os.path.join(BASE,'印尼高层建筑分析_交付包.zip')
    if os.path.exists(zipname): os.remove(zipname)
    with zipfile.ZipFile(zipname,'w',zipfile.ZIP_DEFLATED) as z:
        for root,_,files in os.walk(PKG):
            for f in files:
                fp=os.path.join(root,f); z.write(fp, os.path.relpath(fp,BASE))
    shutil.rmtree(PKG)
    print(f'交付包已生成: {zipname} ({os.path.getsize(zipname)/1024/1024:.1f} MB)')
    print(f'  README数字: 12城合计{total} | 雅加达{jkt}(医院{hosp}) | ≥80m {ge80}')

main()
