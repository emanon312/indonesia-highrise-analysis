# -*- coding: utf-8 -*-
"""12城 Excel：把剔除表里的医院行剪切到新增「医院」sheet。dry-run默认，--run执行。"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')
CITIES = ['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
DROP = '已剔除-公共设施'
HOSP = '医院'
LOG = []
def log(s): LOG.append(s); print(s)

def cat_col(header):
    for key in ['剔除类别','用途分类']:
        for i,h in enumerate(header):
            if h and key in str(h): return i
    return None

def run_city(city, write):
    path = os.path.join(OUT, city+'高层建筑清单.xlsx')
    if not os.path.exists(path):
        log(city+' [缺失]'); return 0
    wbr = load_workbook(path, read_only=True, data_only=True)
    if DROP not in wbr.sheetnames:
        log(city+' [无剔除表]'); wbr.close(); return 0
    rows = list(wbr[DROP].iter_rows(values_only=True))
    wbr.close()
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    ci = cat_col(header)
    ni = next((i for i,h in enumerate(header) if '楼宇名称' in h), 1)
    if ci is None:
        log(city+' [无分类列] '+str(header)); return 0
    hosp_idx = []; samples = []
    for excel_row, r in enumerate(rows[1:], start=2):
        if all(v is None for v in r): continue
        cat = str(r[ci]) if ci < len(r) and r[ci] is not None else ''
        if '医院' in cat:
            hosp_idx.append(excel_row)
            nm = str(r[ni]) if ni < len(r) and r[ni] else '(无名)'
            if len(samples) < 3: samples.append(nm[:22])
    log('%-5s %2d列 分类列第%d列 医院=%-3d %s' % (city, len(header), ci+1, len(hosp_idx), ', '.join(samples)))
    if not write or not hosp_idx:
        return len(hosp_idx)
    wb = load_workbook(path)
    dws = wb[DROP]
    if HOSP in wb.sheetnames: del wb[HOSP]
    hws = wb.create_sheet(HOSP, wb.sheetnames.index(DROP)+1)
    allv = list(dws.iter_rows(values_only=True))
    for c,v in enumerate(header, start=1): hws.cell(row=1,column=c,value=v)
    wr = 2
    for ei in hosp_idx:
        for c,v in enumerate(allv[ei-1], start=1): hws.cell(row=wr,column=c,value=v)
        hws.cell(row=wr,column=1,value=wr-1)
        wr += 1
    for ei in sorted(hosp_idx, reverse=True): dws.delete_rows(ei,1)
    for i,_ in enumerate(list(dws.iter_rows(min_row=2,values_only=True))):
        dws.cell(row=i+2,column=1,value=i+1)
    wb.save(path); wb.close()
    log('      -> 已写入医院sheet %d行，剔除表删%d行' % (len(hosp_idx), len(hosp_idx)))
    return len(hosp_idx)

def main():
    write = '--run' in sys.argv
    log('='*56); log('  医院搬移 - ' + ('执行' if write else 'dry-run预览')); log('='*56)
    tot = 0
    for c in CITIES: tot += run_city(c, write)
    log('-'*56); log('  合计医院: %d' % tot)
    open(os.path.join(OUT,'_医院搬移日志.txt'),'w',encoding='utf-8').write('\n'.join(LOG))

main()
