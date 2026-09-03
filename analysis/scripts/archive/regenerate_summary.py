"""从12城Excel重新生成汇总MD —— 只读第一个sheet(商业高层(纯净))的高度分档列"""
import os
import re
from datetime import datetime
from openpyxl import load_workbook

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")

# 12城列表，顺序无所谓，最后按≥32m降序排列
CITIES = [
    "雅加达", "泗水", "万隆", "唐格朗", "望加锡", "棉兰",
    "巴淡", "三宝垄", "勿加泗", "德波", "茂物", "巨港"
]

# 需要先做一次对比：之前的值 vs 当前值
OLD_VALUES = {
    "雅加达":  {32: 7220, 40: 5261, 48: 4226, 64: 2943, 80: 2105},
    "泗水":    {32: 811,  40: 547,  48: 402,  64: 202,  80: 107},
    "万隆":    {32: 402,  40: 221,  48: 143,  64: 72,   80: 38},
    "唐格朗":  {32: 380,  40: 240,  48: 179,  64: 91,   80: 48},
    "望加锡":  {32: 214,  40: 119,  48: 75,   64: 28,   80: 8},
    "棉兰":    {32: 191,  40: 107,  48: 83,   64: 42,   80: 20},
    "巴淡":    {32: 189,  40: 103,  48: 62,   64: 33,   80: 13},
    "三宝垄":  {32: 139,  40: 77,   48: 54,   64: 23,   80: 4},
    "勿加泗":  {32: 97,   40: 67,   48: 49,   64: 22,   80: 9},
    "德波":    {32: 75,   40: 56,   48: 37,   64: 24,   80: 16},
    "茂物":    {32: 95,   40: 59,   48: 40,   64: 14,   80: 6},
    "巨港":    {32: 77,   40: 35,   48: 24,   64: 9,    80: 2},
}
OLD_TOTALS = {32: 9890, 40: 6892, 48: 5374, 64: 3503, 80: 2376}

rows_out = []
changes = []  # 记录变化

# 两口径累计：每栋单算=主表行数之和；建筑群合算=独立楼+近距簇数
total_rows = 0        # 12城主表数据行数之和（每栋单算）
all_clusters = set()  # 全12城近距簇ID集合（城市代码前缀保证不重复）
cluster_members = 0   # 近距簇成员总数

for city in CITIES:
    xlsx_path = os.path.join(OUT_DIR, f"{city}高层建筑清单.xlsx")
    if not os.path.exists(xlsx_path):
        print(f"  [跳过] 未找到：{xlsx_path}")
        continue

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 第一个sheet = "商业高层(纯净)"
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        print(f"  [跳过] {city} sheet 为空")
        continue

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # 找包含"分档"的列
    bi = next((i for i, h in enumerate(header) if "分档" in h), None)
    if bi is None:
        print(f"  [警告] {city} 未找到分档列，header: {header}")
        wb.close()
        continue

    buckets = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
    # 近距簇列（末列，可能不存在于个别文件，做容错）
    ci = next((i for i, h in enumerate(header) if "近距簇" in h), None)
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        total_rows += 1  # 每栋单算：统计实际数据行
        if ci is not None and ci < len(r) and r[ci] not in (None, ""):
            all_clusters.add(str(r[ci]))
            cluster_members += 1
        band_str = str(r[bi]) if bi < len(r) and r[bi] is not None else ""
        m = re.search(r"(80|64|48|40|32)", band_str)
        if m:
            thr = int(m.group(1))
            if thr in buckets:
                buckets[thr] += 1
    wb.close()

    # 累计统计：≥t = sum(buckets[k] for k >= t)
    cum = {t: sum(v for k, v in buckets.items() if k >= t) for t in (32, 40, 48, 64, 80)}
    src = "官方真值" if city == "雅加达" else "Google ML估算(下限)"
    rows_out.append({"name": city, **cum, "src": src, "_buckets": buckets})

    # 检查变化
    if city in OLD_VALUES:
        for thr in [32, 40, 48, 64, 80]:
            old = OLD_VALUES[city][thr]
            new = cum[thr]
            if old != new:
                changes.append(f"  {city} ≥{thr}m: {old} → {new} ({'+' if new > old else ''}{new - old})")

    print(f"  {city}: 总行数={len(rows)-1}, buckets={buckets}, cum={cum}")

# 按 ≥32m 降序排列
rows_out.sort(key=lambda x: x[32], reverse=True)

# 构建 MD
lines = []
lines.append("# 印尼12城高层建筑汇总（按累计阈值统计栋数）\n")
lines.append("> 口径：4米/层。≥32米=8层 / ≥40=10 / ≥48=12 / ≥64=16 / ≥80=20。")
jkt_row = next((r for r in rows_out if r["name"] == "雅加达"), None)
if jkt_row:
    lines.append(f"> 雅加达为官方真值标杆（官方登记去重后保留商业高层 {jkt_row[32]} 栋，含医院、已剔除其他公共设施）；其余11城为 Google ML 估算**下限**。")
lines.append(f"> 自动生成于 {datetime.now().strftime('%Y-%m-%d %H:%M')}。\n")
# 两口径说明（读12城数据实算）：每栋单算 vs 建筑群合算
_n_cluster = len(all_clusters)
_indep = total_rows - cluster_members
_group_total = _indep + _n_cluster
lines.append(f"> **两口径栋数**：每栋单算 **{total_rows}** 栋（12城主表行数之和）；建筑群合算 **{_group_total}** 个（= 独立楼 {_indep} + 近距簇 {_n_cluster} 个；近距簇为 <30 米相邻楼聚合的疑似建筑群，每簇计 1 个）。\n")
lines.append("| 城市 | ≥32米(8层) | ≥40米(10层) | ≥48米(12层) | ≥64米(16层) | ≥80米(20层) | 数据来源 |")
lines.append("|------|-----------:|-----------:|-----------:|-----------:|-----------:|----------|")

tot = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
for r in rows_out:
    lines.append(f"| {r['name']} | {r[32]} | {r[40]} | {r[48]} | {r[64]} | {r[80]} | {r['src']} |")
    for k in tot:
        tot[k] += r[k]
lines.append(f"| **合计** | **{tot[32]}** | **{tot[40]}** | **{tot[48]}** | **{tot[64]}** | **{tot[80]}** | — |")

md_path = os.path.join(OUT_DIR, "12城高层建筑汇总.md")
with open(md_path, "w", encoding="utf-8") as f:
    f.write("\n".join(lines) + "\n")

# ========== 输出结果 ==========
print(f"\n{'='*60}")
print(f"  汇总 MD 已重新生成")
print(f"{'='*60}")
print(f"文件：{md_path}\n")

print("各城统计：")
for r in rows_out:
    print(f"  {r['name']:6s}  ≥32m={r[32]:>5d}  ≥40m={r[40]:>5d}  ≥48m={r[48]:>5d}  ≥64m={r[64]:>5d}  ≥80m={r[80]:>5d}  ({r['src']})")

print(f"\n合计：")
print(f"  ≥32m={tot[32]}  ≥40m={tot[40]}  ≥48m={tot[48]}  ≥64m={tot[64]}  ≥80m={tot[80]}")

# 两口径核对
print(f"\n两口径：")
print(f"  每栋单算  = {total_rows} 栋（主表行数之和）")
print(f"  建筑群合算 = {_group_total} 个（独立楼 {_indep} + 近距簇 {_n_cluster}，簇成员 {cluster_members}）")

# 与旧值对比
if changes:
    print(f"\n变化（与旧汇总对比）：")
    for c in changes:
        print(c)
else:
    print(f"\n（与旧汇总无变化）")

# 旧合计对比
new_totals_changed = any(tot[t] != OLD_TOTALS[t] for t in [32, 40, 48, 64, 80])
if new_totals_changed:
    print(f"\n合计变化：")
    for t in [32, 40, 48, 64, 80]:
        print(f"  ≥{t}m: {OLD_TOTALS[t]} → {tot[t]} ({'+' if tot[t] > OLD_TOTALS[t] else ''}{tot[t] - OLD_TOTALS[t]})")
