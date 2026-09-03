# -*- coding: utf-8 -*-
"""规范同名合并：距离<40m 且 规范化后同名 → 合并（大小写/标点差异的同一栋楼，之前精确匹配漏网）。
规范化=小写+去非字母数字+去开头the。如 'RITZ CARLTON'=='The Ritz-Carlton'、'Cyber 2 Tower'=='CYBER 2 TOWER'。
保留信息最全+最高的代表点。dry-run默认，--run写回。"""
import sys, os, math, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
R=40.0; SHEET='商业高层(纯净)'

def dist(a,b):return math.hypot((a[0]-b[0])*111000,(a[1]-b[1])*111000*math.cos(math.radians(a[0])))
def unnamed(n): return (not n) or str(n).strip() in('（无名）','(无名)','None','')
def norm(s):
    s=re.sub(r'[^a-z0-9]','',str(s).lower())
    if s.startswith('the'): s=s[3:]
    return s

def process(city, write):
    path=os.path.join(OUT,f'{city}高层建筑清单.xlsx')
    wb=load_workbook(path); ws=wb[SHEET]
    header=[str(c.value).strip() if c.value else '' for c in ws[1]]
    def ci(name): return next(i for i,h in enumerate(header) if name in h)
    li,oi,hi=ci('纬度'),ci('经度'),ci('高度(米)')
    ni,ai,fi=ci('楼宇名称'),ci('建筑面积'),ci('层数')
    ui,gi=ci('用途分类'),ci('建筑全名')
    rows=[[c.value for c in row] for row in ws.iter_rows(min_row=2) if any(c.value is not None for c in row)]
    def num(v):
        try:return float(v)
        except:return None
    def score(r):
        s=0; nm=str(r[ni]) if r[ni] else ''
        if nm and not unnamed(nm):s+=2
        if r[ai] not in(None,'',0):s+=3
        if r[fi] not in(None,''):s+=1
        u=str(r[ui]) if r[ui] else ''
        if u and u!='未分类':s+=1
        g=str(r[gi]) if r[gi] else ''
        if g and g!='—':s+=2
        return s
    # 每点：纬、经、高、规范名、原名、score、idx
    pts=[(num(r[li]),num(r[oi]),num(r[hi]) or 0,
          '' if unnamed(r[ni]) else norm(r[ni]),
          str(r[ni]).strip() if r[ni] else '',score(r),idx) for idx,r in enumerate(rows)]
    # 贪心：按信息完整度+高度降序保留，删除<40m且规范同名的
    order=sorted(range(len(pts)),key=lambda i:(-pts[i][5],-pts[i][2]))
    removed=[False]*len(pts)
    cell=0.0004;grid=defaultdict(list)
    for i,p in enumerate(pts):
        if p[0] is None:continue
        grid[(int(p[0]/cell),int(p[1]/cell))].append(i)
    keep_idx=[]
    merges=[]
    for i in order:
        if removed[i]:continue
        p=pts[i]; keep_idx.append(i)
        if p[0] is None or p[3]=='':continue  # 无名/规范化空 不作合并
        ck=(int(p[0]/cell),int(p[1]/cell))
        for dx in(-1,0,1):
            for dy in(-1,0,1):
                for j in grid.get((ck[0]+dx,ck[1]+dy),[]):
                    if not removed[j] and j!=i:
                        q=pts[j]
                        if q[3]!='' and q[3]==p[3] and dist(p,q)<R:
                            removed[j]=True
                            merges.append((p[4],p[2],q[4],q[2],dist(p,q)))
    keep_idx.sort()
    keep_rows=[rows[i] for i in keep_idx]
    deleted=len(rows)-len(keep_rows)
    if write and deleted>0:
        import shutil
        bak=os.path.join(OUT,'..','backup_规范同名前')
        os.makedirs(bak,exist_ok=True)
        shutil.copy2(path,os.path.join(bak,f'{city}高层建筑清单.xlsx'))
        seq=ci('序号')
        for ridx,rowvals in enumerate(keep_rows, start=2):
            for c,v in enumerate(rowvals, start=1): ws.cell(row=ridx,column=c,value=v)
            ws.cell(row=ridx,column=seq+1,value=ridx-1)
        last=len(keep_rows)+1
        if ws.max_row>last: ws.delete_rows(last+1, ws.max_row-last)
        wb.save(path)
    wb.close()
    return len(rows), len(keep_rows), deleted, merges

def main():
    write='--run' in sys.argv
    print('='*50); print('  规范同名合并(规范同名+<40m) - '+('执行' if write else 'dry-run')); print('='*50)
    t0=t1=0
    for c in CITIES:
        b,a,d,merges=process(c,write); t0+=b;t1+=a
        if d>0:
            print('  %-5s %d → %d (删%d)'%(c,b,a,d))
            for kn,kh,dn,dh,ds in merges:
                print('       保留 %r(%.0fm) ← 删 %r(%.0fm) 距%.0fm'%(kn,kh,dn,dh,ds))
    print('-'*50)
    print('  12城合计: %d → %d (删%d)'%(t0,t1,t0-t1))

main()
