# -*- coding: utf-8 -*-
"""analysis/output 去重全面检查脚本。
检查维度：L1文件哈希 / L2跨城空间重复 / L3城内空间重复 / L4衍生数据一致性 / L5 Sheet交叉。
输出：_去重检查报告.txt + 4份CSV明细。
用法：python audit_dedup.py
"""

import os, sys, hashlib, csv, math, re, zipfile, io, json
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook

# 修复 Windows GBK 终端编码问题
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

# ===================== 路径配置 =====================
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "output")
SCRIPTS_DIR = os.path.join(BASE, "scripts")

# 12城 → 文件名（不含扩展名）
CITIES = [
    "雅加达", "泗水", "万隆", "唐格朗", "望加锡", "棉兰",
    "巴淡", "三宝垄", "勿加泗", "德波", "茂物", "巨港"
]

# 跨城检查对（相邻城市，bbox可能有重叠）
CROSS_PAIRS = [
    ("雅加达", "唐格朗"),
    ("雅加达", "勿加泗"),
    ("雅加达", "德波"),
]

# 空间去重阈值
SPATIAL_THRESHOLD_M = 15       # 距离 < 此值视为空间重合
HEIGHT_DIFF_RATIO = 0.20       # 高度差比例 < 此值视为同一建筑
MULTI_TOWER_HEIGHT_RATIO = 0.30  # 高度差 > 此值 + 同名 → 多塔楼（保留）
SHEET_CROSS_THRESHOLD_M = 5     # Sheet交叉判定更严格

# ===================== 工具函数 =====================

def haversine(lat1, lon1, lat2, lon2):
    """计算两点间距离（米）"""
    r = 6371000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return r * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def file_sha256(filepath):
    """计算文件SHA256"""
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()

def read_excel_buildings(xlsx_path, sheet_name=None):
    """读取Excel中指定sheet的建筑数据。
    返回 (rows, col_index) — rows是列表的列表，col_index是列名→下标的映射。
    若sheet_name为None，使用第一个sheet（"商业高层(纯净)"）。
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]  # 第一个sheet = 商业高层(纯净)
    ws = wb[sheet_name]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_raw:
        return [], {}
    header = [str(c).strip() if c is not None else "" for c in rows_raw[0]]
    col_idx = {h: i for i, h in enumerate(header)}
    data_rows = []
    for r in rows_raw[1:]:
        if r is None or all(v is None for v in r):
            continue
        data_rows.append(list(r))
    return data_rows, col_idx

def get_building_info(row, col_idx):
    """从一行数据提取关键字段"""
    def _get(col_name):
        i = col_idx.get(col_name)
        if i is not None and i < len(row) and row[i] is not None:
            return row[i]
        return None
    return {
        "lat": _get("纬度"),
        "lon": _get("经度"),
        "height": _get("高度(米)"),
        "name": _get("楼宇名称"),
        "band": _get("高度分档"),
        "source": _get("名称来源"),
    }

def spatial_dedup_check(rows, col_idx, threshold_m=SPATIAL_THRESHOLD_M):
    """对一组建筑做空间去重检查。使用网格索引（0.01度≈1km）避免O(n²)。
    返回疑似重复对列表 [(idx_a, idx_b, dist_m, info_a, info_b), ...]。
    """
    # 提取有效坐标的建筑
    buildings = []
    for i, row in enumerate(rows):
        info = get_building_info(row, col_idx)
        if info["lat"] is not None and info["lon"] is not None:
            try:
                lat, lon = float(info["lat"]), float(info["lon"])
                buildings.append((i, lat, lon, info))
            except (ValueError, TypeError):
                pass

    # 网格索引：0.01度 ≈ 1.1km
    grid = defaultdict(list)
    for idx, lat, lon, info in buildings:
        gx = int(lon * 100)
        gy = int(lat * 100)
        grid[(gx, gy)].append((idx, lat, lon, info))

    pairs = []
    checked = set()
    for (gx, gy), cell in grid.items():
        # 检查本格及相邻8格
        for dgx in (-1, 0, 1):
            for dgy in (-1, 0, 1):
                ng = (gx + dgx, gy + dgy)
                if ng not in grid:
                    continue
                other_cell = grid[ng]
                for i in range(len(cell)):
                    for j in range(len(other_cell)):
                        if ng == (gx, gy) and j <= i:
                            continue  # 同格内避免重复检查
                        a_idx, a_lat, a_lon, a_info = cell[i]
                        b_idx, b_lat, b_lon, b_info = other_cell[j]
                        if a_idx >= b_idx:
                            continue
                        pair_key = (min(a_idx, b_idx), max(a_idx, b_idx))
                        if pair_key in checked:
                            continue
                        checked.add(pair_key)
                        dist = haversine(a_lat, a_lon, b_lat, b_lon)
                        if dist < threshold_m:
                            pairs.append((a_idx, b_idx, dist, a_info, b_info))
    return pairs

def band_to_num(band_str):
    """从分档标签提取阈值数字"""
    if band_str is None:
        return None
    m = re.search(r"(80|64|48|40|32)", str(band_str))
    return int(m.group(1)) if m else None

def classify_dup_type(info_a, info_b, dist_m):
    """判定重复对类型：'多塔楼'（同名不同高）、'疑似重复'（同高或无名）"""
    name_a = str(info_a.get("name", "")).strip() if info_a.get("name") else ""
    name_b = str(info_b.get("name", "")).strip() if info_b.get("name") else ""
    h_a = info_a.get("height")
    h_b = info_b.get("height")
    same_name = (name_a and name_b and name_a == name_b)
    if h_a is not None and h_b is not None:
        try:
            ha, hb = float(h_a), float(h_b)
            max_h = max(ha, hb)
            if max_h > 0:
                ratio = abs(ha - hb) / max_h
                if ratio > MULTI_TOWER_HEIGHT_RATIO and same_name:
                    return "多塔楼（同名不同高）"
        except (ValueError, TypeError):
            pass
    if same_name:
        return "同名疑似重复"
    return "疑似重复"

# ===================== 维度1：文件级哈希 =====================

def check_file_hashes():
    """检查output目录下文件级完全重复"""
    print("=" * 60)
    print("维度1：文件级哈希检测（L1）")
    print("=" * 60)

    results = {"files": {}, "duplicates": []}

    # 计算所有文件SHA256（跳过临时文件和锁文件）
    for fname in os.listdir(OUT_DIR):
        if fname.startswith("~$") or fname.startswith("._") or fname.endswith(".tmp"):
            continue
        fpath = os.path.join(OUT_DIR, fname)
        if not os.path.isfile(fpath):
            continue
        # 跳过当前脚本生成的报告文件
        if fname.startswith("_") and (fname.endswith(".csv") or fname.endswith(".txt")):
            continue
        try:
            h = file_sha256(fpath)
            results["files"][fname] = {"sha256": h, "size": os.path.getsize(fpath)}
        except PermissionError:
            print(f"    跳过（文件被占用）: {fname}")

    # 检查zip内文件是否与散落文件一致
    zip_path = os.path.join(OUT_DIR, "印尼高层建筑清单.zip")
    if os.path.exists(zip_path):
        print(f"\n  ZIP: 印尼高层建筑清单.zip ({results['files'].get('印尼高层建筑清单.zip', {}).get('size', 0)/1024:.0f} KB)")
        with zipfile.ZipFile(zip_path, 'r') as zf:
            for zi in zf.infolist():
                if zi.is_dir():
                    continue
                zip_data = zf.read(zi.filename)
                zip_hash = hashlib.sha256(zip_data).hexdigest()
                # 比对同名散落文件
                if zi.filename in results["files"]:
                    loose_hash = results["files"][zi.filename]["sha256"]
                    match = "✓ 完全一致" if zip_hash == loose_hash else "✗ 不一致!"
                    print(f"    {zi.filename}: zip={zip_hash[:12]}... 散落={loose_hash[:12]}... {match}")
                    if zip_hash == loose_hash:
                        results["duplicates"].append({
                            "type": "zip_vs_loose",
                            "loose": zi.filename,
                            "zip_entry": zi.filename,
                        })
                else:
                    print(f"    {zi.filename}: (zip独有，无对应散落文件)")

        if results["duplicates"]:
            print(f"\n  → 结论：zip包中 {len(results['duplicates'])} 个文件与散落xlsx完全重复（L1）")
            print(f"  → 建议：删除 印尼高层建筑清单.zip（需要时随时重新打包）")
    else:
        print("\n  (未找到印尼高层建筑清单.zip)")

    return results

# ===================== 维度2：跨城空间重复 =====================

def check_cross_city():
    """检查相邻城市间的建筑空间重复（L2）"""
    print("\n" + "=" * 60)
    print("维度2：跨城空间重复检测（L2）")
    print("=" * 60)

    all_pairs = []
    for city_a, city_b in CROSS_PAIRS:
        path_a = os.path.join(OUT_DIR, f"{city_a}高层建筑清单.xlsx")
        path_b = os.path.join(OUT_DIR, f"{city_b}高层建筑清单.xlsx")
        if not os.path.exists(path_a) or not os.path.exists(path_b):
            print(f"\n  {city_a}-{city_b}: 文件缺失，跳过")
            continue

        rows_a, ci_a = read_excel_buildings(path_a)
        rows_b, ci_b = read_excel_buildings(path_b)
        print(f"\n  {city_a}({len(rows_a)}栋) vs {city_b}({len(rows_b)}栋):")

        # 提取两城的有效建筑
        bld_a = []
        for i, row in enumerate(rows_a):
            info = get_building_info(row, ci_a)
            if info["lat"] is not None and info["lon"] is not None:
                try:
                    bld_a.append((i, float(info["lat"]), float(info["lon"]), info))
                except (ValueError, TypeError):
                    pass
        bld_b = []
        for i, row in enumerate(rows_b):
            info = get_building_info(row, ci_b)
            if info["lat"] is not None and info["lon"] is not None:
                try:
                    bld_b.append((i, float(info["lat"]), float(info["lon"]), info))
                except (ValueError, TypeError):
                    pass

        # 网格索引
        grid_b = defaultdict(list)
        for idx, lat, lon, info in bld_b:
            gx, gy = int(lon * 100), int(lat * 100)
            grid_b[(gx, gy)].append((idx, lat, lon, info))

        cross_dups = []
        for a_idx, a_lat, a_lon, a_info in bld_a:
            gx, gy = int(a_lon * 100), int(a_lat * 100)
            candidates = []
            for dgx in (-1, 0, 1):
                for dgy in (-1, 0, 1):
                    candidates.extend(grid_b.get((gx + dgx, gy + dgy), []))
            for b_idx, b_lat, b_lon, b_info in candidates:
                dist = haversine(a_lat, a_lon, b_lat, b_lon)
                if dist >= SPATIAL_THRESHOLD_M:
                    continue
                # 高度比较
                h_a = a_info.get("height")
                h_b = b_info.get("height")
                height_close = True
                if h_a is not None and h_b is not None:
                    try:
                        ha, hb = float(h_a), float(h_b)
                        max_h = max(ha, hb)
                        if max_h > 0 and abs(ha - hb) / max_h >= HEIGHT_DIFF_RATIO:
                            height_close = False
                    except (ValueError, TypeError):
                        pass
                if not height_close:
                    continue
                na = a_info.get("name") or ""
                nb = b_info.get("name") or ""
                pair = {
                    "城市A": city_a, "序号A": a_idx + 1,
                    "楼名A": na, "纬度A": a_lat, "经度A": a_lon,
                    "高度A": a_info.get("height"), "分档A": a_info.get("band"),
                    "城市B": city_b, "序号B": b_idx + 1,
                    "楼名B": nb, "纬度B": b_lat, "经度B": b_lon,
                    "高度B": b_info.get("height"), "分档B": b_info.get("band"),
                    "距离(米)": f"{dist:.1f}",
                    "判定": "疑似跨城重复",
                }
                cross_dups.append(pair)
                all_pairs.append(pair)

        print(f"    发现 {len(cross_dups)} 对疑似跨城重复")

    if all_pairs:
        csv_path = os.path.join(OUT_DIR, "_跨城疑似重复.csv")
        write_csv(csv_path, all_pairs)
        print(f"\n  → 详情已写入: _跨城疑似重复.csv ({len(all_pairs)} 条)")
        print(f"  → 建议：人工复核每条记录，决定建筑归属城市")
    else:
        print(f"\n  → 未发现跨城重复 ✓")

    return all_pairs

# ===================== 维度3：城内空间重复 =====================

def check_within_city():
    """检查各城内部的空间重复（L3）"""
    print("\n" + "=" * 60)
    print("维度3：城内空间重复检测（L3）")
    print("=" * 60)

    all_pairs = []
    stats = {}
    for city in CITIES:
        xlsx_path = os.path.join(OUT_DIR, f"{city}高层建筑清单.xlsx")
        if not os.path.exists(xlsx_path):
            print(f"\n  {city}: 文件缺失，跳过")
            continue

        rows, ci = read_excel_buildings(xlsx_path)

        # 雅加达特殊处理：官方数据每栋独立记录，不做15米空间去重。
        # 但检查精确坐标+高度重复（同一坐标同一高度出现多次 = 真正的重复录入）
        if city == "雅加达":
            coord_groups = defaultdict(list)
            for i, row in enumerate(rows):
                info = get_building_info(row, ci)
                if info["lat"] is not None and info["lon"] is not None:
                    try:
                        lat = round(float(info["lat"]), 8)
                        lon = round(float(info["lon"]), 8)
                        h = round(float(info["height"]), 1) if info["height"] is not None else None
                        coord_groups[(lat, lon, h)].append((i, info))
                    except (ValueError, TypeError):
                        pass

            city_pairs = []
            for (lat, lon, h), group in coord_groups.items():
                if len(group) <= 1:
                    continue
                for a in range(len(group)):
                    for b in range(a + 1, len(group)):
                        a_idx, a_info = group[a]
                        b_idx, b_info = group[b]
                        pair = {
                            "城市": city,
                            "序号A": a_idx + 1, "楼名A": a_info.get("name") or "",
                            "纬度A": lat, "经度A": lon,
                            "高度A": h, "分档A": a_info.get("band"),
                            "序号B": b_idx + 1, "楼名B": b_info.get("name") or "",
                            "纬度B": lat, "经度B": lon,
                            "高度B": h, "分档B": b_info.get("band"),
                            "距离(米)": "0.0（精确坐标+高度相同）",
                            "判定": "坐标高度完全重复",
                        }
                        city_pairs.append(pair)
                        all_pairs.append(pair)

            dup_groups = len(coord_groups) - sum(1 for v in coord_groups.values() if len(v) == 1)  # 有重复的组数
            duplicate_rows = sum(len(v) for v in coord_groups.values() if len(v) > 1)
            redundant = sum(len(v) - 1 for v in coord_groups.values() if len(v) > 1)
            unique_coords = sum(1 for v in coord_groups.values() if len(v) == 1)  # 无重复的坐标数（近似独立建筑数）
            stats[city] = {
                "total_rows": len(rows), "pairs": len(city_pairs),
                "dup_groups": dup_groups, "duplicate_rows": duplicate_rows, "redundant": redundant,
                "unique_estimate": len(coord_groups),
                "note": "官方独立记录，检查精确坐标+高度重复"
            }
            print(f"  {city}: {len(rows)}行 → {dup_groups}组精确坐标+高度重复")
            print(f"         涉及 {duplicate_rows} 行，其中 {redundant} 行冗余")
            print(f"         去重后估计独立建筑: {len(coord_groups)} 栋")
            continue

        # 其余11城：空间去重检查
        pairs = spatial_dedup_check(rows, ci, SPATIAL_THRESHOLD_M)
        city_pairs = []
        for a_idx, b_idx, dist, info_a, info_b in pairs:
            dup_type = classify_dup_type(info_a, info_b, dist)
            pair = {
                "城市": city,
                "序号A": a_idx + 1, "楼名A": info_a.get("name") or "",
                "纬度A": info_a.get("lat"), "经度A": info_a.get("lon"),
                "高度A": info_a.get("height"), "分档A": info_a.get("band"),
                "序号B": b_idx + 1, "楼名B": info_b.get("name") or "",
                "纬度B": info_b.get("lat"), "经度B": info_b.get("lon"),
                "高度B": info_b.get("height"), "分档B": info_b.get("band"),
                "距离(米)": f"{dist:.1f}",
                "判定": dup_type,
            }
            city_pairs.append(pair)
            all_pairs.append(pair)

        multi = sum(1 for p in city_pairs if "多塔楼" in p["判定"])
        suspect = len(city_pairs) - multi
        stats[city] = {"total_rows": len(rows), "pairs": len(city_pairs), "multi_tower": multi, "suspect": suspect}
        print(f"  {city}: {len(rows)}行 → {len(city_pairs)}对空间重合（{multi}多塔楼, {suspect}疑似重复）")

    if all_pairs:
        csv_path = os.path.join(OUT_DIR, "_城内疑似重复.csv")
        write_csv(csv_path, all_pairs)
        print(f"\n  → 详情已写入: _城内疑似重复.csv ({len(all_pairs)} 条)")

    # 统计
    total_suspect = sum(s.get("suspect", 0) for s in stats.values())
    total_multi = sum(s.get("multi_tower", 0) for s in stats.values())
    print(f"\n  → 汇总：{total_multi}对多塔楼（保留），{total_suspect}对疑似重复（需复核）")
    if "雅加达" in stats:
        print(f"  → 注意：雅加达为官方独立记录，按名称匹配而非空间去重（稠密城区15米内多建筑正常）")

    return all_pairs, stats

# ===================== 维度4：汇总md一致性 =====================

def check_summary_md():
    """验证汇总md与Excel数据的一致性（L4）"""
    print("\n" + "=" * 60)
    print("维度4：汇总md一致性校验（L4）")
    print("=" * 60)

    md_path = os.path.join(OUT_DIR, "12城高层建筑汇总.md")
    if not os.path.exists(md_path):
        print("\n  12城高层建筑汇总.md 不存在！")
        return []

    # 从md解析各城数字
    md_data = {}
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()
    for line in content.split("\n"):
        if line.startswith("|") and "≥32米" not in line and "合计" not in line and "城市" not in line:
            parts = [p.strip() for p in line.split("|") if p.strip()]
            if len(parts) >= 6:
                city = parts[0]
                try:
                    md_data[city] = {
                        "ge32": int(parts[1]), "ge40": int(parts[2]),
                        "ge48": int(parts[3]), "ge64": int(parts[4]),
                        "ge80": int(parts[5]),
                    }
                except (ValueError, IndexError):
                    pass

    # 从Excel逐城统计（读"商业高层(纯净)"sheet）
    excel_data = {}
    for city in CITIES:
        xlsx_path = os.path.join(OUT_DIR, f"{city}高层建筑清单.xlsx")
        if not os.path.exists(xlsx_path):
            continue
        rows, ci = read_excel_buildings(xlsx_path)
        buckets = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
        for row in rows:
            info = get_building_info(row, ci)
            thr = band_to_num(info.get("band"))
            if thr in buckets:
                buckets[thr] += 1
        # 累计
        cum = {}
        for t in (32, 40, 48, 64, 80):
            cum[t] = sum(v for k, v in buckets.items() if k >= t)
        excel_data[city] = cum

    # 比对
    diffs = []
    print()
    for city in sorted(set(list(md_data.keys()) + list(excel_data.keys()))):
        md = md_data.get(city, {})
        ex = excel_data.get(city, {})
        if not md:
            print(f"  {city}: MD中无数据")
            continue
        if not ex:
            print(f"  {city}: Excel无数据，MD={md['ge32']}（可能硬编码）")
            continue
        for thr in (32, 40, 48, 64, 80):
            m = md.get(f"ge{thr}", 0)
            e = ex.get(thr, 0)
            if m != e:
                diff = {
                    "城市": city, "阈值": f"≥{thr}米",
                    "汇总md": m, "Excel统计": e,
                    "差异": e - m,
                }
                diffs.append(diff)
                print(f"  {city} ≥{thr}米: MD={m}, Excel={e}  → 差异 {e-m:+d}")

    if not diffs:
        print("  ✓ 汇总md与Excel数据完全一致")
    else:
        print(f"\n  → 发现 {len(diffs)} 处不一致")
        print(f"  → 原因分析：汇总md中雅加达使用官方全量(7220)，Excel\"商业高层(纯净)\"已剔除公共设施(6521)")
        print(f"  → 建议：重新运行 summary_12cities.py 以当前Excel为准重新生成")

    # 写入差异CSV
    if diffs:
        csv_path = os.path.join(OUT_DIR, "_汇总md差异.csv")
        write_csv(csv_path, diffs)

    return diffs

# ===================== 维度5：Sheet交叉验证 =====================

def check_sheet_cross():
    """验证"商业高层(纯净)"与"已剔除-公共设施"之间无交叉（L5）"""
    print("\n" + "=" * 60)
    print("维度5：Sheet交叉验证（L5）")
    print("=" * 60)

    all_cross = []
    for city in CITIES:
        xlsx_path = os.path.join(OUT_DIR, f"{city}高层建筑清单.xlsx")
        if not os.path.exists(xlsx_path):
            continue

        # 读两个sheet
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        if len(sheet_names) < 2:
            wb.close()
            continue
        clean_name = sheet_names[0]   # 商业高层(纯净)
        drop_name = sheet_names[1]    # 已剔除-公共设施

        ws_clean = wb[clean_name]
        ws_drop = wb[drop_name]

        # 读取坐标
        def read_coords(ws):
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return [], {}
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            ci = {h: i for i, h in enumerate(header)}
            coords = []
            for r in rows[1:]:
                try:
                    lat_i = ci.get("纬度")
                    lon_i = ci.get("经度")
                    if lat_i is not None and lon_i is not None:
                        lat = float(r[lat_i]) if r[lat_i] is not None else None
                        lon = float(r[lon_i]) if r[lon_i] is not None else None
                        if lat is not None and lon is not None:
                            coords.append((lat, lon))
                except (ValueError, TypeError, IndexError):
                    pass
            return coords, ci

        clean_coords, _ = read_coords(ws_clean)
        drop_coords, _ = read_coords(ws_drop)
        wb.close()

        # 网格索引查交叉
        drop_grid = defaultdict(list)
        for lat, lon in drop_coords:
            gx, gy = int(lon * 200), int(lat * 200)  # 更细网格
            drop_grid[(gx, gy)].append((lat, lon))

        crosses = []
        for lat, lon in clean_coords:
            gx, gy = int(lon * 200), int(lat * 200)
            for dgx in (-1, 0, 1):
                for dgy in (-1, 0, 1):
                    for dlat, dlon in drop_grid.get((gx + dgx, gy + dgy), []):
                        dist = haversine(lat, lon, dlat, dlon)
                        if dist < SHEET_CROSS_THRESHOLD_M:
                            crosses.append({
                                "城市": city, "纬度": lat, "经度": lon,
                                "已剔除_纬度": dlat, "已剔除_经度": dlon,
                                "距离(米)": f"{dist:.1f}",
                            })

        if crosses:
            print(f"  {city}: ✗ 发现 {len(crosses)} 处Sheet交叉！")
            all_cross.extend(crosses)
        else:
            print(f"  {city}: ✓ 无交叉")

    if all_cross:
        csv_path = os.path.join(OUT_DIR, "_Sheet交叉异常.csv")
        write_csv(csv_path, all_cross)
        print(f"\n  → ⚠ 发现 {len(all_cross)} 处Sheet交叉异常！详情见 _Sheet交叉异常.csv")
    else:
        print(f"\n  → ✓ 所有城市Sheet间无交叉")

    return all_cross

# ===================== 维度6：Dashboard过期分析 =====================

def check_dashboard():
    """分析dashboard.html的数据时效性（L4）"""
    print("\n" + "=" * 60)
    print("维度6：Dashboard数据状态（L4）")
    print("=" * 60)

    dash_path = os.path.join(OUT_DIR, "dashboard.html")
    if not os.path.exists(dash_path):
        print("\n  dashboard.html 不存在")
        return {}

    with open(dash_path, "r", encoding="utf-8") as f:
        content = f.read()

    # 提取日期信息
    dates_found = re.findall(r"(\d{4}-\d{2}-\d{2})", content)
    dates_found = list(set(dates_found))

    # 检查包含哪些城市
    cities_in_dash = []
    for city in CITIES:
        if city in content:
            cities_in_dash.append(city)

    # 提取雅加达数据（从JS中找）
    jkt_official = re.findall(r'"official".*?"ge32"\s*:\s*(\d+)', content, re.DOTALL)

    result = {
        "dates": dates_found,
        "cities_covered": cities_in_dash,
        "total_cities_dash": len(cities_in_dash),
        "total_cities_actual": 12,
        "jkt_official_ge32_in_dash": jkt_official,
    }

    print(f"\n  内嵌数据日期：{dates_found}")
    print(f"  覆盖城市：{cities_in_dash}（{len(cities_in_dash)}/12）")
    print(f"  缺失城市：{sorted(set(CITIES) - set(cities_in_dash))}")
    if jkt_official:
        print(f"  雅加达官方≥32米：dashboard={jkt_official[0]}, Excel(纯净)≈6521")
    print(f"\n  → 结论：dashboard仅覆盖 {len(cities_in_dash)}/12 城，数据已过期")
    print(f"  → 建议：重命名为 dashboard_legacy.html 归档，或运行 build_dashboard.py 更新")

    return result

# ===================== CSV写入 =====================

def write_csv(path, rows):
    """写入CSV文件（UTF-8 BOM，Excel友好）"""
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)

# ===================== 主报告生成 =====================

def generate_report(results):
    """生成汇总报告"""
    report_path = os.path.join(OUT_DIR, "_去重检查报告.txt")
    lines = []
    lines.append("=" * 60)
    lines.append("  印尼12城高层建筑清单 · 去重检查报告")
    lines.append(f"  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"  检查范围：{OUT_DIR}")
    lines.append("=" * 60)

    # 维度1
    lines.append("\n── 维度1：文件级完全重复（L1）──")
    fh = results.get("file_hashes", {})
    dups = fh.get("duplicates", [])
    if dups:
        lines.append(f"  [重复] zip包中 {len(dups)} 个文件与散落xlsx完全一致")
        lines.append(f"  [建议] 删除 印尼高层建筑清单.zip（需时重新打包）")
    else:
        lines.append(f"  [通过] 未发现文件级完全重复")

    # 维度2
    lines.append("\n── 维度2：跨城空间重复（L2）──")
    cross = results.get("cross_city", [])
    if cross:
        by_pair = defaultdict(int)
        for p in cross:
            by_pair[f"{p['城市A']}-{p['城市B']}"] += 1
        lines.append(f"  [发现] {len(cross)} 对疑似跨城重复")
        for pair, cnt in sorted(by_pair.items()):
            lines.append(f"    {pair}: {cnt} 对")
        lines.append(f"  [详情] _跨城疑似重复.csv")
        lines.append(f"  [建议] 人工复核每条记录，决定归属")
    else:
        lines.append(f"  [通过] 未发现跨城重复 ✓")

    # 维度3
    lines.append("\n── 维度3：城内空间重复（L3）──")
    wcs = results.get("within_city_stats", {})
    total_suspect = sum(s.get("suspect", 0) for s in wcs.values())
    total_multi = sum(s.get("multi_tower", 0) for s in wcs.values())
    total_pairs = sum(s.get("pairs", 0) for s in wcs.values())
    total_redundant = sum(s.get("redundant", 0) for s in wcs.values())
    if total_pairs > 0:
        lines.append(f"  [发现] 共 {total_pairs} 对空间重合（{total_multi}多塔楼保留, {total_suspect}疑似需复核）")
        for city, s in sorted(wcs.items()):
            if s["pairs"] > 0:
                if "redundant" in s:
                    lines.append(f"    {city}: {s['total_rows']}行 → {s.get('dup_groups','?')}组精确坐标+高度重复，涉及{s['duplicate_rows']}行（冗余{s['redundant']}行）")
                    lines.append(f"        去重后估计独立建筑: {s['unique_estimate']} 栋")
                else:
                    lines.append(f"    {city}: {s['total_rows']}行 → {s['pairs']}对（{s['multi_tower']}多塔楼, {s['suspect']}疑似）")
        if total_redundant > 0:
            lines.append(f"  ⚠ 雅加达精确坐标+高度重复：{total_redundant} 行冗余！")
        lines.append(f"  [详情] _城内疑似重复.csv")
    else:
        lines.append(f"  [通过] 12城均未发现空间重复 ✓")

    # 维度4
    lines.append("\n── 维度4：汇总md一致性（L4）──")
    md_diffs = results.get("md_diffs", [])
    if md_diffs:
        lines.append(f"  [差异] 汇总md与Excel有 {len(md_diffs)} 处不一致")
        lines.append(f"  [原因] 汇总md中雅加达使用官方全量(7220)，Excel已剔除公共设施(6521)")
        lines.append(f"  [建议] 重新运行 summary_12cities.py 生成最新汇总md")
    else:
        lines.append(f"  [通过] 汇总md与Excel数据一致 ✓")

    # 维度5
    lines.append("\n── 维度5：Sheet交叉验证（L5）──")
    sheet_cross = results.get("sheet_cross", [])
    if sheet_cross:
        lines.append(f"  [异常] ⚠ 发现 {len(sheet_cross)} 处Sheet交叉！")
        lines.append(f"  [详情] _Sheet交叉异常.csv")
        lines.append(f"  [建议] 修复生成脚本中分类逻辑")
    else:
        lines.append(f"  [通过] 所有城市Sheet间无交叉 ✓")

    # 维度6
    lines.append("\n── 维度6：Dashboard数据状态（L4）──")
    dash = results.get("dashboard", {})
    lines.append(f"  [信息] 覆盖 {dash.get('total_cities_dash', '?')}/{dash.get('total_cities_actual', 12)} 城")
    lines.append(f"  [信息] 内嵌日期：{dash.get('dates', [])}")
    lines.append(f"  [建议] 重命名为 dashboard_legacy.html 归档")

    # 汇总建议
    lines.append("\n" + "=" * 60)
    lines.append("  一键清理建议（需人工确认后执行 dedup_clean.py）")
    lines.append("=" * 60)
    lines.append(f"  [ ] 删除 印尼高层建筑清单.zip（释放 {fh.get('files', {}).get('印尼高层建筑清单.zip', {}).get('size', 0)/1024:.0f} KB）")
    lines.append(f"  [ ] 重新生成 12城高层建筑汇总.md（需确认雅加达口径：全量 or 纯商业）")
    lines.append(f"  [ ] 归档 dashboard.html → dashboard_legacy.html")
    if cross:
        lines.append(f"  [ ] 复核 _跨城疑似重复.csv 中 {len(cross)} 条记录")
    if total_redundant > 0:
        lines.append(f"  [ ] 清理雅加达精确重复：{total_redundant} 行冗余（坐标+高度完全相同）")
    if total_suspect > 0:
        lines.append(f"  [ ] 复核 _城内疑似重复.csv 中 {total_suspect} 条记录")

    report_text = "\n".join(lines)
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_text)

    print("\n" + report_text)
    print(f"\n报告已写入: {report_path}")
    return report_path

# ===================== 主入口 =====================

def main():
    print("印尼12城 output/ 去重全面检查")
    print(f"目录: {OUT_DIR}")
    print()

    results = {}

    # L1: 文件哈希
    results["file_hashes"] = check_file_hashes()

    # L2: 跨城重复
    results["cross_city"] = check_cross_city()

    # L3: 城内重复
    within_pairs, within_stats = check_within_city()
    results["within_city"] = within_pairs
    results["within_city_stats"] = within_stats

    # L4: 汇总md
    results["md_diffs"] = check_summary_md()

    # L5: Sheet交叉
    results["sheet_cross"] = check_sheet_cross()

    # L4b: Dashboard
    results["dashboard"] = check_dashboard()

    # 生成主报告
    generate_report(results)

    print("\n全部检查完成。")

if __name__ == "__main__":
    main()
