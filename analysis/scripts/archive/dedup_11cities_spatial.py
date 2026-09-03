# -*- coding: utf-8 -*-
"""11城空间去重(R=25m + 高度差<25%)，合并Google ML把同一栋楼拆成的多个点。
与雅加达同口径：贪心，按信息完整度+高度降序保留代表点，删除半径内高度接近的拆点。
雅加达不在此脚本(已单独去重)。dry-run默认，--run写回。"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
CITIES=['泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
R=25.0; HPCT=0.25; SHEET='商业高层(纯净)'

def dist(a,b):return math.hypot((a[0]-b[0])*111000,(a[1]-b[1])*111000*math.cos(math.radians(a[0])))

def process(city, write):
    path=os.path.join(OUT,f'{city}高层建筑清单.xlsx')
    wb=load_workbook(path)
    ws=wb[SHEET]
    header=[str(c.value).strip() if c.value else '' for c in ws[1]]
    def ci(name): return next(i for i,h in enumerate(header) if name in h)
    li,oi,hi=ci('纬度'),ci('经度'),ci('高度(米)')
    ni,ai,fi=ci('楼宇名称'),ci('建筑面积'),ci('层数')
    ui,gi=ci('用途分类'),ci('建筑全名')
    rows=[[c.value for c in row] for row in ws.iter_rows(min_row=2) if any(c.value is not None for c in row)]
    def num(v):
        try:return float(v)
        except:return None
    def score(r):
        s=0; nm=str(r[ni]) if r[ni] else ''
        if nm and nm not in('（无名）','(无名)','None',''):s+=2
        if r[ai] not in(None,'',0):s+=3
        if r[fi] not in(None,''):s+=1
        u=str(r[ui]) if r[ui] else ''
        if u and u!='未分类':s+=1
        g=str(r[gi]) if r[gi] else ''
        if g and g!='—':s+=2
        return s
    pts=[(num(r[li]),num(r[oi]),num(r[hi]) or 0,score(r),idx) for idx,r in enumerate(rows)]
    order=sorted(range(len(pts)),key=lambda i:(-pts[i][3],-pts[i][2]))
    removed=[False]*len(pts)
    cell=0.0003;grid=defaultdict(list)
    for i,p in enumerate(pts):
        if p[0] is None:continue
        grid[(int(p[0]/cell),int(p[1]/cell))].append(i)
    keep_idx=[]
    for i in order:
        if removed[i]:continue
        p=pts[i]
        if p[0] is None:keep_idx.append(i);continue
        keep_idx.append(i)
        ck=(int(p[0]/cell),int(p[1]/cell))
        for dx in(-1,0,1):
            for dy in(-1,0,1):
                for j in grid.get((ck[0]+dx,ck[1]+dy),[]):
                    if not removed[j] and j!=i:
                        q=pts[j];m=max(p[2],q[2])
                        if dist(p,q)<R and (m==0 or abs(p[2]-q[2])/m<HPCT):
                            removed[j]=True
    keep_idx.sort()
    keep_rows=[rows[i] for i in keep_idx]
    deleted=len(rows)-len(keep_rows)
    if write:
        import shutil
        bak=os.path.join(OUT,'..','backup_11城去重前')
        os.makedirs(bak,exist_ok=True)
        shutil.copy2(path,os.path.join(bak,f'{city}高层建筑清单.xlsx'))
        seq=ci('序号')
        for ridx,rowvals in enumerate(keep_rows, start=2):
            for c,v in enumerate(rowvals, start=1): ws.cell(row=ridx,column=c,value=v)
            ws.cell(row=ridx,column=seq+1,value=ridx-1)
        last=len(keep_rows)+1
        if ws.max_row>last: ws.delete_rows(last+1, ws.max_row-last)
        wb.save(path)
    wb.close()
    return len(rows), len(keep_rows), deleted

def main():
    write='--run' in sys.argv
    print('='*54); print('  11城空间去重(25m+高度差<25%) - '+('执行' if write else 'dry-run')); print('='*54)
    t0=t1=0
    for c in CITIES:
        b,a,d=process(c,write)
        t0+=b;t1+=a
        print('  %-5s %d → %d  (删%d)'%(c,b,a,d))
    print('-'*54)
    print('  11城合计: %d → %d (删%d)'%(t0,t1,t0-t1))

main()
