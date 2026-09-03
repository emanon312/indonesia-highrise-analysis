# -*- coding: utf-8 -*-
"""更新说明sheet里「名称来源 与 数据来源」段落，同步细化后的口径。dry-run默认，--run执行。"""
import sys, os, copy
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
TITLE='【名称来源 与 数据来源'

def jkt_lines():
    return [
     ('t','【名称来源 与 数据来源 —— 两列区别，勿混淆】'),
     ('b','  · 名称来源：这栋楼的“名字”是从哪匹配到的 —— 官方 / OSM补充(≤50米) / OSM补充(存疑,50~100米) / 无。'),
     ('b','  · 数据来源：这栋楼的“本体数据”（经纬度、高度、层数、建筑面积、轮廓）来自哪个源 —— 本城全部为『官方(DPMPTSP)』官方三维库，真值。'),
     ('b','  · 两列关系：雅加达数据来源恒为官方；名称来源多为官方，仅原本无名的楼才用 OSM 补名，故两列偶有不同。'),
    ]
def gen_lines(n):
    base=[
     ('t','【名称来源 与 数据来源 —— 两列区别，勿混淆】'),
     ('b','  · 名称来源：这栋楼的“名字”是从哪匹配到的 —— CTBUH地标 / OSM补充(≤40米) / 无。'),
    ]
    if n>0:
        base += [
     ('b','  · 数据来源：这栋楼的“本体数据”（经纬度、高度、层数、建筑面积、轮廓）来自哪个源，本城有两种取值：'),
     ('b','       - GoogleV3坐标+2.5D高度：多数楼，高度为 Google 2.5D 机器学习估算（偏保守，系统性偏低）。'),
     ('b','       - GoogleV3坐标+CTBUH真高：本城 %d 栋地标楼，坐标仍是 Google V3，但高度为 CTBUH 收录的实测真值（更准）。'%n),
        ]
    else:
        base += [
     ('b','  · 数据来源：这栋楼的“本体数据”（经纬度、高度、层数、建筑面积、轮廓）来自哪个源 —— 本城全部为『GoogleV3坐标+2.5D高度』（无 CTBUH 地标真高楼）。'),
        ]
    base += [('b','  · 两列可不同：Google 只提供坐标和高度、不提供楼名，楼名需另去 OSM/CTBUH 匹配。例：数据来源=GoogleV3坐标+2.5D高度，但名字来自 OSM → 名称来源=OSM补充。')]
    return base

def run_city(city, write):
    path=os.path.join(OUT,city+'高层建筑清单.xlsx')
    wb=load_workbook(path); ws=wb['说明']
    # 统计该城CTBUH真高数(读主表+剔除表数据来源列)
    n=0
    for sn in ['商业高层(纯净)','已剔除-公共设施']:
        w=ws.parent[sn]; hh=[str(c.value) for c in w[1]]; di=hh.index('数据来源')
        for row in w.iter_rows(min_row=2,values_only=True):
            if row[di]=='GoogleV3坐标+CTBUH真高': n+=1
    lines = jkt_lines() if city=='雅加达' else gen_lines(n)
    # 样式模板
    t_font=b_font=None
    for r in range(1,ws.max_row+1):
        v=ws.cell(row=r,column=1).value
        if v is None: continue
        s=str(v)
        if t_font is None and s.startswith('【'): t_font=copy.copy(ws.cell(row=r,column=1).font)
        if b_font is None and not s.startswith('【') and r>1: b_font=copy.copy(ws.cell(row=r,column=1).font)
    # 找旧段落起始
    start=None
    for r in range(1,ws.max_row+1):
        v=ws.cell(row=r,column=1).value
        if v and str(v).startswith(TITLE): start=r; break
    if not write:
        print('%-5s CTBUH真高=%d 旧段落起于第%s行 → 重写%d行'%(city,n,start,len(lines)))
        wb.close(); return
    if start:
        ws.delete_rows(start, ws.max_row-start+1)
    else:
        start=ws.max_row+2
    for i,(kind,text) in enumerate(lines):
        c=ws.cell(row=start+i,column=1,value=text)
        f=t_font if kind=='t' else b_font
        if f is not None: c.font=f
    wb.save(path); wb.close()
    print('%-5s CTBUH真高=%d 已重写说明(%d行)'%(city,n,len(lines)))

def main():
    write='--run' in sys.argv
    print('  更新说明段落 - '+('执行' if write else 'dry-run'))
    for c in CITIES: run_city(c,write)

main()
