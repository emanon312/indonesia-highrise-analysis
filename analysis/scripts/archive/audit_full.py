# -*- coding: utf-8 -*-
"""全面复核脚本（只读）—— 输出完整审计报告
检查维度：
  A. 文件清单（含bak、临时文件）
  B. Excel结构（sheet数量/名称/列头）
  C. 数据行数（Sheet1纯净 vs Sheet2剔除 vs Sheet3说明）
  D. 高度分档一致性（分档累计 = 行数）
  E. 汇总MD vs Excel 交叉验证
  F. 已知问题（空行、跨城残余、~$锁文件）
  G. 跨城重复残余（可能重复31对的状态）
"""
import os, re, csv, sys
from datetime import datetime
from openpyxl import load_workbook
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output")
REPORT = os.path.join(OUT, "_完整复核报告.txt")

CITIES = ["雅加达","泗水","万隆","唐格朗","望加锡","棉兰","巴淡","三宝垄","勿加泗","德波","茂物","巨港"]
EXPECTED_SHEETS = ["商业高层(纯净)", "已剔除-公共设施", "说明"]

# ===================== A. 文件清单 =====================
def audit_files():
    lines = []
    lines.append("=" * 60)
    lines.append("A. 文件清单")
    lines.append("=" * 60)

    all_files = {}
    for f in os.listdir(OUT):
        full = os.path.join(OUT, f)
        if os.path.isfile(full):
            size = os.path.getsize(full)
            all_files[f] = size

    # 分类
    xlsx_files = {k:v for k,v in all_files.items() if k.endswith('.xlsx') and not k.startswith('~')}
    bak_files = {k:v for k,v in all_files.items() if k.endswith('.bak')}
    temp_files = {k:v for k,v in all_files.items() if k.startswith('~')}
    csv_files = {k:v for k,v in all_files.items() if k.endswith('.csv')}
    md_files = {k:v for k,v in all_files.items() if k.endswith('.md')}
    txt_files = {k:v for k,v in all_files.items() if k.endswith('.txt')}
    html_files = {k:v for k,v in all_files.items() if k.endswith('.html')}
    py_files = {k:v for k,v in all_files.items() if k.endswith('.py')}
    other = {k:v for k,v in all_files.items() if k not in {**xlsx_files, **bak_files, **temp_files, **csv_files, **md_files, **txt_files, **html_files, **py_files}}

    lines.append(f"\n总计文件: {len(all_files)} 个, {sum(all_files.values())/1024:.0f} KB")
    lines.append(f"  Excel (.xlsx):    {len(xlsx_files):>3} 个  {sum(xlsx_files.values())/1024:.0f} KB")
    lines.append(f"  备份 (.bak):      {len(bak_files):>3} 个  {sum(bak_files.values())/1024:.0f} KB")
    lines.append(f"  临时 (~$):        {len(temp_files):>3} 个  {sum(temp_files.values())/1024:.0f} KB")
    lines.append(f"  CSV:              {len(csv_files):>3} 个  {sum(csv_files.values())/1024:.0f} KB")
    lines.append(f"  Markdown (.md):   {len(md_files):>3} 个  {sum(md_files.values())/1024:.0f} KB")
    lines.append(f"  Text (.txt):      {len(txt_files):>3} 个  {sum(txt_files.values())/1024:.0f} KB")
    lines.append(f"  HTML:             {len(html_files):>3} 个  {sum(html_files.values())/1024:.0f} KB")
    lines.append(f"  Python (.py):     {len(py_files):>3} 个  {sum(py_files.values())/1024:.0f} KB")

    # Excel 清单
    lines.append(f"\nExcel 明细:")
    for f in sorted(xlsx_files.keys()):
        size_kb = xlsx_files[f] / 1024
        city = f.replace("高层建筑清单.xlsx", "")
        lines.append(f"  {city:<6} {size_kb:>8.0f} KB")

    # 异常文件
    issues = []
    if temp_files:
        for f in temp_files:
            issues.append(f"  [异常] 临时锁文件: {f} (Excel未关闭)")
    if py_files:
        for f in py_files:
            issues.append(f"  [注意] Python脚本在output目录: {f}")
    if bak_files:
        for f in bak_files:
            issues.append(f"  [注意] 备份文件: {f}")

    if issues:
        lines.append(f"\n文件问题:")
        for i in issues:
            lines.append(i)

    # 检查12城完整性
    present = set()
    for f in xlsx_files:
        for c in CITIES:
            if c in f:
                present.add(c)
    missing = set(CITIES) - present
    if missing:
        lines.append(f"\n  [严重] 缺失城市: {missing}")
    else:
        lines.append(f"\n  12城 Excel 齐全 ✓")

    return lines, len(issues)

# ===================== B/C/D. Excel逐文件审计 =====================
def audit_excel():
    lines = []
    lines.append("\n" + "=" * 60)
    lines.append("B/C/D. Excel 结构 + 数据行数 + 分档一致性")
    lines.append("=" * 60)

    all_ok = True
    issues = []
    summary_check = {}  # {city: {32: N, 40: N, ...}}

    for city in CITIES:
        xlsx = os.path.join(OUT, f"{city}高层建筑清单.xlsx")
        if not os.path.exists(xlsx):
            lines.append(f"\n  [缺失] {city}")
            continue

        lines.append(f"\n--- {city} ---")
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        sheets = wb.sheetnames

        # B. Sheet 结构
        if len(sheets) != 3:
            lines.append(f"  [结构异常] sheet数={len(sheets)} (预期3), sheets={sheets}")
            issues.append(f"{city}: sheet数量={len(sheets)}")
        else:
            lines.append(f"  Sheet: {sheets[0]} | {sheets[1]} | {sheets[2]}")

        # C. 各Sheet行数
        for sn in sheets:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            data_rows = len(rows) - 1 if rows else 0  # -header

            # 检查空行
            empty_count = 0
            for r in rows[1:]:
                if all(v is None for v in r):
                    empty_count += 1

            extra = ""
            if empty_count > 0:
                extra = f" [含{empty_count}空行!]"
                issues.append(f"{city}/{sn}: {empty_count} 空行")
            lines.append(f"  {sn}: {data_rows} 行{extra}")

            # 检查列头（仅 Sheet1）
            if sn == sheets[0] and rows:
                header = [str(c).strip() if c else "" for c in rows[0]]
                expected_cols = ["序号", "楼宇名称", "名称来源", "OSM匹配距离", "建筑全名",
                                 "纬度", "经度", "高度", "层数", "高度分档", "用途分类",
                                 "行政区", "街道办区", "地址", "建筑面积"]
                found = sum(1 for ec in expected_cols if any(ec in h for h in header))
                lines.append(f"  列数: {len(header)}, 关键列匹配: {found}/{len(expected_cols)}")

        # D. 分档一致性（仅 Sheet1）
        ws1 = wb[sheets[0]]
        rows1 = list(ws1.iter_rows(values_only=True))
        if rows1:
            header = [str(c).strip() if c else "" for c in rows1[0]]
            band_col = next((i for i, h in enumerate(header) if "分档" in h), None)

            if band_col is not None:
                buckets = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
                band_total = 0
                for r in rows1[1:]:
                    if r[band_col] is None:
                        continue
                    band_str = str(r[band_col])
                    m = re.search(r"(80|64|48|40|32)", band_str)
                    if m:
                        thr = int(m.group(1))
                        if thr in buckets:
                            buckets[thr] += 1
                            band_total += 1

                cum = {t: sum(v for k, v in buckets.items() if k >= t) for t in (32, 40, 48, 64, 80)}
                summary_check[city] = cum

                data_rows = len([r for r in rows1[1:] if not all(v is None for v in r)])
                lines.append(f"  分档: >=32m={cum[32]} >=40m={cum[40]} >=48m={cum[48]} >=64m={cum[64]} >=80m={cum[80]}")

                if cum[32] != data_rows:
                    lines.append(f"  [分档不一致!] cum[32]={cum[32]} vs 有效行={data_rows}")
                    issues.append(f"{city}: 分档累计{cum[32]} != 有效行{data_rows}")

        wb.close()

    return lines, issues, summary_check

# ===================== E. 汇总MD交叉验证 =====================
def audit_summary(summary_check):
    lines = []
    lines.append("\n" + "=" * 60)
    lines.append("E. 汇总MD vs Excel 交叉验证")
    lines.append("=" * 60)

    md_path = os.path.join(OUT, "12城高层建筑汇总.md")
    if not os.path.exists(md_path):
        lines.append("  [缺失] 12城高层建筑汇总.md")
        return lines, 1

    # 解析 MD 表格
    md_nums = {}
    with open(md_path, "r", encoding="utf-8") as f:
        content = f.read()

    for line in content.split("\n"):
        # | 城市 | >=32m | >=40m | ...
        if line.startswith("|") and not line.startswith("|--") and not line.startswith("| 城市") and not line.startswith("| **"):
            parts = [p.strip() for p in line.split("|") if p.strip()]
            if len(parts) >= 6:
                city_name = parts[0]
                try:
                    nums = {32: int(parts[1]), 40: int(parts[2]), 48: int(parts[3]),
                            64: int(parts[4]), 80: int(parts[5])}
                    md_nums[city_name] = nums
                except (ValueError, IndexError):
                    pass

    issues = 0
    lines.append(f"\nMD中城市数: {len(md_nums)}")

    for city in CITIES:
        if city not in summary_check:
            continue
        if city not in md_nums:
            lines.append(f"  [缺失] {city} 不在汇总MD中")
            issues += 1
            continue

        xl = summary_check[city]
        md = md_nums[city]
        for thr in [32, 40, 48, 64, 80]:
            if xl[thr] != md[thr]:
                lines.append(f"  [不一致] {city} >={thr}m: Excel={xl[thr]} MD={md[thr]} (差{md[thr]-xl[thr]})")
                issues += 1

    if issues == 0:
        lines.append("  所有城市 MD 与 Excel 一致 ✓")

    # 合计验证
    if "合计" in md_nums:
        md_total = md_nums["合计"]
        xl_total = {t: sum(summary_check.get(c, {}).get(t, 0) for c in CITIES) for t in [32,40,48,64,80]}
        for thr in [32, 40, 48, 64, 80]:
            if xl_total[thr] != md_total[thr]:
                lines.append(f"  [合计不一致] >={thr}m: Excel合计={xl_total[thr]} MD合计={md_total[thr]}")
                issues += 1

    return lines, issues

# ===================== F. 已知问题汇总 =====================
def audit_known_issues():
    lines = []
    lines.append("\n" + "=" * 60)
    lines.append("F. 已知问题盘点")
    lines.append("=" * 60)

    issues = 0

    # 1. 空行问题
    lines.append("\n[问题1] 跨城去重残留空行")
    lines.append("  唐格朗: 8 空行 (已去重但未缩行)")
    lines.append("  勿加泗: 3 空行 (同上)")
    lines.append("  修复: remove_cross_city_dup.py 已加 delete_rows(), 需重跑")
    issues += 1

    # 2. ~$ 临时文件
    temp_files = [f for f in os.listdir(OUT) if f.startswith("~$")]
    if temp_files:
        lines.append(f"\n[问题2] Excel 锁文件: {temp_files}")
        lines.append("  影响: 该文件被 Excel 打开, 脚本可能写入失败")
        issues += 1

    # 3. bak 备份文件
    bak_files = [f for f in os.listdir(OUT) if f.endswith('.bak')]
    if bak_files:
        lines.append(f"\n[问题3] 残留备份: {bak_files}")
        lines.append("  建议: 确认数据无误后可删除")

    # 4. Python脚本在output目录
    py_files = [f for f in os.listdir(OUT) if f.endswith('.py')]
    if py_files:
        lines.append(f"\n[问题4] 脚本混入output目录: {py_files}")
        issues += 1

    # 5. 汇总md口径
    lines.append(f"\n[问题5] 雅加达汇总口径")
    lines.append("  当前MD: 5694 (纯商业高层, 已剔除780公共设施)")
    lines.append("  原始全量: 7220 条官方记录")
    lines.append("  口径选择: 仅统计剔除后的商业高层——合理")

    return lines, issues

# ===================== G. 跨城残余 =====================
def audit_cross_city():
    lines = []
    lines.append("\n" + "=" * 60)
    lines.append("G. 跨城重复残余")
    lines.append("=" * 60)

    csv_path = os.path.join(OUT, "_跨城重复分析报告.csv")
    if not os.path.exists(csv_path):
        lines.append("  报告CSV缺失")
        return lines, 1

    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # 统计
    confirmed = [r for r in rows if r["判定"] == "确定重复"]
    possible = [r for r in rows if r["判定"] == "可能重复"]
    not_dup = [r for r in rows if "不是重复" in r["判定"]]

    lines.append(f"\n总对数: {len(rows)}")
    lines.append(f"  确定重复: {len(confirmed)} 对 → 已删除唐格朗/勿加泗侧 (但有残留空行)")
    lines.append(f"  可能重复: {len(possible)} 对 → 未处理")
    lines.append(f"  非重复:   {len(not_dup)} 对 → 保留")

    # 可能重复中需要人工判的
    manual = [r for r in possible if "需人工判断" in r.get("建议归属", "")]
    lines.append(f"\n可能重复中需人工判断: {len(manual)} 对")
    for r in manual:
        lines.append(f"  {r['楼名A_完整'][:30]} vs {r['楼名B_完整'][:30]} ({r['城市A']}-{r['城市B']}) 距离{r['距离米']}m")

    return lines, len(manual)

# ===================== 主入口 =====================
def main():
    all_lines = []
    all_lines.append(f"完整复核报告")
    all_lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    all_lines.append(f"目录: {OUT}")
    all_lines.append("")

    total_issues = 0

    # A
    lines, n = audit_files()
    all_lines.extend(lines)
    total_issues += n

    # B/C/D
    lines, n, summary_check = audit_excel()
    all_lines.extend(lines)
    total_issues += len([x for x in n if "空行" not in x])  # 空行归类到F

    # E
    lines, n = audit_summary(summary_check)
    all_lines.extend(lines)
    total_issues += n

    # F
    lines, n = audit_known_issues()
    all_lines.extend(lines)
    total_issues += n

    # G
    lines, n = audit_cross_city()
    all_lines.extend(lines)
    total_issues += n

    # 总结
    all_lines.append("\n" + "=" * 60)
    all_lines.append("总结")
    all_lines.append("=" * 60)
    all_lines.append(f"发现问题: {total_issues} 项")

    # 写入报告
    report = "\n".join(all_lines)
    with open(REPORT, "w", encoding="utf-8") as f:
        f.write(report)

    # 同时输出到屏幕
    print(report)

if __name__ == "__main__":
    main()
