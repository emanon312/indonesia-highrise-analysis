# -*- coding: utf-8 -*-
"""细化「数据来源」列，写全混合来源。dry-run默认，--run执行。
取值：
  雅加达            → 官方(DPMPTSP)
  11城·CTBUH真高楼  → GoogleV3坐标+CTBUH真高   (名称来源=CTBUH地标 且 高度真匹配到CTBUH条目)
  11城·其余         → GoogleV3坐标+2.5D高度
"""
import sys, os, json, math
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

HERE=os.path.dirname(os.path.abspath(__file__))
OUT=os.path.join(HERE,'..','output'); DATA=os.path.join(HERE,'..','data')
EN={'泗水':'surabaya','万隆':'bandung','唐格朗':'tangerang','望加锡':'makassar','棉兰':'medan','巴淡':'batam','三宝垄':'semarang','勿加泗':'bekasi','德波':'depok','茂物':'bogor','巨港':'palembang'}
SHEETS=['商业高层(纯净)','已剔除-公共设施']
V_JKT='官方(DPMPTSP)'; V_CTBUH='GoogleV3坐标+CTBUH真高'; V_GOOGLE='GoogleV3坐标+2.5D高度'
LOG=[]
def log(s): LOG.append(s); print(s)

def ctbuh_realh_set(city, sheets_rows):
    """返回该城高度真来自CTBUH的楼坐标集合(round7)"""
    ct=os.path.join(DATA,'%s_ctbuh.json'%EN[city])
    if not os.path.exists(ct): return set()
    lms=[l for l in json.load(open(ct,encoding='utf-8')) if l.get('height_m') and l.get('lat') is not None]
    s=set()
    for rows,si,li,oi in sheets_rows:
        for r in rows:
            if str(r[si])!='CTBUH地标': continue
            la,lo=r[li],r[oi]
            if la is None or lo is None: continue
            bd=1e18;best=None
            for lm in lms:
                dlat=(la-lm['lat'])*111000; dlon=(lo-lm['lon'])*111000*math.cos(math.radians(lm['lat']))
                d2=dlat*dlat+dlon*dlon
                if d2<bd: bd=d2;best=lm
            if best and bd<=150*150: s.add((round(la,7),round(lo,7)))
    return s

def run_city(city, write):
    path=os.path.join(OUT,city+'高层建筑清单.xlsx')
    wb=load_workbook(path)
    # 先收集判定用信息
    sheets_rows=[]
    meta={}
    for sn in SHEETS:
        ws=wb[sn]
        header=[str(c.value).strip() if c.value else '' for c in ws[1]]
        si=header.index('名称来源'); li=header.index('纬度'); oi=header.index('经度'); di=header.index('数据来源')
        rows=[[c.value for c in row] for row in ws.iter_rows(min_row=2) if any(c.value is not None for c in row)]
        meta[sn]=(ws,si,li,oi,di,rows)
        sheets_rows.append((rows,si,li,oi))
    realh = ctbuh_realh_set(city, sheets_rows) if city in EN else set()
    cnt={V_JKT:0,V_CTBUH:0,V_GOOGLE:0}
    for sn in SHEETS:
        ws,si,li,oi,di,rows=meta[sn]
        for ridx,r in enumerate(rows, start=2):
            if city=='雅加达': val=V_JKT
            else:
                la,lo=r[li],r[oi]
                val = V_CTBUH if (la is not None and (round(la,7),round(lo,7)) in realh) else V_GOOGLE
            cnt[val]+=1
            if write: ws.cell(row=ridx, column=di+1, value=val)
    if write: wb.save(path)
    wb.close()
    tag = V_JKT if city=='雅加达' else ('CTBUH真高=%d 其余Google=%d'%(cnt[V_CTBUH],cnt[V_GOOGLE]))
    log('%-5s %s'%(city, (V_JKT+' 全表%d行'%cnt[V_JKT]) if city=='雅加达' else tag))
    return cnt[V_CTBUH]

def main():
    write='--run' in sys.argv
    log('='*56); log('  细化数据来源列 - '+('执行' if write else 'dry-run')); log('='*56)
    tot=0
    for c in ['雅加达']+list(EN.keys()): tot+=run_city(c,write)
    log('-'*56); log('  合计标记CTBUH真高: %d 栋'%tot)
    open(os.path.join(OUT,'_数据来源细化日志.txt'),'w',encoding='utf-8').write('\n'.join(LOG))

main()
