# -*- coding: utf-8 -*-
"""跨城去重：从唐格朗/勿加泗 Excel 中删除确认重复的建筑行
读取 _跨城重复分析报告.csv，筛选 判定=确定重复 且 建议归属=雅加达，
按坐标+高度匹配对应城市 Excel 中的行并删除。
"""
import os, csv
from openpyxl import load_workbook
from copy import copy

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "..", "output")
CSV_PATH = os.path.join(OUT, "_跨城重复分析报告.csv")

def find_col(headers, keywords):
    for i, h in enumerate(headers):
        if h and all(kw in h for kw in keywords):
            return i
    raise ValueError(f"找不到列: {keywords}")

def main():
    # 1) 读取 CSV，筛选确认重复
    to_remove = {}  # {城市B: [(纬度B, 经度B, 高度B, 楼名A, 楼名B), ...]}
    with open(CSV_PATH, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["判定"] != "确定重复":
                continue
            if row["建议归属"] != "雅加达":
                continue
            city_b = row["城市B"]
            lat_b = float(row["纬度B"])
            lon_b = float(row["经度B"])
            h_b = float(row["高度B"])
            name_a = row["楼名A_完整"]
            name_b = row["楼名B_完整"] or row["楼名B_CSV"]
            if city_b not in to_remove:
                to_remove[city_b] = []
            to_remove[city_b].append((lat_b, lon_b, h_b, name_a, name_b))

    print(f"待删除: {sum(len(v) for v in to_remove.values())} 条")
    for city, items in to_remove.items():
        print(f"  {city}: {len(items)} 条")
        for lat, lon, h, na, nb in items:
            print(f"    {nb} ({h}m) <- 重复于 {na}")

    # 2) 逐城市处理
    for city, items in to_remove.items():
        xlsx_path = os.path.join(OUT, f"{city}高层建筑清单.xlsx")
        bak_path = xlsx_path + ".bak"
        if not os.path.exists(xlsx_path):
            print(f"\n[!] {city} Excel 不存在，跳过")
            continue

        print(f"\n处理 {city}...")
        wb = load_workbook(xlsx_path)
        ws = wb[wb.sheetnames[0]]  # 第一个 sheet = 商业高层(纯净)

        # 读 header
        rows_data = list(ws.iter_rows(values_only=True))
        header = [str(c).strip() if c is not None else "" for c in rows_data[0]]

        lat_col = find_col(header, ["纬度"])
        lon_col = find_col(header, ["经度"])
        h_col = find_col(header, ["高度", "米"])

        # 构建查找集合 (round8 lat, round8 lon, round1 height)
        target_keys = set()
        for lat, lon, h, _, _ in items:
            key = (round(lat, 8), round(lon, 8), round(h, 1))
            target_keys.add(key)

        # 筛选保留行
        keep_indices = [0]  # header 始终保留 (1-based in Excel)
        removed_count = 0
        for i, row in enumerate(rows_data[1:], start=2):  # Excel row numbers (1-based)
            lat_val = float(row[lat_col]) if row[lat_col] is not None else 0
            lon_val = float(row[lon_col]) if row[lon_col] is not None else 0
            h_val = float(row[h_col]) if row[h_col] is not None else 0
            key = (round(lat_val, 8), round(lon_val, 8), round(h_val, 1))
            if key in target_keys:
                name = row[1] if len(row) > 1 else ""
                print(f"  X 删除 行{i}: {name} ({h_val}m) @ ({lat_val}, {lon_val})")
                removed_count += 1
            else:
                keep_indices.append(i)

        print(f"  删除 {removed_count} 行，保留 {len(keep_indices)-1} 行")

        if removed_count == 0:
            wb.close()
            print(f"  [!] 未匹配到任何行，可能坐标已变化，跳过")
            continue

        # 重建 sheet：清空后逐行写入保留行（保留样式较难，直接清空重写）
        # 先备份
        wb.save(bak_path)
        print(f"  已备份: {bak_path}")

        # 由于 openpyxl 不方便直接删行，采用：清空 sheet → 重写保留行
        # 收集保留行的数据
        keep_rows = [rows_data[0]]  # header
        for idx in keep_indices[1:]:  # skip header index
            keep_rows.append(rows_data[idx - 1])  # rows_data is 0-based

        # 清空 sheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # 重写保留行
        for r_idx, row_data in enumerate(keep_rows, start=1):
            for c_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # 重新编号序号列（第1列）
        seq_col = 1  # 序号通常是 A 列
        for r_idx in range(2, len(keep_rows) + 1):
            ws.cell(row=r_idx, column=seq_col, value=r_idx - 1)

        # 删除多余空行（清空不会自动缩 max_row）
        if ws.max_row > len(keep_rows):
            ws.delete_rows(len(keep_rows) + 1, ws.max_row - len(keep_rows))

        wb.save(xlsx_path)
        wb.close()
        print(f"  [OK] 已保存: {xlsx_path} (保留 {len(keep_rows)-1} 行)")

    print("\n" + "=" * 50)
    print("  跨城去重完成")
    print("=" * 50)

if __name__ == "__main__":
    main()
