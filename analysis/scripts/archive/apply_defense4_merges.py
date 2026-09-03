# -*- coding: utf-8 -*-
"""
雅加达高层建筑防线④聚簇合并脚本
对"去重待复核"sheet 中的防线④聚簇执行自动合并判定

合并判定规则:
  应合并（满足任一）:
    1. 高度波动<0.5m 且 层数一致 且 聚簇直径<30m
    2. 高度波动<1m 且 层数一致 且 聚簇直径<50m
    3. 楼宇名称="（无名）"（所有无名建筑紧密聚簇都是重复录入）
  应保留:
    高度波动>=5m 或 层数不一致（>=2种不同层数）
  不处理: 其余情况保持不动（标记为"待人工"）

评分函数: 与 dedup_jakarta.py 完全一致
  建筑面积(3分) + 楼宇名称(2分) + 建筑全名(2分) + 用途分类(2分)
  + 地址(1分) + 层数(1分) + 行政区(1分) + 街道办区(1分)
  + 名称来源(1分) + OSM匹配距离(1分) = 15分
"""

import os
import shutil
import re
from datetime import datetime
from math import radians, sin, cos, sqrt, asin
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ===================== 路径配置 =====================
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT = os.path.join(BASE, "output", "雅加达高层建筑清单.xlsx")
BACKUP = INPUT + ".bak2"


# ===================== 工具函数 =====================

def find_col(headers, keywords):
    """根据关键词列表查找列索引（0-based），用关键词匹配而非硬编码列号"""
    for i, h in enumerate(headers):
        if h and all(kw in h for kw in keywords):
            return i
    raise ValueError(f"无法找到包含以下关键词的列: {keywords}")


def has_value(val, exclude_set=None):
    """判断单元格值是否有意义（非空、非排除值）"""
    if val is None:
        return False
    if isinstance(val, str) and val.strip() == '':
        return False
    if exclude_set is not None and val in exclude_set:
        return False
    return True


def safe_float(val):
    """安全转浮点数"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """安全转整数"""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def haversine(lat1, lon1, lat2, lon2):
    """Haversine 距离（米）"""
    R = 6371000
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat / 2) ** 2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2
    return R * 2 * asin(sqrt(a))


def normalize_name(name):
    """标准化名称：去空格、转小写、去特殊字符"""
    if name is None:
        return ""
    n = str(name).strip().lower()
    n = re.sub(r'[^a-z0-9一-鿿]', '', n)
    return n


def score_row(row, col_map):
    """评分函数：满分15分，与 dedup_jakarta.py 完全一致"""
    score = 0
    if has_value(row[col_map['area']]):
        score += 3
    if has_value(row[col_map['name']], exclude_set={'（无名）', '(无名)'}):
        score += 2
    if has_value(row[col_map['fullname']], exclude_set={'—'}):
        score += 2
    if has_value(row[col_map['use']], exclude_set={'未分类'}):
        score += 2
    if has_value(row[col_map['addr']], exclude_set={'—'}):
        score += 1
    if has_value(row[col_map['floors']]):
        score += 1
    if has_value(row[col_map['district']]):
        score += 1
    if has_value(row[col_map['subdistrict']]):
        score += 1
    if has_value(row[col_map['source']], exclude_set={'无'}):
        score += 1
    if has_value(row[col_map['osm_dist']], exclude_set={'—'}):
        score += 1
    return score


# ===================== 主流程 =====================

def main():
    start_time = datetime.now()
    print("=" * 70)
    print("  雅加达高层建筑防线④聚簇自动合并")
    print(f"  时间：{start_time.strftime('%Y-%m-%d %H:%M')}")
    print("=" * 70)

    # ---- 步骤1：备份 ----
    if os.path.exists(BACKUP):
        os.remove(BACKUP)
    shutil.copy2(INPUT, BACKUP)
    print(f"\n[1] 备份完成: {os.path.basename(BACKUP)}")

    # ---- 步骤2：加载工作簿 ----
    wb = load_workbook(INPUT)

    sheet_names = wb.sheetnames
    print(f"\n[2] 工作簿包含 {len(sheet_names)} 个Sheet: {sheet_names}")

    ws_main = wb[sheet_names[0]]   # 商业高层(纯净)
    ws_drop = wb[sheet_names[1]]   # 已剔除-公共设施
    ws_note = wb[sheet_names[2]]   # 说明

    review_sheet_name = "去重待复核"
    if review_sheet_name not in sheet_names:
        print(f"\n[错误] 未找到Sheet「{review_sheet_name}」，请先运行 dedup_jakarta.py")
        wb.close()
        return
    ws_review = wb[review_sheet_name]

    # ---- 步骤3：读取 Sheet1 表头，建立列索引映射 ----
    header_cells = list(ws_main[1])
    headers_main = [c.value if c.value is not None else "" for c in header_cells]
    print(f"\n[3] Sheet1 表头 ({len(headers_main)} 列):")
    for i, h in enumerate(headers_main):
        print(f"      col{i+1}: {h}")

    col_map = {
        'seq':        find_col(headers_main, ['序']),
        'name':       find_col(headers_main, ['楼宇', '名称']),
        'source':     find_col(headers_main, ['名称', '来源']),
        'osm_dist':   find_col(headers_main, ['OSM']),
        'fullname':   find_col(headers_main, ['建筑', '全名']),
        'lat':        find_col(headers_main, ['纬']),
        'lon':        find_col(headers_main, ['经']),
        'height':     find_col(headers_main, ['高', '米']),
        'floors':     find_col(headers_main, ['层']),
        'band':       find_col(headers_main, ['分档']),
        'use':        find_col(headers_main, ['用途', '分类']),
        'district':   find_col(headers_main, ['行政']),
        'subdistrict': find_col(headers_main, ['街道']),
        'addr':       find_col(headers_main, ['地址']),
        'area':       find_col(headers_main, ['建筑', '面积']),
    }
    print(f"\n    列映射完成，{len(col_map)} 个字段")

    # ---- 步骤4：读取 Sheet1 所有数据行 ----
    data_rows = []
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, values_only=True):
        if all(v is None for v in row):
            continue
        data_rows.append(list(row))

    total_original = len(data_rows)
    print(f"\n[4] Sheet1「{sheet_names[0]}」数据行: {total_original}")

    # ---- 步骤5：构建 Sheet1 索引，用于匹配去重待复核中的记录 ----
    # 使用 (归一化名称, 纬度round4, 经度round4) 作为匹配键
    # round4 精度约 11m，可容忍去重前后的坐标微小漂移
    s1_index = {}  # key -> data_rows 中的索引
    for idx, row in enumerate(data_rows):
        name = row[col_map['name']]
        lat = safe_float(row[col_map['lat']])
        lon = safe_float(row[col_map['lon']])
        if name and lat is not None and lon is not None:
            key = (normalize_name(name), round(lat, 4), round(lon, 4))
            # 如果同一个key有多个匹配（去重后不应出现），保留先出现的
            if key not in s1_index:
                s1_index[key] = idx

    print(f"    Sheet1 索引条目: {len(s1_index)}")

    # ---- 步骤6：读取「去重待复核」sheet，按聚簇ID分组 ----
    clusters = defaultdict(list)

    for r in range(2, ws_review.max_row + 1):
        cid = ws_review.cell(row=r, column=1).value
        if not cid:
            continue
        name = ws_review.cell(row=r, column=2).value
        lat = safe_float(ws_review.cell(row=r, column=6).value)
        lon = safe_float(ws_review.cell(row=r, column=7).value)
        diameter = safe_float(ws_review.cell(row=r, column=4).value)

        # 尝试匹配 Sheet1 中的行
        s1_idx = None
        if name and lat is not None and lon is not None:
            key = (normalize_name(name), round(lat, 4), round(lon, 4))
            s1_idx = s1_index.get(key, None)

        clusters[cid].append({
            'review_row': r,
            'name': name,
            's1_idx': s1_idx,
            'diameter': diameter,
        })

    total_clusters = len(clusters)
    total_review_records = sum(len(v) for v in clusters.values())
    print(f"\n[5] 「去重待复核」聚簇: {total_clusters} 组, {total_review_records} 条记录")

    # 统计匹配情况
    matched_records = sum(1 for recs in clusters.values() for r in recs if r['s1_idx'] is not None)
    print(f"    匹配到Sheet1: {matched_records}/{total_review_records} "
          f"({matched_records / total_review_records * 100:.1f}%)")

    # ---- 步骤7：对每个聚簇执行自动判定 ----
    print(f"\n[6] 逐聚簇执行自动判定...")

    stats = {'merged': 0, 'kept': 0, 'manual': 0}
    to_delete = set()          # 待删除的 data_rows 索引
    merge_details = {}         # {cluster_id: {result, reason, keep_s1_idx, delete_count, ...}}
    cluster_results = {}       # {cluster_id: result_string} 用于填回Sheet4

    for cid in sorted(clusters.keys()):
        recs = clusters[cid]

        # 只考虑能匹配到 Sheet1 的记录
        found_recs = [r for r in recs if r['s1_idx'] is not None]
        # 去重：同一个 s1_idx 可能被多条记录匹配到（round4碰撞）
        seen_idx = set()
        unique_found = []
        for r in found_recs:
            if r['s1_idx'] not in seen_idx:
                seen_idx.add(r['s1_idx'])
                unique_found.append(r)
        found_recs = unique_found

        if len(found_recs) < 2:
            # 记录数不足以合并，标记为待人工（含匹配不足的情况）
            cluster_results[cid] = '待人工判断（匹配记录不足，无法自动判定）'
            merge_details[cid] = {
                'action': 'manual',
                'reason': '匹配记录数<2' if len(found_recs) < 2 else '无匹配',
                'found_count': len(found_recs),
            }
            stats['manual'] += 1
            continue

        # ---- 计算聚簇属性 ----
        heights = []
        floors_set = set()
        names_raw = []

        for rec in found_recs:
            row = data_rows[rec['s1_idx']]
            h = safe_float(row[col_map['height']])
            f = safe_int(row[col_map['floors']])
            n = row[col_map['name']]
            if h is not None:
                heights.append(h)
            if f is not None:
                floors_set.add(f)
            names_raw.append(str(n) if n else '')

        h_range = max(heights) - min(heights) if len(heights) >= 2 else 0.0
        floors_consistent = len(floors_set) == 1
        all_unnamed = all(n == '（无名）' for n in names_raw)
        diameter = found_recs[0]['diameter']  # 同一聚簇所有行直径相同

        # ---- 判定逻辑 ----
        action = None   # 'merge', 'keep', 'manual'
        reason = ''

        # 应保留规则（优先级最高）
        if h_range >= 5:
            action = 'keep'
            reason = f'高度波动{h_range:.1f}m >= 5m，疑似不同塔楼'
        elif not floors_consistent and len(floors_set) >= 2:
            action = 'keep'
            reason = f'层数不一致（{sorted(floors_set)}），疑似不同塔楼'
        # 应合并 - 规则1
        elif h_range < 0.5 and floors_consistent and diameter is not None and diameter < 30:
            action = 'merge'
            reason = (f'规则1: 高度波动{h_range:.2f}m<0.5m + 层数一致({list(floors_set)[0]}层)'
                      f' + 直径{diameter:.0f}m<30m')
        # 应合并 - 规则2
        elif h_range < 1 and floors_consistent and diameter is not None and diameter < 50:
            action = 'merge'
            reason = (f'规则2: 高度波动{h_range:.2f}m<1m + 层数一致({list(floors_set)[0]}层)'
                      f' + 直径{diameter:.0f}m<50m')
        # 应合并 - 规则3
        elif all_unnamed:
            action = 'merge'
            reason = f'规则3: 全部{len(found_recs)}条记录为"（无名）"，紧密聚簇即重复录入'
        else:
            action = 'manual'
            # 构造待人工原因
            parts = []
            if h_range >= 0.5:
                parts.append(f'高度波动{h_range:.2f}m')
            if not floors_consistent:
                parts.append(f'层数不一致({len(floors_set)}种)')
            if diameter is not None and diameter >= 30:
                parts.append(f'直径{diameter:.0f}m>=30m')
            if not parts:
                parts.append('不满足任一自动合并条件')
            reason = '; '.join(parts)

        # ---- 执行动作 ----
        if action == 'merge':
            # 对聚簇内所有记录评分
            scored = []
            for rec in found_recs:
                row = data_rows[rec['s1_idx']]
                s = score_row(row, col_map)
                area_val = safe_float(row[col_map['area']])
                area = area_val if area_val is not None else 0.0
                scored.append((rec['s1_idx'], s, area))

            # 排序：分数降序，面积降序，索引升序
            scored.sort(key=lambda x: (-x[1], -x[2], x[0]))
            keep_idx = scored[0][0]
            keep_score = scored[0][1]
            delete_indices = [x[0] for x in scored[1:]]

            # 标记删除（暂存，后面统一处理）
            for di in delete_indices:
                to_delete.add(di)

            merge_details[cid] = {
                'action': 'merge',
                'reason': reason,
                'keep_s1_idx': keep_idx,
                'keep_score': keep_score,
                'delete_indices': delete_indices,
                'found_count': len(found_recs),
            }
            stats['merged'] += 1

        elif action == 'keep':
            merge_details[cid] = {
                'action': 'keep',
                'reason': reason,
                'found_count': len(found_recs),
            }
            stats['kept'] += 1

        else:  # action == 'manual'
            merge_details[cid] = {
                'action': 'manual',
                'reason': reason,
                'found_count': len(found_recs),
            }
            stats['manual'] += 1

    print(f"    应合并: {stats['merged']} 组")
    print(f"    应保留: {stats['kept']} 组")
    print(f"    待人工: {stats['manual']} 组")

    # ---- 步骤8：计算删除后新的序号（用于填写判定结果） ----
    # 构建索引映射: 原始 s1_idx -> 新序号（从1开始）
    kept_flags = [i for i in range(len(data_rows)) if i not in to_delete]
    new_seq_map = {}
    for new_pos, old_idx in enumerate(kept_flags):
        new_seq_map[old_idx] = new_pos + 1  # 序号从1开始

    # ---- 步骤9：从 Sheet1 删除标记行 ----
    print(f"\n[7] 从 Sheet1 删除标记行: {len(to_delete)} 行")

    kept_rows = [data_rows[i] for i in range(len(data_rows)) if i not in to_delete]

    # 清空并重写 Sheet1 数据区
    if ws_main.max_row > 1:
        ws_main.delete_rows(2, ws_main.max_row - 1)

    for i, row_data in enumerate(kept_rows):
        for j, val in enumerate(row_data):
            ws_main.cell(row=i + 2, column=j + 1, value=val)

    # 重新编序号
    seq_col_idx = col_map['seq'] + 1  # openpyxl 列号从1开始
    for i in range(len(kept_rows)):
        ws_main.cell(row=i + 2, column=seq_col_idx, value=i + 1)

    total_deleted = len(to_delete)
    print(f"    Sheet1: {total_original} → {len(kept_rows)} 行 (删除 {total_deleted} 行)")

    # ---- 步骤10：生成每个聚簇的判定结果字符串（含新序号） ----
    for cid, detail in merge_details.items():
        if detail['action'] == 'merge':
            keep_idx = detail['keep_s1_idx']
            new_seq = new_seq_map.get(keep_idx, '?')
            cluster_results[cid] = (
                f"已合并-保留序号{new_seq}（{detail['reason']}，"
                f"{detail['found_count']}条→1条，保留评分{detail['keep_score']}分）"
            )
        elif detail['action'] == 'keep':
            cluster_results[cid] = f"已确认保留-不同塔楼（{detail['reason']}）"
        elif detail['action'] == 'manual':
            cluster_results[cid] = f"待人工判断（{detail['reason']}）"

    # ---- 步骤11：更新 Sheet4「去重待复核」 ----
    print(f"\n[8] 更新 Sheet「去重待复核」...")

    # 检查是否已有"判定结果"列
    result_col_idx = None
    for c in range(1, ws_review.max_column + 1):
        header_val = ws_review.cell(row=1, column=c).value
        if header_val and '判定结果' in str(header_val):
            result_col_idx = c
            break

    if result_col_idx is None:
        # 新增"判定结果"列（第17列）
        result_col_idx = ws_review.max_column + 1
        header_cell = ws_review.cell(row=1, column=result_col_idx, value='判定结果')
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 为每行填入判定结果（按聚簇ID）
    for cid, recs in clusters.items():
        result_str = cluster_results.get(cid, '未判定')
        for rec in recs:
            ws_review.cell(row=rec['review_row'], column=result_col_idx, value=result_str)

    # 调整新列宽度
    from openpyxl.utils import get_column_letter
    ws_review.column_dimensions[get_column_letter(result_col_idx)].width = 55

    print(f"    判定结果列: 第{result_col_idx}列")

    # ---- 步骤12：打印合并前后对比（关键聚簇） ----
    print(f"\n[9] 关键聚簇合并前后对比:")
    print("-" * 70)

    merged_clusters = [(cid, d) for cid, d in merge_details.items() if d['action'] == 'merge']
    # 按删除数量降序，展示前10个
    merged_clusters.sort(key=lambda x: -len(x[1].get('delete_indices', [])))

    for rank, (cid, detail) in enumerate(merged_clusters[:10]):
        recs = clusters[cid]
        found = [r for r in recs if r['s1_idx'] is not None]
        keep_idx = detail['keep_s1_idx']
        keep_row = data_rows[keep_idx]
        keep_name = keep_row[col_map['name']]
        keep_height = safe_float(keep_row[col_map['height']])
        keep_floors = safe_int(keep_row[col_map['floors']])

        delete_count = len(detail.get('delete_indices', []))
        new_seq = new_seq_map.get(keep_idx, '?')

        print(f"\n  [{rank + 1}] {cid} | {keep_name}")
        print(f"       {detail['reason']}")
        print(f"       保留: 序号{new_seq} 高度{keep_height}m 层数{keep_floors}层 评分{detail['keep_score']}分")
        print(f"       删除: {delete_count} 条 → 合并为1条")

        # 打印被删除记录的关键信息
        for di in detail.get('delete_indices', [])[:5]:
            drow = data_rows[di]
            dh = safe_float(drow[col_map['height']])
            df = safe_int(drow[col_map['floors']])
            ds = score_row(drow, col_map)
            print(f"         - 原序号{di + 1} 高度{dh}m 层数{df}层 评分{ds}分")
        if delete_count > 5:
            print(f"         ... 等共{delete_count}条")

    # 也应展示几个应保留的聚簇
    print(f"\n  --- 应保留聚簇示例 ---")
    kept_clusters = [(cid, d) for cid, d in merge_details.items() if d['action'] == 'keep']
    for cid, detail in kept_clusters[:5]:
        recs = clusters[cid]
        found = [r for r in recs if r['s1_idx'] is not None]
        name = found[0]['name'] if found else '?'
        print(f"  {cid} | {name} | {detail['reason']} | {detail['found_count']}条保留")

    # ---- 步骤13：更新 Sheet3「说明」 ----
    print(f"\n[10] 更新「说明」sheet...")

    last_row = ws_note.max_row
    # 找到最后一个有内容的行
    while last_row > 0 and ws_note.cell(row=last_row, column=1).value is None:
        last_row -= 1
    if last_row == 0:
        last_row = 0
    nr = last_row + 2

    ws_note.cell(row=nr, column=1, value="【防线④聚簇自动合并记录】")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    合并时间：{start_time.strftime('%Y-%m-%d %H:%M')}")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="    合并策略（自动判定防线④去重待复核聚簇）：")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      应合并规则1: 高度波动<0.5m + 层数一致 + 聚簇直径<30m")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      应合并规则2: 高度波动<1m + 层数一致 + 聚簇直径<50m")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      应合并规则3: 楼宇名称全部为\"（无名）\"，紧密聚簇即重复录入")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      应保留: 高度波动>=5m 或 层数不一致(>=2种不同层数) → 确认不同塔楼")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      待人工: 其余情况保持不动，标记为待人工判断")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    合并统计：共 {total_clusters} 组聚簇 → "
                       f"合并 {stats['merged']} 组（删除 {total_deleted} 行），"
                       f"保留 {stats['kept']} 组，"
                       f"待人工 {stats['manual']} 组")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    Sheet1 变化：{total_original} 行 → {len(kept_rows)} 行 "
                       f"（删除 {total_deleted} 行）")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="    评分标准（选优依据）：建筑面积(3分) + 楼宇名称(2分) + 建筑全名(2分) + "
                       "用途分类(2分) + 地址(1分) + 层数(1分) + 行政区(1分) + "
                       "街道办区(1分) + 名称来源(1分) + OSM匹配距离(1分) = 15分")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    备份文件: {os.path.basename(BACKUP)}")

    # ---- 步骤14：保存 ----
    print(f"\n[11] 保存文件...")
    wb.save(INPUT)
    wb.close()

    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"\n{'=' * 70}")
    print(f"  防线④聚簇合并完成！（耗时 {elapsed:.1f}s）")
    print(f"  Sheet1: {total_original} → {len(kept_rows)} 行 (删除 {total_deleted} 行)")
    print(f"  文件: {os.path.basename(INPUT)}")
    print(f"  备份: {os.path.basename(BACKUP)}")
    print(f"\n  合并统计:")
    print(f"    聚簇总数: {total_clusters}")
    print(f"    已合并:   {stats['merged']} 组（自动合并为1条）")
    print(f"    已保留:   {stats['kept']} 组（确认不同塔楼）")
    print(f"    待人工:   {stats['manual']} 组（需人工逐组判断）")
    print(f"    删除行数: {total_deleted}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
