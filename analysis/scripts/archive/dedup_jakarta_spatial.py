# -*- coding: utf-8 -*-
"""雅加达空间去重(R=25m + 高度差<25%)，去除官方多条重复登记。仅雅加达，11城不动。
贪心：按信息完整度+高度降序，保留代表点，删除其半径内高度接近的重复。
dry-run默认，--run写回。"""
import sys, os, math
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
PATH=os.path.join(OUT,'雅加达高层建筑清单.xlsx')
R=25.0; HPCT=0.25; SHEET='商业高层(纯净)'

def dist(a,b):return math.hypot((a[0]-b[0])*111000,(a[1]-b[1])*111000*math.cos(math.radians(a[0])))

def main():
    write='--run' in sys.argv
    wb=load_workbook(PATH)
    ws=wb[SHEET]
    header=[str(c.value).strip() if c.value else '' for c in ws[1]]
    def ci(name): return next(i for i,h in enumerate(header) if name in h)
    li,oi,hi=ci('纬度'),ci('经度'),ci('高度(米)')
    ni,ai,fi=ci('楼宇名称'),ci('建筑面积'),ci('层数')
    ui,gi=ci('用途分类'),ci('建筑全名')
    # 读全部数据行(保留整行值)
    rows=[[c.value for c in row] for row in ws.iter_rows(min_row=2) if any(c.value is not None for c in row)]
    def num(v):
        try:return float(v)
        except:return None
    def score(r):
        s=0
        nm=str(r[ni]) if r[ni] else ''
        if nm and nm not in('（无名）','(无名)','None',''):s+=2
        if r[ai] not in(None,'',0):s+=3
        if r[fi] not in(None,''):s+=1
        u=str(r[ui]) if r[ui] else ''
        if u and u!='未分类':s+=1
        g=str(r[gi]) if r[gi] else ''
        if g and g!='—':s+=2
        return s
    pts=[(num(r[li]),num(r[oi]),num(r[hi]) or 0,score(r),idx) for idx,r in enumerate(rows)]
    # 贪心: 按(score,height)降序保留，删R内高度接近
    order=sorted(range(len(pts)),key=lambda i:(-pts[i][3],-pts[i][2]))
    removed=[False]*len(pts)
    cell=0.0003;grid=defaultdict(list)
    for i,p in enumerate(pts):
        if p[0] is None:continue
        grid[(int(p[0]/cell),int(p[1]/cell))].append(i)
    keep_idx=[]
    for i in order:
        if removed[i]:continue
        p=pts[i]
        if p[0] is None:keep_idx.append(i);continue
        keep_idx.append(i)
        ck=(int(p[0]/cell),int(p[1]/cell))
        for dx in(-1,0,1):
            for dy in(-1,0,1):
                for j in grid.get((ck[0]+dx,ck[1]+dy),[]):
                    if not removed[j] and j!=i:
                        q=pts[j];m=max(p[2],q[2])
                        if dist(p,q)<R and (m==0 or abs(p[2]-q[2])/m<HPCT):
                            removed[j]=True
    keep_idx.sort()  # 恢复原序
    keep_rows=[rows[i] for i in keep_idx]
    print('雅加达 %s: %d → %d  (删%d重复, %.0f%%)'%(SHEET,len(rows),len(keep_rows),
          len(rows)-len(keep_rows),100*(len(rows)-len(keep_rows))/len(rows)))
    if not write:
        print('(dry-run，未写回)');wb.close();return
    # 备份
    import shutil
    bak=os.path.join(OUT,'..','backup_雅加达空间去重前')
    os.makedirs(bak,exist_ok=True)
    shutil.copy2(PATH,os.path.join(bak,'雅加达高层建筑清单.xlsx'))
    print('已备份 →', bak)
    # 重写主表: 清空数据行→写keep→重编号→删多余
    seq=ci('序号')
    for ridx,rowvals in enumerate(keep_rows, start=2):
        for c,v in enumerate(rowvals, start=1): ws.cell(row=ridx,column=c,value=v)
        ws.cell(row=ridx,column=seq+1,value=ridx-1)
    last=len(keep_rows)+1
    if ws.max_row>last: ws.delete_rows(last+1, ws.max_row-last)
    wb.save(PATH);wb.close()
    print('已写回主表 %d 行'%len(keep_rows))

main()
