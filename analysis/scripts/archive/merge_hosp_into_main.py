# -*- coding: utf-8 -*-
"""把剔除表里的医院行并入主表「商业高层(纯净)」。dry-run默认，--run执行。
- 按列名映射（主表15列 vs 剔除表15列，列顺序不同）
- 用途分类填「医院」，高度分档按高度算，雅加达面积从源JSON回填
- 主表按高度降序重排+重编号；剔除表删除医院行+重编号
- 不新建独立sheet
"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '..', 'output')
CITIES = ['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
MAIN, DROP = '商业高层(纯净)', '已剔除-公共设施'
LOG = []
def log(s): LOG.append(s); print(s)

def band(h):
    if h is None: return ''
    if h >= 80: return '≥80米(约20层)'
    if h >= 64: return '≥64米(约16层)'
    if h >= 48: return '≥48米(约12层)'
    if h >= 40: return '≥40米(约10层)'
    if h >= 32: return '≥32米(约8层)'
    return '<32米'

def jakarta_area_map():
    raw = json.load(open(os.path.join(HERE,'..','data','jakarta_official_geom_raw.json'), encoding='utf-8'))
    m = {}
    for rec in raw:
        la, lo, ar = rec.get('_lat'), rec.get('_lon'), rec.get('luas_bgn')
        if la and lo and ar:
            m[(round(la,6), round(lo,6))] = ar
    return m

def idx_of(header, name):
    for i,h in enumerate(header):
        if h and name in str(h): return i
    return None

def run_city(city, write, area_map):
    path = os.path.join(OUT, city+'高层建筑清单.xlsx')
    wbr = load_workbook(path, read_only=True, data_only=True)
    mrows = list(wbr[MAIN].iter_rows(values_only=True))
    drows = list(wbr[DROP].iter_rows(values_only=True))
    wbr.close()
    mh = [str(c).strip() if c is not None else '' for c in mrows[0]]
    dh = [str(c).strip() if c is not None else '' for c in drows[0]]
    # 剔除表列索引
    d = {name: idx_of(dh, name) for name in ['楼宇名称','名称来源','建筑全名(官方)','纬度','经度','高度(米)','层数','行政区','街道办区','地址']}
    dcat = idx_of(dh, '剔除类别')
    # 主表列索引
    m = {name: idx_of(mh, name) for name in mh}
    # 识别医院行
    hosp = []
    keep_drop = [drows[0]]  # 保留表头
    for r in drows[1:]:
        if all(v is None for v in r): continue
        cat = str(r[dcat]) if dcat is not None and dcat < len(r) and r[dcat] is not None else ''
        if '医院' in cat:
            hosp.append(r)
        else:
            keep_drop.append(r)
    # 医院行 → 主表格式
    def get(r, name):
        i = d.get(name)
        return r[i] if i is not None and i < len(r) else None
    new_main_rows = []
    for r in hosp:
        row = [None]*len(mh)
        row[m['楼宇名称']]      = get(r,'楼宇名称')
        row[m['名称来源']]      = get(r,'名称来源')
        if m.get('OSM匹配距离(米)') is not None: row[m['OSM匹配距离(米)']] = None
        row[m['建筑全名(官方)']] = get(r,'建筑全名(官方)')
        row[m['纬度']]          = get(r,'纬度')
        row[m['经度']]          = get(r,'经度')
        h = get(r,'高度(米)')
        row[m['高度(米)']]      = h
        row[m['层数']]          = get(r,'层数')
        row[m['高度分档']]      = band(h)
        row[m['用途分类']]      = '医院'
        row[m['行政区']]        = get(r,'行政区')
        row[m['街道办区']]      = get(r,'街道办区')
        row[m['地址']]          = get(r,'地址')
        # 面积回填(仅雅加达)
        area = None
        if area_map is not None:
            la, lo = get(r,'纬度'), get(r,'经度')
            if la is not None and lo is not None:
                area = area_map.get((round(la,6), round(lo,6)))
        row[m['建筑面积(㎡)']] = area
        new_main_rows.append(row)
    # 合并 + 按高度降序（稳定）
    body = [list(r) for r in mrows[1:] if not all(v is None for v in r)]
    hi = m['高度(米)']
    merged = body + new_main_rows
    merged.sort(key=lambda x: (x[hi] if x[hi] is not None else -1), reverse=True)
    # 重编号
    for i,row in enumerate(merged, 1): row[0] = i
    filled_area = sum(1 for r in new_main_rows if r[m['建筑面积(㎡)']] is not None)
    log('%-5s 医院=%-3d 主表 %d→%d  剔除 %d→%d  面积回填=%d'
        % (city, len(hosp), len(body), len(merged), len(drows)-1, len(keep_drop)-1, filled_area))
    if not write or not hosp:
        return len(hosp)
    # 写回
    wb = load_workbook(path)
    mws = wb[MAIN]; dws = wb[DROP]
    # 主表：覆盖写(变多，无需删行)
    for i,row in enumerate(merged, start=2):
        for c,v in enumerate(row, start=1): mws.cell(row=i, column=c, value=v)
    # 剔除表：覆盖写 + 删尾部多余
    for j,row in enumerate(keep_drop[1:], start=2):
        vals = list(row)
        vals[0] = j-1
        for c,v in enumerate(vals, start=1): dws.cell(row=j, column=c, value=v)
    last = len(keep_drop)  # 最后有效行(含表头)
    if dws.max_row > last:
        dws.delete_rows(last+1, dws.max_row-last)
    wb.save(path); wb.close()
    return len(hosp)

def main():
    write = '--run' in sys.argv
    log('='*60); log('  医院并入主表 - ' + ('执行' if write else 'dry-run预览')); log('='*60)
    amap = jakarta_area_map()
    tot = 0
    for c in CITIES:
        tot += run_city(c, write, amap if c=='雅加达' else None)
    log('-'*60); log('  合计并入医院: %d' % tot)
    open(os.path.join(OUT,'_医院并入主表日志.txt'),'w',encoding='utf-8').write('\n'.join(LOG))

main()
