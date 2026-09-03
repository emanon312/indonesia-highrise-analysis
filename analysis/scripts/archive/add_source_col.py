# -*- coding: utf-8 -*-
"""在12城 Excel 主表+剔除表「名称来源」列后插入「数据来源」列。dry-run默认，--run执行。
- 雅加达=官方；其余11城=Google估算（整表常量）
- 表头蓝色样式沿用；分档等其他列按列名被下游脚本引用，插列不影响
"""
import sys, os, copy
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '..', 'output')
CITIES = ['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']
SHEETS = ['商业高层(纯净)', '已剔除-公共设施']
NEWCOL = '数据来源'
LOG=[]
def log(s): LOG.append(s); print(s)

def process(ws, src):
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    body = [list(r) for r in rows[1:] if not all(v is None for v in r)]
    pos = next(i for i,h in enumerate(header) if h and '名称来源' in str(h))
    at = pos+1
    old_w = [ws.column_dimensions[get_column_letter(i+1)].width for i in range(len(header))]
    tmpl = ws.cell(row=1, column=1)
    tf,tl,ta,tb = copy.copy(tmpl.font),copy.copy(tmpl.fill),copy.copy(tmpl.alignment),copy.copy(tmpl.border)
    new_header = header[:at] + [NEWCOL] + header[at:]
    new_body = [r[:at] + [src] + r[at:] for r in body]
    # 写表头
    for c,v in enumerate(new_header, 1): ws.cell(row=1, column=c, value=v)
    lc = ws.cell(row=1, column=len(new_header)); lc.font,lc.fill,lc.alignment,lc.border = tf,tl,ta,tb
    # 写数据
    for i,row in enumerate(new_body, 2):
        for c,v in enumerate(row, 1): ws.cell(row=i, column=c, value=v)
    # 删尾部多余行(若过滤了空行)
    last = len(new_body)+1
    if ws.max_row > last: ws.delete_rows(last+1, ws.max_row-last)
    # 列宽
    new_w = old_w[:at] + [12] + old_w[at:]
    for i,w in enumerate(new_w, 1):
        if w: ws.column_dimensions[get_column_letter(i)].width = w
    return len(new_header), len(new_body)

def run_city(city, write):
    path = os.path.join(OUT, city+'高层建筑清单.xlsx')
    src = '官方' if city=='雅加达' else 'Google估算'
    if not write:
        wb = load_workbook(path, read_only=True)
        info=[]
        for sn in SHEETS:
            h=[str(x) for x in next(wb[sn].iter_rows(min_row=1,max_row=1,values_only=True))]
            info.append('%s:%d列→%d列'%(sn,len(h),len(h)+1))
        wb.close()
        log('%-5s src=%-10s %s'%(city,src,' | '.join(info)))
        return
    wb = load_workbook(path)
    r=[]
    for sn in SHEETS:
        nc,nb = process(wb[sn], src)
        r.append('%s(%d列,%d行)'%(sn,nc,nb))
    wb.save(path); wb.close()
    log('%-5s src=%-10s %s'%(city,src,' | '.join(r)))

def main():
    write='--run' in sys.argv
    log('='*60); log('  新增数据来源列 - '+('执行' if write else 'dry-run')); log('='*60)
    for c in CITIES: run_city(c, write)
    open(os.path.join(OUT,'_数据来源列日志.txt'),'w',encoding='utf-8').write('\n'.join(LOG))

main()
