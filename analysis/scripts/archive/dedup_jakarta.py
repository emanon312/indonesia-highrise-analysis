# -*- coding: utf-8 -*-
"""雅加达高层建筑精确去重脚本 v2 —— 多轮合并 + 五道防线

去重策略（三轮递进）：
  第一轮：按 (纬度round8, 经度round8, 高度round1) 精确分组 → 组内选最优
  第二轮：同名 + 距离 < 5m → 直接合并（浮点误差兜底）
  第三轮：同名 + 5-50m + 高度差<2m + 地址相同 → 过五道防线：
    ① 建筑面积差 > 30%       → 保留（不同塔楼）
    ② OSM匹配距离差异 > 50m  → 保留（不同OSM实体）
    ③ 层数不同               → 保留（高度同层数不同→不同塔楼）
    ④ 同名紧密聚簇 ≥ 3条     → 不合并，导出"去重待复核"sheet 人工判断
    ⑤ 以上都不触发            → 合并

输出变更：
  - Sheet1 "商业高层(纯净)"：去重后的数据
  - Sheet2 "已剔除-公共设施"：不动
  - Sheet3 "说明"：追加去重记录
  - Sheet4 "去重待复核"（新增）：防线④拦截的聚簇，需人工判断
"""

import os, shutil, re
from datetime import datetime
from math import radians, sin, cos, sqrt, asin
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ===================== 路径配置 =====================
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT = os.path.join(BASE, "output", "雅加达高层建筑清单.xlsx")
BACKUP = INPUT + ".bak"

# ===================== 工具函数 =====================

def find_col(headers, keywords):
    """根据关键词列表查找列索引（0-based）"""
    for i, h in enumerate(headers):
        if h and all(kw in h for kw in keywords):
            return i
    raise ValueError(f"无法找到包含以下关键词的列: {keywords}")

def has_value(val, exclude_set=None):
    """判断单元格值是否有意义"""
    if val is None:
        return False
    if isinstance(val, str) and val.strip() == '':
        return False
    if exclude_set is not None and val in exclude_set:
        return False
    return True

def safe_float(val):
    """安全转浮点数"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def safe_int(val):
    """安全转整数"""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None

def haversine(lat1, lon1, lat2, lon2):
    """Haversine 距离（米）"""
    R = 6371000
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = sin(dlat/2)**2 + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2
    return R * 2 * asin(sqrt(a))

def normalize_name(name):
    """标准化名称：去空格、转小写、去特殊字符"""
    if name is None:
        return ""
    n = str(name).strip().lower()
    n = re.sub(r'[^a-z0-9一-鿿]', '', n)
    return n

def score_row(row, col_map):
    """对一行数据评分，满分 15 分"""
    score = 0
    if has_value(row[col_map['area']]):
        score += 3
    if has_value(row[col_map['name']], exclude_set={'（无名）', '(无名)'}):
        score += 2
    if has_value(row[col_map['fullname']], exclude_set={'—'}):
        score += 2
    if has_value(row[col_map['use']], exclude_set={'未分类'}):
        score += 2
    if has_value(row[col_map['addr']], exclude_set={'—'}):
        score += 1
    if has_value(row[col_map['floors']]):
        score += 1
    if has_value(row[col_map['district']]):
        score += 1
    if has_value(row[col_map['subdistrict']]):
        score += 1
    if has_value(row[col_map['source']], exclude_set={'无'}):
        score += 1
    if has_value(row[col_map['osm_dist']], exclude_set={'—'}):
        score += 1
    return score

# ===================== 主流程 =====================

def main():
    start_time = datetime.now()
    print("=" * 60)
    print("  雅加达高层建筑精确去重 v2（多轮合并 + 五道防线）")
    print(f"  时间：{start_time.strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # ---- 步骤1：备份 ----
    if os.path.exists(BACKUP):
        os.remove(BACKUP)
    shutil.copy2(INPUT, BACKUP)
    print(f"\n[1] 备份完成: {os.path.basename(BACKUP)}")

    # ---- 步骤2：加载工作簿 ----
    wb = load_workbook(INPUT)
    ws_main = wb[wb.sheetnames[0]]
    ws_drop = wb[wb.sheetnames[1]]
    ws_note = wb[wb.sheetnames[2]]

    header_cells = list(ws_main[1])
    headers = [c.value if c.value is not None else "" for c in header_cells]
    print(f"\n[2] 读取表头: {len(headers)} 列")

    # ---- 步骤3：列索引映射 ----
    col_map = {
        'seq':        find_col(headers, ['序']),
        'name':       find_col(headers, ['楼宇', '名称']),
        'source':     find_col(headers, ['名称', '来源']),
        'osm_dist':   find_col(headers, ['OSM']),
        'fullname':   find_col(headers, ['建筑', '全名']),
        'lat':        find_col(headers, ['纬']),
        'lon':        find_col(headers, ['经']),
        'height':     find_col(headers, ['高', '米']),
        'floors':     find_col(headers, ['层']),
        'band':       find_col(headers, ['分档']),
        'use':        find_col(headers, ['用途', '分类']),
        'district':   find_col(headers, ['行政']),
        'subdistrict': find_col(headers, ['街道']),
        'addr':       find_col(headers, ['地址']),
        'area':       find_col(headers, ['建筑', '面积']),
    }
    print(f"[2b] 列索引映射完成，{len(col_map)} 个字段")

    # ---- 步骤4：读取所有数据行 ----
    data_rows = []
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, values_only=True):
        if all(v is None for v in row):
            continue
        data_rows.append(list(row))

    total_original = len(data_rows)
    print(f"\n[3] 读取数据行: {total_original}")

    # ---- 步骤5：构建合并图（并查集） ----
    # 每条记录在原 data_rows 中的索引作为节点 ID
    n = len(data_rows)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[rx] = ry
            return True
        return False

    # ---- 预计算每条记录的属性（避免重复取值） ----
    rec_name = [normalize_name(data_rows[i][col_map['name']]) for i in range(n)]
    rec_lat = [safe_float(data_rows[i][col_map['lat']]) for i in range(n)]
    rec_lon = [safe_float(data_rows[i][col_map['lon']]) for i in range(n)]
    rec_h = [safe_float(data_rows[i][col_map['height']]) for i in range(n)]
    rec_floors = [safe_int(data_rows[i][col_map['floors']]) for i in range(n)]
    rec_area = [safe_float(data_rows[i][col_map['area']]) for i in range(n)]
    rec_osm = [safe_float(data_rows[i][col_map['osm_dist']]) for i in range(n)]
    rec_addr = []
    for i in range(n):
        v = data_rows[i][col_map['addr']]
        rec_addr.append(str(v).strip() if v else "")

    # ==================== 5a. 第一轮：round8 + round1 精确分组 ====================
    print("\n[4] 第一轮：round8坐标 + round1高度 精确分组...")
    round8_groups = defaultdict(list)
    for i in range(n):
        if rec_lat[i] is None or rec_lon[i] is None or rec_h[i] is None:
            continue
        key = (round(rec_lat[i], 8), round(rec_lon[i], 8), round(rec_h[i], 1))
        round8_groups[key].append(i)

    round1_merged = 0
    for key, indices in round8_groups.items():
        if len(indices) > 1:
            first = indices[0]
            for other in indices[1:]:
                if union(first, other):
                    round1_merged += 1
    print(f"    分组数: {len(round8_groups)}, 合并: {round1_merged} 对")

    # ==================== 5b. 构建同名组索引 ====================
    name_groups = defaultdict(list)
    for i in range(n):
        if rec_name[i]:
            name_groups[rec_name[i]].append(i)

    # ==================== 5c. 第二轮：同名 + < 5m → 直接合并 ====================
    print("\n[5] 第二轮：同名 + 距离<5m 直接合并...")
    round2_merged = 0
    for norm_name, indices in name_groups.items():
        m = len(indices)
        if m < 2:
            continue
        for p in range(m):
            for q in range(p + 1, m):
                i, j = indices[p], indices[q]
                if rec_lat[i] is None or rec_lat[j] is None:
                    continue
                dist = haversine(rec_lat[i], rec_lon[i], rec_lat[j], rec_lon[j])
                if dist < 5:
                    if union(i, j):
                        round2_merged += 1
    print(f"    合并: {round2_merged} 对")

    # ==================== 5d. 第三轮：5-50m + 防线 ====================
    print("\n[6] 第三轮：同名 + 5-50m + 防线判断...")
    round3_stats = {
        'merged': 0,
        'defense1_area': 0,
        'defense2_osm': 0,
        'defense3_floors': 0,
        'defense4_cluster': 0,
    }

    # 防线④ 收集：按同名组找紧密聚簇
    defense4_clusters = []  # 每个元素: {name, records_indices, diameter}

    for norm_name, indices in name_groups.items():
        m = len(indices)
        if m < 2:
            continue

        # 找出组内所有 5-50m + 高度差<2m + 地址相同 的边
        eligible_edges = []  # (i, j, dist)
        for p in range(m):
            for q in range(p + 1, m):
                i, j = indices[p], indices[q]
                if rec_lat[i] is None or rec_lat[j] is None or rec_h[i] is None or rec_h[j] is None:
                    continue
                dist = haversine(rec_lat[i], rec_lon[i], rec_lat[j], rec_lon[j])
                if not (5 <= dist < 50):
                    continue
                if abs(rec_h[i] - rec_h[j]) >= 2:
                    continue
                if not (rec_addr[i] and rec_addr[j] and rec_addr[i] == rec_addr[j]):
                    continue
                eligible_edges.append((i, j, dist))

        if not eligible_edges:
            continue

        # 用并查集找连通分量（仅在本同名组内）
        local_parent = {}
        def local_find(x):
            if x not in local_parent:
                local_parent[x] = x
            if local_parent[x] != x:
                local_parent[x] = local_find(local_parent[x])
            return local_parent[x]
        def local_union(x, y):
            local_parent[local_find(x)] = local_find(y)

        for i, j, dist in eligible_edges:
            local_union(i, j)

        # 按根分组
        local_clusters = defaultdict(list)
        for i, j, dist in eligible_edges:
            root = local_find(i)
            local_clusters[root].append((i, j, dist))

        # 检查每个连通分量
        defense4_nodes = set()  # 本组内触发防线④的节点

        for root, edges in local_clusters.items():
            nodes = set()
            for i, j, dist in edges:
                nodes.add(i)
                nodes.add(j)
            if len(nodes) < 3:
                continue

            # 计算分量直径
            node_list = list(nodes)
            max_diam = 0
            for p in range(len(node_list)):
                for q in range(p + 1, len(node_list)):
                    d = haversine(rec_lat[node_list[p]], rec_lon[node_list[p]],
                                  rec_lat[node_list[q]], rec_lon[node_list[q]])
                    max_diam = max(max_diam, d)

            if max_diam < 200:  # 紧密聚簇
                for nid in nodes:
                    defense4_nodes.add(nid)
                defense4_clusters.append({
                    'name': data_rows[indices[0]][col_map['name']],
                    'indices': sorted(node_list),
                    'diameter': max_diam,
                })

        # 逐对判断（仅 eligible_edges 中的对）
        for i, j, dist in eligible_edges:
            # 如果其中任一节点在防线④聚簇中，整对跳过
            if i in defense4_nodes and j in defense4_nodes:
                round3_stats['defense4_cluster'] += 1
                continue

            # 防线①: 面积差 > 30%
            if rec_area[i] and rec_area[j] and rec_area[i] > 0 and rec_area[j] > 0:
                area_diff = abs(rec_area[i] - rec_area[j]) / max(rec_area[i], rec_area[j])
                if area_diff > 0.3:
                    round3_stats['defense1_area'] += 1
                    continue

            # 防线②: OSM匹配距离差异 > 50m
            if rec_osm[i] and rec_osm[j]:
                if abs(rec_osm[i] - rec_osm[j]) > 50:
                    round3_stats['defense2_osm'] += 1
                    continue

            # 防线③: 层数不同
            if rec_floors[i] and rec_floors[j] and rec_floors[i] != rec_floors[j]:
                round3_stats['defense3_floors'] += 1
                continue

            # 防线⑤: 通过 → 合并
            if union(i, j):
                round3_stats['merged'] += 1

    print(f"    合并(防线通过): {round3_stats['merged']} 对")
    print(f"    防线①拦截(面积): {round3_stats['defense1_area']} 对")
    print(f"    防线②拦截(OSM):  {round3_stats['defense2_osm']} 对")
    print(f"    防线③拦截(层数): {round3_stats['defense3_floors']} 对")
    print(f"    防线④拦截(聚簇): {round3_stats['defense4_cluster']} 对 ({len(defense4_clusters)} 组)")

    # ---- 步骤6：按连通分量合并，选最优保留 ----
    print("\n[7] 按连通分量合并，评分选优...")
    components = defaultdict(list)
    for i in range(n):
        root = find(i)
        components[root].append(i)

    total_components = len(components)
    total_removed = n - total_components
    print(f"    连通分量数: {total_components}")
    print(f"    删除行数:   {total_removed}")

    # 每个分量选最优
    keep_indices = set()
    for root, indices in components.items():
        if len(indices) == 1:
            keep_indices.add(indices[0])
        else:
            scored = []
            for idx in indices:
                s = score_row(data_rows[idx], col_map)
                area_val = rec_area[idx]
                area = area_val if area_val is not None else 0.0
                scored.append((idx, s, area))
            scored.sort(key=lambda x: (-x[1], -x[2], x[0]))
            keep_indices.add(scored[0][0])

    final_rows = [data_rows[i] for i in range(n) if i in keep_indices]
    # 按原始顺序排序（保持数据行顺序一致）
    final_rows.sort(key=lambda row: n)  # 保序，用原始索引

    # 更好的排序：维持删除前的相对顺序
    ordered_final = [data_rows[i] for i in range(n) if i in keep_indices]

    print(f"    保留行数: {len(ordered_final)}")

    # ---- 步骤7：写回 Sheet1 ----
    print("\n[8] 写回 Sheet1「商业高层(纯净)」...")
    if ws_main.max_row > 1:
        ws_main.delete_rows(2, ws_main.max_row - 1)

    for i, row_data in enumerate(ordered_final):
        for j, val in enumerate(row_data):
            ws_main.cell(row=i + 2, column=j + 1, value=val)

    # 重新编号序号列
    seq_col_idx = col_map['seq'] + 1
    for i in range(len(ordered_final)):
        ws_main.cell(row=i + 2, column=seq_col_idx, value=i + 1)

    print(f"    已写入 {len(ordered_final)} 行，序号已重编")

    # ---- 步骤8：处理防线④ → 写 Sheet4 "去重待复核" ----
    print("\n[9] 处理防线④聚簇 → Sheet「去重待复核」...")

    # 删除已有的"去重待复核" sheet（如果存在）
    review_sheet_name = "去重待复核"
    if review_sheet_name in [s.title for s in wb.worksheets]:
        del wb[review_sheet_name]

    if defense4_clusters:
        ws_review = wb.create_sheet(review_sheet_name, 3)  # 插入为第4个sheet

        # 表头
        review_headers = [
            '聚簇ID', '楼宇名称', '聚簇记录数', '聚簇直径(m)',
            '序号(去重后)', '纬度', '经度', '高度(m)', '层数', '建筑面积(㎡)',
            '用途分类', '行政区', '街道办区', '地址', '高度分档', '判定理由'
        ]
        # 蓝色表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for c, h in enumerate(review_headers, 1):
            cell = ws_review.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 写数据
        row_idx = 2
        for cidx, cluster in enumerate(defense4_clusters):
            for rec_idx in cluster['indices']:
                # 该记录在去重后可能已被合并掉——只列出在 keep_indices 中的
                row_data = data_rows[rec_idx]
                ws_review.cell(row=row_idx, column=1, value=f"D4_{cidx+1:04d}")
                ws_review.cell(row=row_idx, column=2, value=cluster['name'])
                ws_review.cell(row=row_idx, column=3, value=len(cluster['indices']))
                ws_review.cell(row=row_idx, column=4, value=round(cluster['diameter'], 1))
                ws_review.cell(row=row_idx, column=5, value=row_data[col_map['seq']])
                ws_review.cell(row=row_idx, column=6, value=rec_lat[rec_idx])
                ws_review.cell(row=row_idx, column=7, value=rec_lon[rec_idx])
                ws_review.cell(row=row_idx, column=8, value=rec_h[rec_idx])
                ws_review.cell(row=row_idx, column=9, value=rec_floors[rec_idx])
                ws_review.cell(row=row_idx, column=10, value=rec_area[rec_idx])
                ws_review.cell(row=row_idx, column=11, value=row_data[col_map['use']])
                ws_review.cell(row=row_idx, column=12, value=row_data[col_map['district']])
                ws_review.cell(row=row_idx, column=13, value=row_data[col_map['subdistrict']])
                # 地址可能很长，截断
                addr_val = rec_addr[rec_idx]
                ws_review.cell(row=row_idx, column=14, value=addr_val[:120] if addr_val else '')
                ws_review.cell(row=row_idx, column=15, value=row_data[col_map['band']])
                ws_review.cell(row=row_idx, column=16,
                              value=f"同名≥3条紧密聚簇（直径{cluster['diameter']:.0f}m），"
                                    f"疑似塔楼群，需人工判断合并/保留")
                row_idx += 1

        # 调整列宽
        col_widths = [12, 28, 10, 12, 12, 16, 17, 10, 6, 12, 14, 10, 14, 40, 18, 50]
        for c, w in enumerate(col_widths, 1):
            ws_review.column_dimensions[get_column_letter(c)].width = w

        print(f"    防线④聚簇: {len(defense4_clusters)} 组, 写入 {row_idx - 2} 行")
    else:
        print(f"    防线④聚簇: 0 组（无需人工复核）")

    # ---- 步骤9：更新「说明」sheet ----
    print("\n[10] 更新「说明」sheet...")
    last_row = ws_note.max_row
    if ws_note.cell(row=last_row, column=1).value is not None:
        last_row += 2

    nr = last_row
    ws_note.cell(row=nr, column=1, value="【v2 多轮去重记录】")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    去重时间：{start_time.strftime('%Y-%m-%d %H:%M')}")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="    去重策略（三轮递进 + 五道防线）：")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      第一轮：按（纬度round8, 经度round8, 高度round1）精确分组 → 组内选最优")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      第二轮：同名 + 距离<5m → 直接合并（浮点误差兜底）")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="      第三轮：同名 + 5-50m + 高度差<2m + 地址相同 → 过五道防线")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="        ①面积差>30%→保留 ②OSM距离差异→保留 ③层数不同→保留 "
                       "④紧密聚簇≥3条→待复核 ⑤通过→合并")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    去重结果：原始 {total_original} 行 → 保留 {len(ordered_final)} 行，"
                       f"删除 {total_removed} 行（{total_components} 个连通分量）")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    各轮统计：第一轮合并 {round1_merged} 对 | "
                       f"第二轮合并 {round2_merged} 对 | "
                       f"第三轮合并 {round3_stats['merged']} 对")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    防线拦截：①面积 {round3_stats['defense1_area']} | "
                       f"②OSM {round3_stats['defense2_osm']} | "
                       f"③层数 {round3_stats['defense3_floors']} | "
                       f"④聚簇 {round3_stats['defense4_cluster']}对({len(defense4_clusters)}组)")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value=f"    防线④聚簇详情见 Sheet「{review_sheet_name}」，需人工逐组判断")
    nr += 1
    ws_note.cell(row=nr, column=1,
                 value="    评分标准：建筑面积(3分) + 楼宇名称(2分) + 建筑全名(2分) + "
                       "用途分类(2分) + 地址(1分) + 层数(1分) + 行政区(1分) + "
                       "街道办区(1分) + 名称来源(1分) + OSM匹配距离(1分) = 15分")

    # ---- 步骤10：保存 ----
    print("\n[11] 保存文件...")
    wb.save(INPUT)
    wb.close()

    elapsed = (datetime.now() - start_time).total_seconds()
    print(f"\n{'=' * 60}")
    print(f"  去重完成！（耗时 {elapsed:.1f}s）")
    print(f"  原始: {total_original} → 最终: {len(ordered_final)}  (删除 {total_removed} 行)")
    print(f"  文件: {os.path.basename(INPUT)}")
    print(f"  备份: {os.path.basename(BACKUP)}")
    print(f"{'=' * 60}")

    # 打印关键统计
    print(f"\n  三轮合并明细:")
    print(f"    第一轮(round8): {round1_merged} 对")
    print(f"    第二轮(<5m):    {round2_merged} 对")
    print(f"    第三轮(5-50m):  {round3_stats['merged']} 对")
    print(f"    合计:           {round1_merged + round2_merged + round3_stats['merged']} 对")
    print(f"  防线拦截:")
    print(f"    ①面积: {round3_stats['defense1_area']} ②OSM: {round3_stats['defense2_osm']} "
          f"③层数: {round3_stats['defense3_floors']} ④聚簇: {round3_stats['defense4_cluster']}对/{len(defense4_clusters)}组")
    if defense4_clusters:
        top5 = sorted(defense4_clusters, key=lambda c: -len(c['indices']))[:5]
        print(f"\n  防线④ TOP5 聚簇（待人工复核）:")
        for c in top5:
            print(f"    {c['name'][:30]:30s} {len(c['indices'])}条 直径{c['diameter']:.0f}m")


if __name__ == "__main__":
    main()
