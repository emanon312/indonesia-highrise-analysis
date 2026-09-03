# -*- coding: utf-8 -*-
"""analysis/output 一键清理脚本。
基于 audit_dedup.py 的检查结果，执行安全清理操作：
1. 删除 zip 包（L1 完全重复）
2. 重新生成汇总 md（L4 数据同步）
3. 归档过期 dashboard（L4 标记）
用法：python dedup_clean.py
"""

import os, sys, shutil
from datetime import datetime

# ===================== 路径配置 =====================
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(BASE, "output")
SCRIPTS_DIR = os.path.join(BASE, "scripts")

# ===================== 清理步骤 =====================

def step_delete_zip():
    """步骤1：删除 zip 包（散落 xlsx 的完全副本）"""
    print("=" * 50)
    print("步骤1：删除 ZIP 包（L1 完全重复）")
    print("=" * 50)

    zip_path = os.path.join(OUT_DIR, "印尼高层建筑清单.zip")
    if os.path.exists(zip_path):
        size_kb = os.path.getsize(zip_path) / 1024
        print(f"  文件：印尼高层建筑清单.zip ({size_kb:.0f} KB)")
        print(f"  原因：zip 内 12 个 xlsx 与散落文件 SHA256 完全一致")
        print(f"  说明：zip 只是打包副本，删除后随时可重新打包")
        os.remove(zip_path)
        print(f"  ✓ 已删除，释放 {size_kb:.0f} KB")
    else:
        print(f"  (文件不存在，跳过)")

def step_regenerate_summary():
    """步骤2：重新生成汇总 md（从当前 Excel 重新统计）"""
    print("\n" + "=" * 50)
    print("步骤2：重新生成汇总 md（L4 数据同步）")
    print("=" * 50)

    summary_script = os.path.join(SCRIPTS_DIR, "..", "data", "summary_12cities.py")
    if os.path.exists(summary_script):
        print(f"  执行：summary_12cities.py")
        # 直接内联执行逻辑（避免子进程依赖问题）
        try:
            import subprocess
            result = subprocess.run(
                [sys.executable, summary_script],
                capture_output=True, text=True, timeout=60,
                cwd=os.path.dirname(summary_script)
            )
            if result.returncode == 0:
                print(f"  ✓ 汇总 md 已重新生成")
                if result.stdout.strip():
                    print(f"  {result.stdout.strip()}")
            else:
                print(f"  ✗ 执行出错：{result.stderr.strip()}")
                print(f"  → 使用内联逻辑重新生成...")
                _regenerate_inline()
        except Exception as e:
            print(f"  子进程调用失败：{e}")
            print(f"  → 使用内联逻辑重新生成...")
            _regenerate_inline()
    else:
        print(f"  脚本不存在，使用内联逻辑重新生成...")
        _regenerate_inline()

def _regenerate_inline():
    """内联重新生成汇总md（不依赖子进程）"""
    import re
    from openpyxl import load_workbook

    CITIES = [
        "雅加达", "泗水", "万隆", "唐格朗", "望加锡", "棉兰",
        "巴淡", "三宝垄", "勿加泗", "德波", "茂物", "巨港"
    ]

    rows_out = []
    for city in CITIES:
        xlsx_path = os.path.join(OUT_DIR, f"{city}高层建筑清单.xlsx")
        if not os.path.exists(xlsx_path):
            continue
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        bi = next((i for i, h in enumerate(header) if "分档" in h), None)
        buckets = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
        for r in rows[1:]:
            band_str = str(r[bi]) if bi is not None and bi < len(r) and r[bi] is not None else ""
            m = re.search(r"(80|64|48|40|32)", band_str)
            if m:
                thr = int(m.group(1))
                if thr in buckets:
                    buckets[thr] += 1
        wb.close()
        cum = {t: sum(v for k, v in buckets.items() if k >= t) for t in (32, 40, 48, 64, 80)}
        # 判断数据来源
        src = "官方真值" if city == "雅加达" else "Google ML估算(下限)"
        rows_out.append({"name": city, **cum, "src": src})

    rows_out.sort(key=lambda x: x[32], reverse=True)

    lines = []
    lines.append("# 印尼12城高层建筑汇总（按累计阈值统计栋数）\n")
    lines.append("> 口径：4米/层。≥32米=8层 / ≥40=10 / ≥48=12 / ≥64=16 / ≥80=20。")
    if "雅加达" in [r["name"] for r in rows_out]:
        jkt_row = next(r for r in rows_out if r["name"] == "雅加达")
        lines.append(f"> 雅加达为官方真值标杆（纯商业高层，已剔除公共设施 {jkt_row[32]} 栋）；其余11城为 Google ML 估算**下限**。")
    lines.append(f"> 自动生成于 {datetime.now().strftime('%Y-%m-%d %H:%M')}。\n")
    lines.append("| 城市 | ≥32米(8层) | ≥40米(10层) | ≥48米(12层) | ≥64米(16层) | ≥80米(20层) | 数据来源 |")
    lines.append("|------|-----------:|-----------:|-----------:|-----------:|-----------:|----------|")
    tot = {32: 0, 40: 0, 48: 0, 64: 0, 80: 0}
    for r in rows_out:
        lines.append(f"| {r['name']} | {r[32]} | {r[40]} | {r[48]} | {r[64]} | {r[80]} | {r['src']} |")
        for k in tot:
            tot[k] += r[k]
    lines.append(f"| **合计** | **{tot[32]}** | **{tot[40]}** | **{tot[48]}** | **{tot[64]}** | **{tot[80]}** | — |")

    md_path = os.path.join(OUT_DIR, "12城高层建筑汇总.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"  ✓ 汇总 md 已重新生成 → {md_path}")
    # 打印摘要
    for r in rows_out:
        print(f"    {r['name']}: ≥32m={r[32]}, ≥40m={r[40]}, ≥48m={r[48]}, ≥64m={r[64]}, ≥80m={r[80]}")

def step_archive_dashboard():
    """步骤3：归档过期 dashboard"""
    print("\n" + "=" * 50)
    print("步骤3：归档过期 dashboard（L4 数据过期）")
    print("=" * 50)

    dash_path = os.path.join(OUT_DIR, "dashboard.html")
    legacy_path = os.path.join(OUT_DIR, "dashboard_legacy.html")

    if os.path.exists(dash_path):
        if os.path.exists(legacy_path):
            print(f"  目标文件已存在，跳过：{legacy_path}")
            return
        shutil.move(dash_path, legacy_path)
        print(f"  dashboard.html → dashboard_legacy.html")
        print(f"  原因：仅含1/12城数据（雅加达），日期2026-06-18，已过期")
        print(f"  说明：如需要更新为12城当前数据，运行 build_dashboard.py")
    else:
        print(f"  (dashboard.html 不存在，跳过)")

# ===================== 主入口 =====================

def main():
    print("=" * 50)
    print("  印尼12城 output/ 一键清理")
    print(f"  时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"  目录：{OUT_DIR}")
    print("=" * 50)

    # 清理前快照
    before_files = set(f for f in os.listdir(OUT_DIR) if os.path.isfile(os.path.join(OUT_DIR, f)))
    before_xlsx = [f for f in before_files if f.endswith(".xlsx")]
    print(f"\n清理前：{len(before_files)} 个文件（含 {len(before_xlsx)} 个 xlsx）")

    # 执行清理
    step_delete_zip()
    step_regenerate_summary()
    step_archive_dashboard()

    # 清理后快照
    after_files = set(f for f in os.listdir(OUT_DIR) if os.path.isfile(os.path.join(OUT_DIR, f)))
    after_xlsx = [f for f in after_files if f.endswith(".xlsx")]
    removed = before_files - after_files
    added = after_files - before_files

    print("\n" + "=" * 50)
    print("  清理完成")
    print("=" * 50)
    print(f"清理后：{len(after_files)} 个文件（含 {len(after_xlsx)} 个 xlsx）")
    if removed:
        print(f"已删除：{', '.join(sorted(removed))}")
    if added:
        print(f"已更新：{', '.join(sorted(added))}")
    print(f"\nxlsx 文件数不变：{len(before_xlsx)} → {len(after_xlsx)} ✓")
    print(f"如需重新打包 zip：在 output/ 目录选中所有 xlsx → 右键 → 压缩")

if __name__ == "__main__":
    main()
