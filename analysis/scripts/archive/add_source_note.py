# -*- coding: utf-8 -*-
"""在12城说明sheet末尾追加「名称来源 vs 数据来源」区别说明。dry-run默认，--run执行。"""
import sys, os, copy
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

HERE=os.path.dirname(os.path.abspath(__file__))
OUT=os.path.join(HERE,'..','output')
CITIES=['雅加达','泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']

JKT=[
 ('t','【名称来源 与 数据来源 —— 两列区别，勿混淆】'),
 ('b','  · 名称来源：这栋楼的“名字”是从哪匹配到的 —— 官方 / OSM补充(≤50米) / OSM补充(存疑,50~100米) / 无。'),
 ('b','  · 数据来源：这栋楼的“本体数据”（经纬度、高度、层数、建筑面积、轮廓）来自哪个源 —— 本城全部为『官方』(DPMPTSP 官方三维库)。'),
 ('b','  · 两列关系：雅加达数据来源恒为『官方』；名称来源多为官方，仅原本无名的楼才用 OSM 补名，故两列偶有不同。'),
]
GEN=[
 ('t','【名称来源 与 数据来源 —— 两列区别，勿混淆】'),
 ('b','  · 名称来源：这栋楼的“名字”是从哪匹配到的 —— CTBUH地标 / OSM补充(≤40米) / 无。'),
 ('b','  · 数据来源：这栋楼的“本体数据”（经纬度、高度、层数、建筑面积、轮廓）来自哪个源 —— 本城全部为『Google估算』(Google V3 坐标与轮廓 + 2.5D 高度)。'),
 ('b','  · 两列可不同：Google 只提供坐标和高度、不提供楼名，楼名需另去 OSM/CTBUH 匹配。例：某楼数据来源=Google估算，但名字来自 OSM → 名称来源=OSM补充。'),
]

def run_city(city, write):
    path=os.path.join(OUT,city+'高层建筑清单.xlsx')
    lines = JKT if city=='雅加达' else GEN
    wb=load_workbook(path)
    ws=wb['说明']
    # 样式模板：找第一个【标题行 和 一个正文行
    t_font=b_font=None
    for r in range(1, ws.max_row+1):
        v=ws.cell(row=r,column=1).value
        if v is None: continue
        s=str(v)
        if t_font is None and s.startswith('【'): t_font=copy.copy(ws.cell(row=r,column=1).font)
        if b_font is None and not s.startswith('【') and r>1: b_font=copy.copy(ws.cell(row=r,column=1).font)
    # 找最后有内容行
    last=1
    for r in range(1, ws.max_row+1):
        if ws.cell(row=r,column=1).value not in (None,''): last=r
    start=last+2
    if not write:
        print('%-5s 说明末行=%d → 从第%d行追加 %d 行'%(city,last,start,len(lines)))
        wb.close(); return
    for i,(kind,text) in enumerate(lines):
        cell=ws.cell(row=start+i, column=1, value=text)
        f = t_font if kind=='t' else b_font
        if f is not None: cell.font=f
    wb.save(path); wb.close()
    print('%-5s 已在第%d行追加 %d 行说明'%(city,start,len(lines)))

def main():
    write='--run' in sys.argv
    print('='*56); print('  追加两列区别说明 - '+('执行' if write else 'dry-run')); print('='*56)
    for c in CITIES: run_city(c, write)

main()
