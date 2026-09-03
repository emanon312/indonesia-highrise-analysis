# -*- coding: utf-8 -*-
"""清洗11城「用途分类」列污染（雅加达官方数据不动）。dry-run默认，--run写回。
污染两类：①城市名/地名占位(值==同行行政区) ②OSM英文标签未翻译(bank/mall等)。
清洗优先级：已规范 > 中文别名归一 > OSM标签映射 > 楼名关键词推断 > 未分类。"""
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

HERE=os.path.dirname(os.path.abspath(__file__)); OUT=os.path.join(HERE,'..','output')
CITIES=['泗水','万隆','唐格朗','望加锡','棉兰','巴淡','三宝垄','勿加泗','德波','茂物','巨港']  # 不含雅加达

STD={'综合体','写字楼','公寓','酒店','商场','商业','商业(店铺)','住宅','医院',
     '仓库','宿舍','工业','政府公共','宗教','会展','未分类'}
# 中文别名归一 → 规范集
CANON={'住宅类':'住宅','商业类':'商业','商业/店铺':'商业(店铺)','店铺':'商业(店铺)',
       '办公住宅':'综合体','旅游休闲':'商业','其他':'未分类','银行':'商业',
       '政府':'政府公共','政府/公共服务':'政府公共','宗教场所':'宗教'}
# OSM英文标签 → 规范集（'__guess__' 表示转楼名推断）
OSM_MAP={'bank':'商业','mall':'商场','supermarket':'商业','marketplace':'商业',
         'townhall':'政府公共','government_office':'政府公共','governor_office':'政府公共',
         'convention':'会展','chapel':'宗教','construction':'__guess__',
         'guardhouse':'未分类','security_post':'未分类','roof':'未分类','ruins':'未分类',
         'office':'写字楼','commercial':'商业','retail':'商业','hotel':'酒店',
         'apartments':'公寓','residential':'住宅','hospital':'医院','warehouse':'仓库',
         'industrial':'工业','dormitory':'宿舍'}
# 已知地名（占位兜底，主判据是 值==同行行政区）
PLACE={'surabaya','bandung','kota bandung','batam','makassar','semarang','kota semarang',
       'kabupaten deli serdang','medan','kota medan','tangerang','tangerang selatan',
       'bogor','kab bogor','depok','palembang','banyuasin','ogan ilir',
       'jakarta selatan','jakarta barat','jakarta timur','jakarta utara','jakarta pusat'}
# 楼名关键词推断（顺序敏感：专指性强的在前，泛词tower/gedung在后）
NAME_GUESS=[
    (['rumah sakit','rsud','rsup','rsia',' rsu',' rs ','klinik','hospital'],'医院'),
    (['hotel','resort',' inn '],'酒店'),
    (['apartemen','apartment','kondominium','condominium','condo','residence','rusun','suites'],'公寓'),
    (['mall',' plaza',' square','itc','mtc',' trade center','shopping','citywalk'],'商场'),
    (['ruko','rukan'],'商业(店铺)'),
    (['asrama','dormitory','kost',' kos '],'宿舍'),
    (['pabrik','gudang','warehouse','factory','industri'],'工业'),
    (['bank'],'商业'),
    # 保守模式：Tower/Office/Gedung/Wisma/Graha 等泛词不推断，归未分类（避免误判）
]

def guess(name):
    low=' '+str(name).lower()+' '
    for kws,cat in NAME_GUESS:
        if any(k in low for k in kws): return cat,'guess'
    return '未分类','unc'

def clean_one(cur, name, admin):
    v=str(cur).strip() if cur is not None else ''
    if v in STD: return v,'keep'
    if v in CANON: return CANON[v],'canon'
    low=v.lower()
    if low in OSM_MAP:
        m=OSM_MAP[low]
        return guess(name) if m=='__guess__' else (m,'osm')
    # 占位：空/None/==行政区/已知地名
    if v in ('','None') or (admin and v==str(admin).strip()) or low in PLACE:
        return guess(name)
    # 其他未知中文/杂值 → 楼名推断兜底
    return guess(name)

def process(city, write, stats):
    path=os.path.join(OUT,f'{city}高层建筑清单.xlsx')
    wb=load_workbook(path)
    ws=wb['商业高层(纯净)']
    h=[str(c.value).strip() if c.value else '' for c in ws[1]]
    ci=h.index('用途分类'); ni=h.index('楼宇名称'); di=h.index('行政区')
    cnt={'keep':0,'canon':0,'osm':0,'guess':0,'unc':0}
    samples=[]
    for row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in row): continue
        cur=row[ci].value; name=row[ni].value; admin=row[di].value
        new,how=clean_one(cur,name,admin)
        cnt[how]+=1
        if how!='keep' and str(cur)!=new:
            if len(samples)<8: samples.append((str(name)[:30],str(cur),new,how))
            if write: row[ci].value=new
    if write: wb.save(path)
    wb.close()
    stats[city]=cnt
    print(f'  {city:<4} 保留{cnt["keep"]:<4} 别名{cnt["canon"]:<3} OSM{cnt["osm"]:<3} '
          f'楼名推断{cnt["guess"]:<4} 未分类{cnt["unc"]:<4}')
    for n,o,nw,how in samples[:4]:
        print(f'        [{how}] {n:<30} {o} → {nw}')

def main():
    write='--run' in sys.argv
    print('='*66); print('  用途分类清洗 - '+('执行写回' if write else 'dry-run预览')+' (雅加达不动)'); print('='*66)
    stats={}
    for c in CITIES: process(c, write, stats)
    tot={k:sum(s[k] for s in stats.values()) for k in ['keep','canon','osm','guess','unc']}
    print('-'*66)
    print(f'  11城合计: 保留{tot["keep"]} 别名归一{tot["canon"]} OSM救回{tot["osm"]} '
          f'楼名推断救回{tot["guess"]} 未分类{tot["unc"]}')
    救回=tot['canon']+tot['osm']+tot['guess']
    print(f'  从污染中救回真实用途: {救回} 栋；最终未分类: {tot["unc"]} 栋')

main()
